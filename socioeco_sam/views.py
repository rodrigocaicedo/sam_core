# -*- coding: utf-8 -*-
from __future__ import unicode_literals

from django.shortcuts import render, redirect

from socioeco_sam.models import Evaluacion_Socioeco, Representante, Estudiante, Domicilio, Integrante_Familia, Situacion_Habitacional, Responsable_Gastos, Propiedades, Ingresos, Gastos, Aceptacion

from socioeco_sam.forms import EvaluacionForm, RepresentanteForm, EstudianteForm, DomicilioForm, IntegranteForm, HabitacionalForm, ResponsableForm, PropiedadesForm, IngresosForm, GastosForm, FinalizarForm

from usuarios_sam.models import CustomUser as User, Students

from general_sam.models import Matricula, Matricula_Grupo

from django.template import Template, Context

from django.core.mail import send_mail

from django.db import IntegrityError

from django.contrib.auth.decorators import login_required, user_passes_test

from django.http import HttpResponse

from admisiones_sam.models import Applications

from general_sam.models import Clase

from usuarios_sam.models import CustomUser as User

from datetime import date

import openpyxl


def is_financiero(user):
    return user.groups.filter(name = "Financiero").exists()

"""
@login_required(login_url = "/user/login")
def Exportar_Requeridos_Completados(request):

  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename={0}.xlsx'.format(date.today())
  wb = openpyxl.Workbook()
  ws = wb.get_active_sheet()
  row = 0
  columns = [(u"CODIGO",250),(u"ESTUDIANTE",250),(u"REPRESENTANTE",250),(u"FAMILIA",250),(u"FORM COMPLETO",250),(u"DOCS RECIBIDOS",250),(u"INFO ADIC.",250),(u"ESTUDIANTES EN MA",250),(u"INGRESOS",250),(u"EGRESOS",250),
  (u"25% DE INGR/EGR",250),(u"PENSION ACTUAL",250)]
  row_num = 0
  for col_num in xrange(len(columns)):
    c = ws.cell(row = row_num+1, column = col_num +1)
    c.value = columns[col_num][0]
  creados = Evaluacion_Socioeco.objects.filter(requerido = True).exclude(enviado = None).order_by("enviado")

  for creado in creados:
    for estudiante in creado.estudiante_set.all():
      row_num += 1
      row = [u"", u"{0} {1} {2}".format(estudiante.apellido_paterno, estudiante.apellido_materno, estudiante.nombres), enviado.solicitante, enviado.familia]
      if creado.enviado == None:
        row.append("SI")
      else:
        row.append("NO")
      row.append("")
      row.append("")
      estudiantes = creado.integrante_familia_set.filter(actividad__iexact = "estudiante", parentesco__iexact = in ["hijo", "hija"]).count()
      montebello = creado.estudiante_set.all().count()
"""
@login_required(login_url = "/user/login")
def Exportar_Admisiones_Xls(request):
  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename=ADMISIONES_{0}.xlsx'.format(date.today())

  wb = openpyxl.Workbook()
  ws = wb.get_active_sheet()
  row = 0
  columns = [(u"SOLICITADO",100),(u"ESTUDIANTE",100),(u"NIVEL",100),(u"REPRESENTANTE",100),(u"CORREO",100),(u"ID_FORMULARIO",100),(u"FORM INICIADO",100),(u"FORM TERMINADO",100)]
  row_num = 0
  for col_num in xrange(len(columns)):
    c = ws.cell(row = row_num+1, column = col_num +1)
    c.value = columns[col_num][0]

  admisiones = Applications.objects.filter(state_applications__state = "Contabilidad").exclude(state_applications__state = "Terminado", state_applications__final_date = None).distinct()

  for admision in admisiones:
    row_num += 1
    estudiante = "{0} {1}".format(admision.surname_student, admision.name_student)
    nivel = admision.applied_grade
    representante = "{0} {1}".format(admision.surname_tutor, admision.name_tutor)
    correo = admision.mail_tutor
    if admision.state_applications_set.filter(state = "Secretaria", initial_date__lt = "2018-06-19").exists():
      solicitado = "NO"
      form = ""
      form_iniciado = ""
      form_enviado = ""
    else:
      solicitado = "SI"
      if Evaluacion_Socioeco.objects.filter(email = admision.mail_tutor).exists():
        evaluacion = Evaluacion_Socioeco.objects.get(email = admision.mail_tutor)
      elif Evaluacion_Socioeco.objects.filter(solicitante = representante).exists():
        evaluacion = Evaluacion_Socioeco.objects.get(solicitante = representante)
      else:
        evaluacion = None
      if evaluacion:
        form = "{0}".format(evaluacion.pk)
        if evaluacion.enviado == None:
          form_enviado = "NO"
        else:
          form_enviado = "SI"
        try:
          form_iniciado = "{0}".format(evaluacion.aceptacion)
          form_iniciado = "SI"
        except:
          form_iniciado = "NO"
      else:
        form = ""
        form_iniciado = ""
        form_enviado = ""
    row = [solicitado, estudiante, nivel, representante, correo, form, form_iniciado, form_enviado]        
    for col_num in xrange(len(row)):
      c = ws.cell(row = row_num +1, column = col_num +1)
      c.value = row[col_num]
  wb.save(response)
  return response





@login_required(login_url = "/user/login")
def Exportar_Enviados_Xls(request):

  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename={0}.xlsx'.format(date.today())

  wb = openpyxl.Workbook()
  ws = wb.get_active_sheet()
  row = 0
  columns = [(u"RAZON DE INICIO",250),(u"MOTIVO",250),(u"SOLICITANTE",250),(u"CORREO DE SOLICITANTE",250),(u"FECHA DE CREACION",50),(u"FECHA DE INGRESO DE INFORMACION",50),(u"ESTUDIANTE", 100), (u"GRADO", 50), (u"GRUPO FAMILIAR", 50), (u"FAMILIA DIRECTA", 50),(u"HIJOS ESTUDIANTES", 50),
  (U"FAMILIARES ADICIONALES", 50), (U"PARENTESCO", 50), (u"ENFERMEDADES CRONICAS",50), (u"DISCAPACIDADES", 50),
  (u"PROPIA",50),(u"ARRENDADA",50),(u"CEDIDAD TRABAJO",50),(u"CEDIDA FAMILIAR",50),(u"OTROS",50),(u"VILLA",50), (u"CASA",50),
  (u"DEPARTAMENTO",50), (u"CUARTOS",50), (u"OTROS",50), (u"CONCRETO",50), (u"ADOBE",50),(u"MADERA",50), (u"AGUA POTABLE",50), 
  (u"ALCANTARILLADO",50), (u"ENERGIA ELECTRICA",50), (u"TELEFONO FIJO",50), (u"INTERNET",50), (u"TV POR CABLE",50),
  (u"NO. INMUEBLES",50), (u"VALOR INMUEBLES",50), (u"NO. VEHICULOS", 50), (u"VALOR VEHICULOS", 50), (u"GASTOS VIVIENDA",50),
  (u"GASTOS ALIMENTACION",50), (u"GASTOS EDUCACION",50), (u"GASTOS TRANSPORTE",50), (u"GASTOS SALUD",50),(u"GASTOS VESTIMENTA",50),
  (u"OTROS GASTOS",50), (u"TOTAL GASTOS",50), (u"TOTAL INGRESOS", 50), (u"DIFERENCIA",50), (u"CAPACIDAD DE PAGO DE EDUCACION TOTAL",50),
  (U"CAPACIDAD DE PAGO DE EDUCACION POR ESTUDIANTE",50), (u"PENSION SOLICITADA", 50)]
  row_num = 0
  for col_num in xrange(len(columns)):
    c = ws.cell(row = row_num+1, column = col_num +1)
    c.value = columns[col_num][0]


  enviados = Evaluacion_Socioeco.objects.all().exclude(enviado = None).order_by("enviado")
  
  for enviado in enviados:
    for estudiante in enviado.estudiante_set.all():
      row_num += 1
      if enviado.requerido == False:
        razon_inicio = "Solicitado"
      else:
        razon_inicio = "Requerido"
      row = [razon_inicio, enviado.motivo_solicitud, enviado.solicitante, enviado.email, enviado.inicio, enviado.enviado, u"{0} {1} {2}".format(estudiante.apellido_paterno, estudiante.apellido_materno, estudiante.nombres), u"{0}".format(estudiante.nivel)]
      integrantes = 0
      integrantes_directos = 0
      integrantes_estudiantes = 0
      integrantes_adicionales = 0
      parentesco = ""
      enfermedad_cronica = ""
      discapacidad = ""
      for integrante in enviado.integrante_familia_set.all():
        integrantes += 1
        if integrante.parentesco.lower() in ["padre", "madre", "hijo", "hija", "papá", "mamá"]:
          integrantes_directos +=1
        else:
          integrantes_adicionales +=1
          if parentesco == "":
            parentesco += u"{0}".format(integrante.parentesco)
          else:
            parentesco += u" - {0}".format(integrante.parentesco)
        if integrante.actividad.lower() == "estudiante" and integrante.parentesco.lower() in ["hijo", "hija"]:
          integrantes_estudiantes +=1
        if integrante.enfermedad_cronica:
          if enfermedad_cronica == "":
            enfermedad_cronica += u"{0} - {1}".format(integrante.parentesco, integrante.enfermedad_cronica)
          else:
            enfermedad_cronica += u", {0} - {1}".format(integrante.parentesco, integrante.enfermedad_cronica)
        if integrante.discapacidad:
          if discapacidad == "":
            discapacidad += u"{0} - {1}".format(integrante.parentesco, integrante.discapacidad)
          else:
            discapacidad += u", {0} - {1}".format(integrante.parentesco, integrante.discapacidad)
      row.append(integrantes)
      row.append(integrantes_directos)
      row.append(integrantes_estudiantes)
      row.append(integrantes_adicionales)
      row.append(parentesco)
      row.append(enfermedad_cronica)
      row.append(discapacidad)

      if enviado.situacion_habitacional.posesion_vivienda == "PROPIA":
        row.append("SI")
        row.append("NO")
        row.append("NO")
        row.append("NO")
        row.append("NO")
      else:
        row.append("NO")
        if enviado.situacion_habitacional.posesion_vivienda == "ARRENDADA":
          row.append("SI")
          row.append("NO")
          row.append("NO")
          row.append("NO")
        else:
          row.append("NO")
          if enviado.situacion_habitacional.posesion_vivienda == "TRABAJO":
            row.append("SI")
            row.append("NO")
            row.append("NO")
          else:
            row.append("NO")
            if enviado.situacion_habitacional.posesion_vivienda == "FAMILIAR":
              row.append("SI")
              row.append("NO")
            else:
              row.append("NO")
              if enviado.situacion_habitacional.posesion_vivienda == "OTRA":
                row.append("SI")

      if enviado.situacion_habitacional.tipo_vivienda == "VILLA":
        row.append("SI")
        row.append("NO")
        row.append("NO")
        row.append("NO")
        row.append("NO")
      else:
        row.append("NO")
        if enviado.situacion_habitacional.tipo_vivienda == "CASA":
          row.append("SI")
          row.append("NO")
          row.append("NO")
          row.append("NO")
        else:
          row.append("NO")
          if enviado.situacion_habitacional.tipo_vivienda == "DEPARTAMENTO":
            row.append("SI")
            row.append("NO")
            row.append("NO")
          else:
            row.append("NO")
            if enviado.situacion_habitacional.tipo_vivienda == "CUARTO(S)":
              row.append("SI")
              row.append("NO")
            else:
              row.append("NO")
              row.append("SI")

      if enviado.situacion_habitacional.estructura_vivienda == "CONCRETO":
        row.append("SI")
        row.append("NO")
        row.append("NO")
      else:
        row.append("NO")
        if enviado.situacion_habitacional.estructura_vivienda == "ADOBE":
          row.append("SI")
          row.append("NO")
        else:
          row.append("NO")
          if enviado.situacion_habitacional.estructura_vivienda == "MADERA":
            row.append("SI")
          else:
            row.append("NO")

      if enviado.situacion_habitacional.agua_potable == True:
        row.append("SI")
      else:
        row.append("NO")

      if enviado.situacion_habitacional.alcantarillado == True:
        row.append("SI")
      else:
        row.append("NO")

      if enviado.situacion_habitacional.electricidad == True:
        row.append("SI")
      else:
        row.append("NO")

      if enviado.situacion_habitacional.telefono_fijo == True:
        row.append("SI")
      else:
        row.append("NO")

      if enviado.situacion_habitacional.internet == True:
        row.append("SI")
      else:
        row.append("NO")

      if enviado.situacion_habitacional.tv_cable == True:
        row.append("SI")
      else:
        row.append("NO")

      row.append(enviado.propiedades.cantidad_inmuebles)
      row.append(enviado.propiedades.valor_inmuebles)
      row.append(enviado.propiedades.cantidad_vehiculos)
      row.append(enviado.propiedades.valor_vehiculos)

      row.append(enviado.gastos.gastos_vivienda)
      row.append(enviado.gastos.gastos_alimentacion)
      row.append(enviado.gastos.gastos_educacion)
      row.append(enviado.gastos.gastos_transporte)
      row.append(enviado.gastos.gastos_salud)
      row.append(enviado.gastos.gastos_vestimenta)
      row.append(enviado.gastos.gastos_otros)

      total_gastos = enviado.gastos.gastos_vivienda + enviado.gastos.gastos_alimentacion + enviado.gastos.gastos_educacion + enviado.gastos.gastos_transporte + enviado.gastos.gastos_salud + enviado.gastos.gastos_vestimenta + enviado.gastos.gastos_otros
      row.append(total_gastos)

      total_ingresos = enviado.ingresos.dependencias + enviado.ingresos.negocios + enviado.ingresos.inversiones + enviado.ingresos.arriendos + enviado.ingresos.otros_ingreso 
      row.append(total_ingresos)

      row.append(total_ingresos - total_gastos)

      capacidad_total = total_ingresos / 4
      row.append(capacidad_total)
      try:
        row.append(capacidad_total / integrantes_estudiantes)
      except:
        row.append("")

      row.append(enviado.capacidad_pago)



      for col_num in xrange(len(row)):
        c = ws.cell(row = row_num +1, column = col_num +1)
        c.value = row[col_num]

  wb.save(response)
  return response


@login_required(login_url = "/user/login")
def Exportar_Enviados_Matriculados_Xls(request):

  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename={0}-MATRCOMP.xlsx'.format(date.today())

  wb = openpyxl.Workbook()
  ws = wb.get_active_sheet()
  row = 0
  columns = [(u"RAZON DE INICIO",250),(u"SOLICITANTE",250),(u"CORREO DE SOLICITANTE",250),(u"FECHA DE CREACION",50),(u"FECHA DE INGRESO DE INFORMACION",50),(u"ESTUDIANTE", 100), (u"GRADO", 50), (u"GRUPO FAMILIAR", 50), (u"FAMILIA DIRECTA", 50),(u"HIJOS ESTUDIANTES", 50),
  (U"FAMILIARES ADICIONALES", 50), (U"PARENTESCO", 50), (u"ENFERMEDADES CRONICAS",50), (u"DISCAPACIDADES", 50),
  (u"PROPIA",50),(u"ARRENDADA",50),(u"CEDIDAD TRABAJO",50),(u"CEDIDA FAMILIAR",50),(u"OTROS",50),(u"VILLA",50), (u"CASA",50),
  (u"DEPARTAMENTO",50), (u"CUARTOS",50), (u"OTROS",50), (u"CONCRETO",50), (u"ADOBE",50),(u"MADERA",50), (u"AGUA POTABLE",50), 
  (u"ALCANTARILLADO",50), (u"ENERGIA ELECTRICA",50), (u"TELEFONO FIJO",50), (u"INTERNET",50), (u"TV POR CABLE",50),
  (u"NO. INMUEBLES",50), (u"VALOR INMUEBLES",50), (u"NO. VEHICULOS", 50), (u"VALOR VEHICULOS", 50), (u"GASTOS VIVIENDA",50),
  (u"GASTOS ALIMENTACION",50), (u"GASTOS EDUCACION",50), (u"GASTOS TRANSPORTE",50), (u"GASTOS SALUD",50),(u"GASTOS VESTIMENTA",50),
  (u"OTROS GASTOS",50), (u"TOTAL GASTOS",50), (u"TOTAL INGRESOS", 50), (u"DIFERENCIA",50), (u"CAPACIDAD DE PAGO DE EDUCACION TOTAL",50),
  (U"CAPACIDAD DE PAGO DE EDUCACION POR ESTUDIANTE",50), (u"PENSION SOLICITADA", 50)]
  row_num = 0
  for col_num in xrange(len(columns)):
    c = ws.cell(row = row_num+1, column = col_num +1)
    c.value = columns[col_num][0]


  enviados = Evaluacion_Socioeco.objects.all().exclude(enviado = None).order_by("enviado")
  
  for enviado in enviados:
    for estudiante in enviado.estudiante_set.all():
      try:
        Matricula_Grupo.objects.get(matricula__estudiante__id_students__identity = estudiante.numero_id)
        row_num += 1
        if enviado.requerido == False:
          razon_inicio = "Solicitado"
        else:
          razon_inicio = "Requerido"
        row = [razon_inicio, enviado.solicitante, enviado.email, enviado.inicio, enviado.enviado, u"{0} {1} {2}".format(estudiante.apellido_paterno, estudiante.apellido_materno, estudiante.nombres), u"{0}".format(estudiante.nivel)]
        integrantes = 0
        integrantes_directos = 0
        integrantes_estudiantes = 0
        integrantes_adicionales = 0
        parentesco = ""
        enfermedad_cronica = ""
        discapacidad = ""
        for integrante in enviado.integrante_familia_set.all():
          integrantes += 1
          if integrante.parentesco.lower() in ["padre", "madre", "hijo", "hija", "papá", "mamá"]:
            integrantes_directos +=1
          else:
            integrantes_adicionales +=1
            if parentesco == "":
              parentesco += u"{0}".format(integrante.parentesco)
            else:
              parentesco += u" - {0}".format(integrante.parentesco)
          if integrante.actividad.lower() == "estudiante" and integrante.parentesco.lower() in ["hijo", "hija"]:
            integrantes_estudiantes +=1
          if integrante.enfermedad_cronica:
            if enfermedad_cronica == "":
              enfermedad_cronica += u"{0} - {1}".format(integrante.parentesco, integrante.enfermedad_cronica)
            else:
              enfermedad_cronica += u", {0} - {1}".format(integrante.parentesco, integrante.enfermedad_cronica)
          if integrante.discapacidad:
            if discapacidad == "":
              discapacidad += u"{0} - {1}".format(integrante.parentesco, integrante.discapacidad)
            else:
              discapacidad += u", {0} - {1}".format(integrante.parentesco, integrante.discapacidad)
        row.append(integrantes)
        row.append(integrantes_directos)
        row.append(integrantes_estudiantes)
        row.append(integrantes_adicionales)
        row.append(parentesco)
        row.append(enfermedad_cronica)
        row.append(discapacidad)

        if enviado.situacion_habitacional.posesion_vivienda == "PROPIA":
          row.append("SI")
          row.append("NO")
          row.append("NO")
          row.append("NO")
          row.append("NO")
        else:
          row.append("NO")
          if enviado.situacion_habitacional.posesion_vivienda == "ARRENDADA":
            row.append("SI")
            row.append("NO")
            row.append("NO")
            row.append("NO")
          else:
            row.append("NO")
            if enviado.situacion_habitacional.posesion_vivienda == "TRABAJO":
              row.append("SI")
              row.append("NO")
              row.append("NO")
            else:
              row.append("NO")
              if enviado.situacion_habitacional.posesion_vivienda == "FAMILIAR":
                row.append("SI")
                row.append("NO")
              else:
                row.append("NO")
                if enviado.situacion_habitacional.posesion_vivienda == "OTRA":
                  row.append("SI")

        if enviado.situacion_habitacional.tipo_vivienda == "VILLA":
          row.append("SI")
          row.append("NO")
          row.append("NO")
          row.append("NO")
          row.append("NO")
        else:
          row.append("NO")
          if enviado.situacion_habitacional.tipo_vivienda == "CASA":
            row.append("SI")
            row.append("NO")
            row.append("NO")
            row.append("NO")
          else:
            row.append("NO")
            if enviado.situacion_habitacional.tipo_vivienda == "DEPARTAMENTO":
              row.append("SI")
              row.append("NO")
              row.append("NO")
            else:
              row.append("NO")
              if enviado.situacion_habitacional.tipo_vivienda == "CUARTO(S)":
                row.append("SI")
                row.append("NO")
              else:
                row.append("NO")
                row.append("SI")

        if enviado.situacion_habitacional.estructura_vivienda == "CONCRETO":
          row.append("SI")
          row.append("NO")
          row.append("NO")
        else:
          row.append("NO")
          if enviado.situacion_habitacional.estructura_vivienda == "ADOBE":
            row.append("SI")
            row.append("NO")
          else:
            row.append("NO")
            if enviado.situacion_habitacional.estructura_vivienda == "MADERA":
              row.append("SI")
            else:
              row.append("NO")

        if enviado.situacion_habitacional.agua_potable == True:
          row.append("SI")
        else:
          row.append("NO")

        if enviado.situacion_habitacional.alcantarillado == True:
          row.append("SI")
        else:
          row.append("NO")

        if enviado.situacion_habitacional.electricidad == True:
          row.append("SI")
        else:
          row.append("NO")

        if enviado.situacion_habitacional.telefono_fijo == True:
          row.append("SI")
        else:
          row.append("NO")

        if enviado.situacion_habitacional.internet == True:
          row.append("SI")
        else:
          row.append("NO")

        if enviado.situacion_habitacional.tv_cable == True:
          row.append("SI")
        else:
          row.append("NO")

        row.append(enviado.propiedades.cantidad_inmuebles)
        row.append(enviado.propiedades.valor_inmuebles)
        row.append(enviado.propiedades.cantidad_vehiculos)
        row.append(enviado.propiedades.valor_vehiculos)

        row.append(enviado.gastos.gastos_vivienda)
        row.append(enviado.gastos.gastos_alimentacion)
        row.append(enviado.gastos.gastos_educacion)
        row.append(enviado.gastos.gastos_transporte)
        row.append(enviado.gastos.gastos_salud)
        row.append(enviado.gastos.gastos_vestimenta)
        row.append(enviado.gastos.gastos_otros)

        total_gastos = enviado.gastos.gastos_vivienda + enviado.gastos.gastos_alimentacion + enviado.gastos.gastos_educacion + enviado.gastos.gastos_transporte + enviado.gastos.gastos_salud + enviado.gastos.gastos_vestimenta + enviado.gastos.gastos_otros
        row.append(total_gastos)

        total_ingresos = enviado.ingresos.dependencias + enviado.ingresos.negocios + enviado.ingresos.inversiones + enviado.ingresos.arriendos + enviado.ingresos.otros_ingreso 
        row.append(total_ingresos)

        row.append(total_ingresos - total_gastos)

        capacidad_total = total_ingresos / 4
        row.append(capacidad_total)
        try:
          row.append(capacidad_total / integrantes_estudiantes)
        except:
          row.append("")

        row.append(enviado.capacidad_pago)

        for col_num in xrange(len(row)):
          c = ws.cell(row = row_num +1, column = col_num +1)
          c.value = row[col_num]

      except:
        pass

  wb.save(response)
  return response



@login_required(login_url = "/user/login")
def Exportar_Resumen_Xls(request):

  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename={0}_RESTOTAL.xlsx'.format(date.today())

  wb = openpyxl.Workbook()
  ws = wb.get_active_sheet()
  row = 0
  columns = [(u"ID FORM", 100), (u"SOLICITANTE",100), (u"CORREO",100), (u"ACEPTADO", 100) , (u"COMPLETADO", 100)]
  row_num = 0
  for col_num in xrange(len(columns)):
    c = ws.cell(row = row_num+1, column = col_num +1)
    c.value = columns[col_num][0]


  enviados = Evaluacion_Socioeco.objects.all()
  
  for enviado in enviados:
    row_num += 1
    row = []
    row.append(u"{}".format(enviado.pk))
    row.append(enviado.solicitante)
    row.append(enviado.email)
    try:
      if enviado.aceptacion.aceptacion == True:
        row.append("X")
      else:
        row.append("-")
    except:
      row.append("-")
    if enviado.enviado == None:
      row.append("-")
    else:
      row.append("X")
    for col_num in xrange(len(row)):
      c = ws.cell(row = row_num +1, column = col_num +1)
      c.value = row[col_num]
  wb.save(response)
  return response


@login_required(login_url = "/user/login")
def Exportar_Resumen_Req_Xls(request):

  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename={0}_RESREQUE.xlsx'.format(date.today())

  wb = openpyxl.Workbook()
  ws = wb.get_active_sheet()
  row = 0
  columns = [(u"ID FORM", 100), (u"SOLICITANTE",100), (u"CORREO",100), (u"ACEPTADO", 100) , (u"COMPLETADO", 100)]
  row_num = 0
  for col_num in xrange(len(columns)):
    c = ws.cell(row = row_num+1, column = col_num +1)
    c.value = columns[col_num][0]


  enviados = Evaluacion_Socioeco.objects.filter(requerido = True)
  
  for enviado in enviados:
    row_num += 1
    row = []
    row.append(u"{}".format(enviado.pk))
    row.append(enviado.solicitante)
    row.append(enviado.email)
    try:
      if enviado.aceptacion.aceptacion == True:
        row.append("X")
      else:
        row.append("-")
    except:
      row.append("-")
    if enviado.enviado == None:
      row.append("-")
    else:
      row.append("X")
    for col_num in xrange(len(row)):
      c = ws.cell(row = row_num +1, column = col_num +1)
      c.value = row[col_num]
  wb.save(response)
  return response


@login_required(login_url = "/user/login")
def Exportar_Resumen_No_Req_Xls(request):

  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename={0}_RESNOREQUE.xlsx'.format(date.today())

  wb = openpyxl.Workbook()
  ws = wb.get_active_sheet()
  row = 0
  columns = [(u"ID FORM", 100), (u"SOLICITANTE",100), (u"CORREO",100), (u"ACEPTADO", 100) , (u"COMPLETADO", 100)]
  row_num = 0
  for col_num in xrange(len(columns)):
    c = ws.cell(row = row_num+1, column = col_num +1)
    c.value = columns[col_num][0]


  enviados = Evaluacion_Socioeco.objects.all().exclude(requerido = True)
  
  for enviado in enviados:
    row_num += 1
    row = []
    row.append(u"{}".format(enviado.pk))
    row.append(enviado.solicitante)
    row.append(enviado.email)
    try:
      if enviado.aceptacion.aceptacion == True:
        row.append("X")
      else:
        row.append("-")
    except:
      row.append("-")
    if enviado.enviado == None:
      row.append("-")
    else:
      row.append("X")
    for col_num in xrange(len(row)):
      c = ws.cell(row = row_num +1, column = col_num +1)
      c.value = row[col_num]
  wb.save(response)
  return response


@login_required(login_url = "/user/login")
def Exportar_Est_1718_Resumen(request):
  response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  response['Content-Disposition'] = 'attachment; filename={0}_RESEST.xlsx'.format(date.today())

  wb = openpyxl.Workbook()
  ws = wb.get_active_sheet()
  row = 0
  columns = [(u"ID ESTUDIANTE", 100), (u"NOMBRES COMPLETOS",100), (u"GRADO",100), (u"# FORMS", 100) , (u"# INICIADOS", 100), (u"# COMPLETADOS", 100)]
  row_num = 0
  for col_num in xrange(len(columns)):
    c = ws.cell(row = row_num+1, column = col_num +1)
    c.value = columns[col_num][0]

  estudiantes = Matricula_Grupo.objects.filter(fin = None)
  for x in estudiantes:
    row_num += 1
    row = [u"{0}".format(x.matricula.estudiante.id_students.identity), u"{0}".format(x.matricula.estudiante.id_students.get_legal_name()), u"{0}".format(x.grupo.nombre)]
    registros = Evaluacion_Socioeco.objects.filter(estudiante__numero_id = x.matricula.estudiante.id_students.identity)
    if registros.exists():
      no_forms = registros.count()
      no_iniciados = registros.filter(aceptacion__aceptacion = True).count()
      no_completados = registros.exclude(enviado = None).count()
      row.append(u"{0}".format(no_forms))
      row.append(u"{0}".format(no_iniciados))
      row.append(u"{0}".format(no_completados))
    else:
      row.append("0")
      row.append("-")
      row.append("-")
    for col_num in xrange(len(row)):
      c = ws.cell(row = row_num +1, column = col_num +1)
      c.value = row[col_num]
  wb.save(response)
  return response      






@login_required(login_url = '/user/login/')
@user_passes_test(is_financiero)
def Detalle_Enviado(request, pedido_pk):
  enviado = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
  return render(request, "socioeco_sam/detalle_enviado.html", {"enviado":enviado})



@login_required(login_url = '/user/login/')
@user_passes_test(is_financiero)

def Enviados(request):
  enviados = Evaluacion_Socioeco.objects.all().exclude(enviado = None).order_by("enviado")
  return render(request, "socioeco_sam/enviados.html", {"enviados":enviados})

@login_required(login_url = '/user/login/')
@user_passes_test(is_financiero)

def Nuevo_Pedido(request):
	if request.method == "POST":
		form = EvaluacionForm(request.POST)
		if form.is_valid():
			#try:

			pedido = form.save(commit=False)
			pedido.capacidad_pago = 0
			pedido.save()
			html_template = '<p>{{pedido.solicitante}}:<br>Hemos recibido un pedido de análisis financiero de su parte. Por favor, de click en este <a href="https://canaldocente.montebelloacademy.org/socioeconomico/{{pedido.pk}}/">enlace</a> para cargar la información correspondiente.</p><br><p>Si existe algún inconveniente, por favor responda a este correo para resolverlo.</p><br><br><p>Saludos cordiales,</p><br><p>Departamento Financiero - Montebello Academy</p>'
			context = Context({"pedido":pedido})
			template = Template(html_template)
			content = template.render(context)
			send_mail(subject = 'Registro de Asistencia Financiera',
				message = "Hola",
				from_email = 'financiero@montebelloacademy.org',
				recipient_list = [pedido.email, "rodrigo@montebelloacademy.org", "ximena@montebelloacademy.org"],
				auth_user = "financiero@montebelloacademy.org",
				auth_password = "Montebello1",
				fail_silently=False,
				html_message = content
			)
				#,
				#auth_user = "financiero@montebelloacademy.org",
				#auth_password = "Montebello1")
			#except:
				#return render(request, "socioeco_sam/error.html", {"pedido":pedido})
			return render(request, "socioeco_sam/exito.html", {"pedido":pedido})
		else:
			print form.errors
			return render(request, "socioeco_sam/inicio.html", {"form":form})
	else:
		form = EvaluacionForm()
		return render(request, "socioeco_sam/inicio.html", {"form":form})


def Resumen(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
	
	if pedido.enviado:
		if pedido.requerido == False:
			return render(request, "socioeco_sam/resultados.html", {"pedido":pedido})
		else:
			return render(request, "socioeco_sam/resultados_requerido.html", {"pedido":pedido})
	
	try:
		Aceptacion.objects.get(evaluacion_socioeco = pedido, aceptacion = True)
		if Integrante_Familia.objects.filter(evaluacion_socioeco = pedido, validado = False).exists():
			sin_validar = True
		else:
			sin_validar = False
		errores = []
		if not pedido.representante_set.exists():
			errores.append("Representantes")
		if not pedido.estudiante_set.exists():
			errores.append("Estudiantes")
		try:
			pedido.domicilio
		except:
			errores.append("Domicilio")
		if not pedido.integrante_familia_set.exists() or pedido.integrante_familia_set.filter(validado = False).exists():	
			errores.append("Grupo Familiar")
		try:
			pedido.situacion_habitacional
		except:
			errores.append("Situación Habitacional")
		try:
			pedido.responsable_gastos
		except:
			errores.append("Responsable de Gastos")
		try:
			pedido.propiedades
		except:
			errores.append("Propiedades")
		try:
			pedido.ingresos
		except:
			errores.append("Ingresos")
		try:
			pedido.gastos
		except:
			errores.append("Gastos")
		if pedido.motivo_solicitud == "":
			errores.append("Motivo de Solicitud")
		return render(request, "socioeco_sam/resumen.html", {"pedido":pedido, "sin_validar":sin_validar, "errores":errores})

	except Aceptacion.DoesNotExist:
		if pedido.requerido == False:
			return render(request, "socioeco_sam/aceptacion.html", {"pedido":pedido})
		else:
			return render(request, "socioeco_sam/aceptacion_requerido.html", {"pedido":pedido})


def Confirmar(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)

	try:
		aceptacion = Aceptacion.objects.create(evaluacion_socioeco = pedido, aceptacion = True)
	except:
		aceptacion = Aceptacion.objects.get(evaluacion_socioeco = pedido)
		aceptacion.aceptacion = True
		aceptacion.save()
	return redirect("resumen", pedido_pk = pedido.pk)


def Terminar(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)

	errores = []

	if Integrante_Familia.objects.filter(evaluacion_socioeco = pedido, validado = False).exists():
		sin_validar = True
		errores.append("Grupo Familiar")
	else:
		sin_validar = False

	if pedido.enviado:
		if pedido.requerido == False:
			return render(request, "socioeco_sam/resultados.html", {"pedido":pedido})
		else:
			return render(request, "socioeco_sam/resultados_requerido.html", {"pedido":pedido})

	#errores = []

	if not pedido.representante_set.exists():
		errores.append("Representantes")
	if not pedido.estudiante_set.exists():
		errores.append("Estudiantes")
	try:
		pedido.domicilio
	except:
		errores.append("Domicilio")
	if not pedido.integrante_familia_set.exists():
		errores.append("Grupo Familiar")
	try:
		pedido.situacion_habitacional
	except:
		errores.append("Situación Habitacional")
	try:
		pedido.responsable_gastos
	except:
		errores.append("Responsable de Gastos")
	try:
		pedido.propiedades
	except:
		errores.append("Propiedades")
	try:
		pedido.ingresos
	except:
		errores.append("Ingresos")
	try:
		pedido.gastos
	except:
		errores.append("Gastos")
	if pedido.motivo_solicitud == "":
		errores.append("Motivo de Solicitud")

	if errores:
		errores2 = True

	if len(errores) != 0:
		return render(request, "socioeco_sam/resumen.html", {"pedido":pedido, "errores":errores, "sin_validar":sin_validar, "errores2":errores2})

	else:
		pedido.enviado = date.today()
		pedido.save()
		html_requerido = """
			<p>Ha finalizado el ingreso de la información para la calificación de subsidio familiar.</p>

              <br>
              <p><strong>Nombre del Representante:</strong> {{pedido.solicitante}}</p>
              <p><strong>Correo Electrónico:</strong> {{pedido.email}}</p>
              <p><strong>Número de trámite:</strong> {{pedido.pk}}</p>
              <h3><strong>Información de los Representantes:</strong></h3>


              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Nombres y Apellidos</th>
                    <th scope="col">Relación</th>
                    <th scope="col">Fecha de Nacimiento</th>
                    <th scope="col">Teléfono Fijo</th>
                    <th scope="col">Teléfono Celular</th>
                  </tr>
                </thead>
                <tbody>
                {% for representante in pedido.representante_set.all %}
                <tr>
                  <td>{{representante.nombres}} {{representante.apellido_paterno}}</td>
                  <td>{{representante.relacion}}</td>
                  <td>{{representante.fecha_nacimiento}}</td>
                  <td>{{representante.telefono_casa}}</td>
                  <td>{{representante.telefono_celular}}</td>
                </tr>
                {% endfor %}
	              </tbody>
	            </table>


                  <h3><strong>Información de los Estudiantes:</strong></h3>


              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Nombres y Apellidos</th>
                    <th scope="col">Fecha de Nacimiento</th>
                    <th scope="col">Grado / Nivel</th>
                  </tr>
                </thead>
                <tbody>
              {% for estudiante in pedido.estudiante_set.all %}
                <tr>
                  <td>{{estudiante.nombres}} {{estudiante.apellido_paterno}}</td>
                  <td>{{estudiante.fecha_nacimiento}}</td>
                  <td>{{estudiante.nivel.nivel}}{{estudiante.nivel.paralelo}}</td>
                </tr>
              {% endfor %}
                </tbody>
              </table>

                  <h3><strong>Información del Domicilio:</strong></h3>

              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Calle Principal</th>
                    <th scope="col">Calle Secundaria</th>
                    <th scope="col">Número</th>
                    <th scope="col">Barrio</th>
                    <th scope="col">Teléfono</th>
                    
                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.domicilio.direccion_principal}}</td>
                  <td>{{pedido.domicilio.direccion_secundaria}}</td>
                  <td>{{pedido.domicilio.direccion_numero}}</td>
                  <td>{{pedido.domicilio.barrio}}</td>
                  <td>{{pedido.domicilio.telefono_casa}}</td>

                </tr>
              
              </tbody>
            </table>


                  <h3><strong>Información de los Integrantes de la Familia:</strong></h3>

              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Nombres Completos</th>
                    <th scope="col">Fecha de Nacimiento</th>
                    <th scope="col">Documento de Identidad</th>
                    <th scope="col">Estado Civil</th>
                    <th scope="col">Parentesco</th>
                    <th scope="col">Nivel de Educación</th>
                    <th scope="col">Título</th>
                    <th scope="col">Ocupación</th>
                    </tr>
                </thead>
                <tbody>           
              {% for integrante in pedido.integrante_familia_set.all %}
                <tr>
                  <td>{{integrante.nombres_completos}}</td>
                  <td>{{integrante.fecha_nacimiento}}</td>
                  <td>{{integrante.numero_id}}</td>
                  <td>{{integrante.estado_civil}}</td>
                  <td>{{integrante.parentesco}}</td>
                  <td>{{integrante.nivel_estudios}}</td>
                  <td>{{integrante.titulo}}</td>
                  <td>{{integrante.actividad}}</td>
                  </tr>
              {% endfor %}
              </tbody>
            </table>


                  <h3><strong>Información de la Situación Habitacional:</strong></h3>

              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Posesión de Vivienda</th>
                    <th scope="col">Tipo de Vivienda</th>
                    <th scope="col">Estructura de Vivienda</th>

                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.situacion_habitacional.posesion_vivienda}}</td>
                  <td>{{pedido.situacion_habitacional.tipo_vivienda}}</td>
                  <td>{{pedido.situacion_habitacional.estructura_vivienda}}</td>

                </tr>
              </tbody>
            </table>

            <h3><strong>Información del Responsable de los Gastos de los Estudiantes:</strong></h3> 

              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Responsable de Gastos Educativos</th>
                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.responsable_gastos.responsable_gastos}}</td>


               </tr>
              </tbody>
            </table>


                  <h3><strong>Información de las Propiedades:</strong></h3>
              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Cantidad de Propiedades</th>
                    <th scope="col">Valor Total de Propiedades</th>
                    <th scope="col">Cantidad de Vehículos</th>
                    <th scope="col">Valor Total de Propiedades</th>
                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.propiedades.cantidad_inmuebles}}</td>
                  <td>{{pedido.propiedades.valor_inmuebles}}</td>
                  <td>{{pedido.propiedades.cantidad_vehiculos}}</td>
                  <td>{{pedido.propiedades.valor_vehiculos}}</td>
                </tr>
              </tbody>
            </table>

                  <h3><strong>Información de los Ingresos:</strong></h3>
              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Relación de Dependencia</th>
                    <th scope="col">Negocios</th>
                    <th scope="col">Inversiones</th>
                    <th scope="col">Arriendos</th>
                    <th scope="col">Otros</th>
                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.ingresos.dependencias}}</td>
                  <td>{{pedido.ingresos.negocios}}</td>
                  <td>{{pedido.ingresos.inversiones}}</td>
                  <td>{{pedido.ingresos.arriendos}}</td>
                  <td>{{pedido.ingresos.otros_ingreso}}</td>
                </tr>
              </tbody>
            </table>


                  <h3><strong>Información de los Gastos:</strong></h3>
              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Vivienda</th>
                    <th scope="col">Alimentación</th>
                    <th scope="col">Educación</th>
                    <th scope="col">Transporte</th>
                    <th scope="col">Salud</th>
                    <th scope="col">Vestimenta</th>
                    <th scope="col">Otros Gastos</th>
                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.gastos.gastos_vivienda}}</td>
                  <td>{{pedido.gastos.gastos_alimentacion}}</td>
                  <td>{{pedido.gastos.gastos_educacion}}</td>
                  <td>{{pedido.gastos.gastos_transporte}}</td>
                  <td>{{pedido.gastos.gastos_salud}}</td>
                  <td>{{pedido.gastos.gastos_vestimenta}}</td>
                  <td>{{pedido.gastos.gastos_otros}}</td>
                </tr>
              </tbody>
            </table>

                  <br><br> 



                  <p>Hemos enviado esta información a su correo electrónico, {{pedido.email}}. Por favor imprímala y entréguela en el Departamento Fianciero con la documentación que demuestre los datos ingresados.</br></br>
                  Muchas gracias por su colaboración.
                  </br></br>
                  Saludos cordiales,
                  </br>
                  Montebello Academy
                  </p>

		"""
		html_norequerido = """              
			<p>Ha finalizado de ingresar la información para enviar el pedido de evaluación socio-económica que usted ha solicitado. Vamos a analizar su requerimiento y le daremos una respuesta.</p>

        <br>
        <p><strong>Nombre del Solicitante:</strong> {{pedido.solicitante}}</p>
        <p><strong>Correo Electrónico:</strong> {{pedido.email}}</p>
        <p><strong>Número de solicitud:</strong> {{pedido.pk}}</p>
        <h3><strong>Información de los Representantes:</strong></h3>


              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Nombres y Apellidos</th>
                    <th scope="col">Relación</th>
                    <th scope="col">Fecha de Nacimiento</th>
                    <th scope="col">Teléfono Fijo</th>
                    <th scope="col">Teléfono Celular</th>
                  </tr>
                </thead>
                <tbody>
                {% for representante in pedido.representante_set.all %}
                <tr>
                  <td>{{representante.nombres}} {{representante.apellido_paterno}}</td>
                  <td>{{representante.relacion}}</td>
                  <td>{{representante.fecha_nacimiento}}</td>
                  <td>{{representante.telefono_casa}}</td>
                  <td>{{representante.telefono_celular}}</td>
                </tr>
                {% endfor %}
              </tbody>
            </table>


                  <h3><strong>Información de los Estudiantes:</strong></h3>


              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Nombres y Apellidos</th>
                    <th scope="col">Fecha de Nacimiento</th>
                    <th scope="col">Grado / Nivel</th>
                  </tr>
                </thead>
                <tbody>
              {% for estudiante in pedido.estudiante_set.all %}
                <tr>
                  <td>{{estudiante.nombres}} {{estudiante.apellido_paterno}}</td>
                  <td>{{estudiante.fecha_nacimiento}}</td>
                  <td>{{estudiante.nivel.nivel}}{{estudiante.nivel.paralelo}}</td>
                </tr>
              {% endfor %}
                </tbody>
              </table>

                  <h3><strong>Información del Domicilio:</strong></h3>

              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Calle Principal</th>
                    <th scope="col">Calle Secundaria</th>
                    <th scope="col">Número</th>
                    <th scope="col">Barrio</th>
                    <th scope="col">Teléfono</th>
                    
                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.domicilio.direccion_principal}}</td>
                  <td>{{pedido.domicilio.direccion_secundaria}}</td>
                  <td>{{pedido.domicilio.direccion_numero}}</td>
                  <td>{{pedido.domicilio.barrio}}</td>
                  <td>{{pedido.domicilio.telefono_casa}}</td>

                </tr>
              
              </tbody>
            </table>


                  <h3><strong>Información de los Integrantes de la Familia:</strong></h3>

              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Nombres Completos</th>
                    <th scope="col">Fecha de Nacimiento</th>
                    <th scope="col">Documento de Identidad</th>
                    <th scope="col">Estado Civil</th>
                    <th scope="col">Parentesco</th>
                    <th scope="col">Nivel de Educación</th>
                    <th scope="col">Título</th>
                    <th scope="col">Ocupación</th>
                    </tr>
                </thead>
                <tbody>           
              {% for integrante in pedido.integrante_familia_set.all %}
                <tr>
                  <td>{{integrante.nombres_completos}}</td>
                  <td>{{integrante.fecha_nacimiento}}</td>
                  <td>{{integrante.numero_id}}</td>
                  <td>{{integrante.estado_civil}}</td>
                  <td>{{integrante.parentesco}}</td>
                  <td>{{integrante.nivel_estudios}}</td>
                  <td>{{integrante.titulo}}</td>
                  <td>{{integrante.actividad}}</td>
                  </tr>
              {% endfor %}
              </tbody>
            </table>


                  <h3><strong>Información de la Situación Habitacional:</strong></h3>

              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Posesión de Vivienda</th>
                    <th scope="col">Tipo de Vivienda</th>
                    <th scope="col">Estructura de Vivienda</th>

                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.situacion_habitacional.posesion_vivienda}}</td>
                  <td>{{pedido.situacion_habitacional.tipo_vivienda}}</td>
                  <td>{{pedido.situacion_habitacional.estructura_vivienda}}</td>

                </tr>
              </tbody>
            </table>

            <h3><strong>Información del Responsable de los Gastos de los Estudiantes:</strong></h3> 

              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Responsable de Gastos Educativos</th>
                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.responsable_gastos.responsable_gastos}}</td>


               </tr>
              </tbody>
            </table>


                  <h3><strong>Información de las Propiedades:</strong></h3>
              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Cantidad de Propiedades</th>
                    <th scope="col">Valor Total de Propiedades</th>
                    <th scope="col">Cantidad de Vehículos</th>
                    <th scope="col">Valor Total de Propiedades</th>
                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.propiedades.cantidad_inmuebles}}</td>
                  <td>{{pedido.propiedades.valor_inmuebles}}</td>
                  <td>{{pedido.propiedades.cantidad_vehiculos}}</td>
                  <td>{{pedido.propiedades.valor_vehiculos}}</td>
                </tr>
              </tbody>
            </table>

                  <h3><strong>Información de los Ingresos:</strong></h3>
              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Relación de Dependencia</th>
                    <th scope="col">Negocios</th>
                    <th scope="col">Inversiones</th>
                    <th scope="col">Arriendos</th>
                    <th scope="col">Otros</th>
                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.ingresos.dependencias}}</td>
                  <td>{{pedido.ingresos.negocios}}</td>
                  <td>{{pedido.ingresos.inversiones}}</td>
                  <td>{{pedido.ingresos.arriendos}}</td>
                  <td>{{pedido.ingresos.otros_ingreso}}</td>
                </tr>
              </tbody>
            </table>


                  <h3><strong>Información de los Gastos:</strong></h3>
              <table class="table">
                <thead>
                  <tr>
                    <th scope="col">Vivienda</th>
                    <th scope="col">Alimentación</th>
                    <th scope="col">Educación</th>
                    <th scope="col">Transporte</th>
                    <th scope="col">Salud</th>
                    <th scope="col">Vestimenta</th>
                    <th scope="col">Otros Gastos</th>
                  </tr>
                </thead>
                <tbody>
                <tr>
                  <td>{{pedido.gastos.gastos_vivienda}}</td>
                  <td>{{pedido.gastos.gastos_alimentacion}}</td>
                  <td>{{pedido.gastos.gastos_educacion}}</td>
                  <td>{{pedido.gastos.gastos_transporte}}</td>
                  <td>{{pedido.gastos.gastos_salud}}</td>
                  <td>{{pedido.gastos.gastos_vestimenta}}</td>
                  <td>{{pedido.gastos.gastos_otros}}</td>
                </tr>
              </tbody>
            </table>

                  <h3><strong>Motivo de la Solicitud</strong></h3>
                  <p><strong>Motivo:</strong>{{pedido.motivo_solicitud}}</p>
                  <p><strong>Capacidad de pago:</strong>USD {{pedido.capacidad_pago}} por cada estudiante</p>
                  <br><br> 



                  <p>Hemos enviado esta información a su correo electrónico, {{pedido.email}}. Por favor imprímala y entréguela en el Departamento Fianciero con la documentación que demuestre los datos ingresados.</br></br>
                  Muchas gracias por su colaboración.
                  </br></br>
                  Saludos cordiales,
                  </br>
                  Montebello Academy
                  </p>"""
		formato_requerido = "socioeco_sam/resultados_requerido.html"
		formato_norequerido = "socioeco_sam/resultados.html"

		if pedido.requerido == False:
			html = html_norequerido
			formato = formato_norequerido

		else:
			html = html_requerido
			formato = formato_requerido

		template = Template(html)
		context = Context({"pedido":pedido})
		content = template.render(context)
		send_mail(subject = 'Subsidio Familiar - Registro Finalizado',
			message = content,
			from_email = 'financiero@montebelloacademy.org',
			recipient_list = [pedido.email],
			auth_user = "financiero@montebelloacademy.org",
			auth_password = "Montebello1",
			fail_silently=False,
			html_message = content
		)

		return render(request, formato , {"pedido":pedido})


def Finalizar(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)

	if request.method=="POST":
		form = FinalizarForm(request.POST, instance = pedido)
		if form.is_valid():
			registro = form.save(commit = False)
			registro.evaluacion_socioeco = pedido
			registro.save()
			return redirect("resumen", pedido_pk = pedido.pk)

		else:
			print form.errors
			return render(request, "socioeco_sam/finalizar.html", {"form":form, "pedido":pedido, "error":error})

	else:
		form = FinalizarForm(instance = pedido)
		return render(request, "socioeco_sam/finalizar.html", {"form":form, "pedido":pedido})

def Editar_Finalizar(request, pedido_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		if request.method == "POST":
			form = FinalizarForm(request.POST, instance = pedido)
			if form.is_valid():
				registro = form.save(commit = False)
				registro.evaluacion_socioeco = pedido
				registro.save()
				return redirect("resumen", pedido_pk = pedido.pk)
			else:
				print form.errors
				return render(request, "socioeco_sam/finalizar.html", {"form":form, "pedido":pedido})
		else:
			form = FinalizarForm(instance = pedido)
			return render(request, "socioeco_sam/finalizar.html", {"form":form, "pedido":pedido})
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Registrar_Representante(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
	if request.method == "POST":
		form = RepresentanteForm(request.POST)
		if form.is_valid():
			registro_representante = form.save(commit = False)
			registro_representante.evaluacion_socioeco = pedido
			try:
				registro_representante.save()
			except IntegrityError:
				error = "Ya se ha creado un representante con estos datos. Por favor, ingrese la información de otro representante."
				print form.errors
				return render(request, "socioeco_sam/representante.html", {"form":form, "pedido":pedido, "error":error})
			try:
				integrante = Integrante_Familia.objects.create(
					evaluacion_socioeco = pedido,
					nombres_completos = "{0} {1} {2}".format(registro_representante.nombres, registro_representante.apellido_paterno, registro_representante.apellido_materno),
					fecha_nacimiento = registro_representante.fecha_nacimiento,
					numero_id = registro_representante.numero_id,
					estado_civil = registro_representante.estado_civil,
					parentesco = registro_representante.relacion,
					nivel_estudios = registro_representante.nivel_estudios,
					titulo = registro_representante.titulo,
					actividad = registro_representante.cargo
					)
			except IntegrityError:
				integrante = Integrante_Familia.objects.get(nombres_completos = "{0} {1} {2}".format(registro_representante.nombres, registro_representante.apellido_paterno, registro_representante.apellido_materno))
				integrante.fecha_nacimiento = registro_representante.fecha_nacimiento
				integrante.numero_id = registro_representante.numero_id
				integrante.estado_civil = registro_representante.estado_civil
				integrante.parentesco = registro_representante.relacion
				integrante.nivel_estudios = registro_representante.nivel_estudios
				integrante.titulo = registro_representante.titulo
				integrante.actividad = registro_representante.cargo
				integrante.save()			
			return redirect("resumen", pedido_pk = pedido.pk)
		else:
			print form.errors
			return render(request, "socioeco_sam/representante.html", {"form":form, "pedido":pedido})
	else:
		form = RepresentanteForm()
		return render(request, "socioeco_sam/representante.html", {"form":form, "pedido":pedido})


def Editar_Representante(request, pedido_pk, representante_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		representante = Representante.objects.get(pk = representante_pk)
		if request.method == "POST":
			form = RepresentanteForm(request.POST, instance = representante)
			if form.is_valid():
				registro_representante = form.save(commit = False)
				registro_representante.evaluacion_socioeco = pedido
				registro_representante.save()
				try:
					integrante = Integrante_Familia.objects.create(
						evaluacion_socioeco = pedido,
						nombres_completos = "{0} {1} {2}".format(registro_representante.nombres, registro_representante.apellido_paterno, registro_representante.apellido_materno),
						fecha_nacimiento = registro_representante.fecha_nacimiento,
						numero_id = registro_representante.numero_id,
						estado_civil = registro_representante.estado_civil,
						parentesco = registro_representante.relacion,
						nivel_estudios = registro_representante.nivel_estudios,
						titulo = registro_representante.titulo,
						actividad = registro_representante.cargo
						)
				except IntegrityError:
					integrante = Integrante_Familia.objects.get(nombres_completos = "{0} {1} {2}".format(registro_representante.nombres, registro_representante.apellido_paterno, registro_representante.apellido_materno))
					integrante.fecha_nacimiento = registro_representante.fecha_nacimiento
					integrante.numero_id = registro_representante.numero_id
					integrante.estado_civil = registro_representante.estado_civil
					integrante.parentesco = registro_representante.relacion
					integrante.nivel_estudios = registro_representante.nivel_estudios
					integrante.titulo = registro_representante.titulo
					integrante.actividad = registro_representante.cargo
					integrante.save()
				return redirect("resumen", pedido_pk = pedido.pk)
			else:
				print form.errors
				return render(request, "socioeco_sam/representante.html", {"form":form, "pedido":pedido})
		else:
			form = RepresentanteForm(instance = representante)
			return render(request, "socioeco_sam/representante.html", {"form":form, "pedido":pedido})
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Representante.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Eliminar_Representante(request, pedido_pk, representante_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		representante = Representante.objects.get(pk = representante_pk)
		try:
			integrante = Integrante_Familia.objects.get(nombres_completos = "{0} {1} {2}".format(representante.nombres, representante.apellido_paterno, representante.apellido_materno))
			integrante.delete()
		except:
			pass
		representante.delete()
		return redirect("resumen", pedido_pk = pedido.pk)
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Representante.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Registrar_Estudiante(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
	if request.method == "POST":
		form = EstudianteForm(request.POST)
		form.fields["nivel"].queryset = Clase.objects.filter(periodo_lectivo__actual = True).order_by("nivel__nombre", "paralelo")
		if form.is_valid():
			registro_estudiante = form.save(commit = False)
			registro_estudiante.evaluacion_socioeco = pedido
			try:
				registro_estudiante.save()
			except IntegrityError:
				error = "Ya se ha creado un estudiante con estos datos. Por favor, ingrese la información de otro estudiante."
				print form.errors
				return render(request, "socioeco_sam/estudiante.html", {"form":form, "pedido":pedido, "error":error})			
			try:
				integrante = Integrante_Familia.objects.create(
					evaluacion_socioeco = pedido,
					nombres_completos = "{0} {1} {2}".format(registro_estudiante.nombres, registro_estudiante.apellido_paterno, registro_estudiante.apellido_materno),
					fecha_nacimiento = registro_estudiante.fecha_nacimiento,
					numero_id = registro_estudiante.numero_id,
					estado_civil = "SOLTERO",
					parentesco = "Hijo",
					nivel_estudios = "BASICA",
					titulo = "",
					actividad = "Estudiante"
					)
			except IntegrityError:
				integrante = Integrante_Familia.objects.get(nombres_completos = "{0} {1} {2}".format(registro_estudiante.nombres, registro_estudiante.apellido_paterno, registro_estudiante.apellido_materno))
				integrante.fecha_nacimiento = registro_estudiante.fecha_nacimiento
				integrante.numero_id = registro_estudiante.numero_id
				integrante.estado_civil = "SOLTERO"
				integrante.parentesco = "Hijo"
				integrante.nivel_estudios = "BASICA"
				integrante.actividad = "Estudiante"
				integrante.save()
			return redirect("resumen", pedido_pk = pedido.pk)
		else:
			print form.errors
			return render(request, "socioeco_sam/estudiante.html", {"form":form, "pedido":pedido})
	else:
		form = EstudianteForm()
		form.fields["nivel"].queryset = Clase.objects.filter(periodo_lectivo__actual = True).order_by("nivel__nombre", "paralelo")
		return render(request, "socioeco_sam/estudiante.html", {"form":form, "pedido":pedido})


def Editar_Estudiante(request, pedido_pk, estudiante_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		estudiante = Estudiante.objects.get(pk = estudiante_pk)
		if request.method == "POST":
			form = EstudianteForm(request.POST, instance = estudiante)
			form.fields["nivel"].queryset = Clase.objects.filter(periodo_lectivo__actual = True).order_by("nivel__nombre", "paralelo")
			if form.is_valid():
				registro_estudiante = form.save(commit = False)
				registro_estudiante.evaluacion_socioeco = pedido
				registro_estudiante.save()
				try:
					integrante = Integrante_Familia.objects.create(
						evaluacion_socioeco = pedido,
						nombres_completos = "{0} {1} {2}".format(registro_estudiante.nombres, registro_estudiante.apellido_paterno, registro_estudiante.apellido_materno),
						fecha_nacimiento = registro_estudiante.fecha_nacimiento,
						numero_id = registro_estudiante.numero_id,
						estado_civil = "SOLTERO",
						parentesco = "Hijo",
						nivel_estudios = "BASICA",
						titulo = "",
						actividad = "Estudiante"
						)
				except IntegrityError:
					integrante = Integrante_Familia.objects.get(nombres_completos = "{0} {1} {2}".format(registro_estudiante.nombres, registro_estudiante.apellido_paterno, registro_estudiante.apellido_materno))
					integrante.fecha_nacimiento = registro_estudiante.fecha_nacimiento
					integrante.numero_id = registro_estudiante.numero_id
					integrante.estado_civil = "SOLTERO"
					integrante.parentesco = "Hijo"
					integrante.nivel_estudios = "BASICA"
					integrante.actividad = "Estudiante"
					integrante.save()
				return redirect("resumen", pedido_pk = pedido.pk)
			else:
				print form.errors
				return render(request, "socioeco_sam/estudiante.html", {"form":form, "pedido":pedido})
		else:
			form = EstudianteForm(instance = estudiante)
			form.fields["nivel"].queryset = Clase.objects.filter(periodo_lectivo__actual = True).order_by("nivel.nombre")
			return render(request, "socioeco_sam/estudiante.html", {"form":form, "pedido":pedido})
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Estudiante.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Eliminar_Estudiante(request, pedido_pk, estudiante_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		estudiante = Estudiante.objects.get(pk = estudiante_pk)
		try:
			integrante = Integrante_Familia.objects.get(nombres_completos = "{0} {1} {2}".format(estudiante.nombres, estudiante.apellido_paterno, estudiante.apellido_materno))
			integrante.delete()
		except:
			pass
		estudiante.delete()
		return redirect("resumen", pedido_pk = pedido.pk)
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Estudiante.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Registrar_Domicilio(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
	if request.method == "POST":
		form = DomicilioForm(request.POST)
		if form.is_valid():
			registro_domicilio = form.save(commit = False)
			registro_domicilio.evaluacion_socioeco = pedido
			registro_domicilio.save()
			return redirect("resumen", pedido_pk = pedido.pk)
		else:
			print form.errors
			return render(request, "socioeco_sam/domicilio.html", {"form":form, "pedido":pedido})
	else:
		form = DomicilioForm()
		return render(request, "socioeco_sam/domicilio.html", {"form":form, "pedido":pedido})


def Editar_Domicilio(request, pedido_pk, domicilio_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		domicilio = Domicilio.objects.get(pk = domicilio_pk)
		if request.method == "POST":
			form = DomicilioForm(request.POST, instance = domicilio)
			if form.is_valid():
				registro_domicilio = form.save(commit = False)
				registro_domicilio.evaluacion_socioeco = pedido
				registro_domicilio.save()
				return redirect("resumen", pedido_pk = pedido.pk)
			else:
				print form.errors
				return render(request, "socioeco_sam/domicilio.html", {"form":form, "pedido":pedido})
		else:
			form = DomicilioForm(instance = domicilio)
			return render(request, "socioeco_sam/domicilio.html", {"form":form, "pedido":pedido})
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Domicilio.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Registrar_Integrante(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
	if request.method == "POST":
		form = IntegranteForm(request.POST)
		if form.is_valid():
			registro_integrante = form.save(commit = False)
			registro_integrante.evaluacion_socioeco = pedido
			registro_integrante.validado = True
			try:
				registro_integrante.save()
			except IntegrityError:
				error = "Ya se ha creado un miembro familiar con estos datos. Por favor, ingrese la información de otro miembro familiar."
				print form.errors
				return render(request, "socioeco_sam/integrante_familia.html", {"form":form, "pedido":pedido, "error":error})

			return redirect("resumen", pedido_pk = pedido.pk)
		else:
			print form.errors
			return render(request, "socioeco_sam/integrante_familia.html", {"form":form, "pedido":pedido})
	else:
		form = IntegranteForm()
		return render(request, "socioeco_sam/integrante_familia.html", {"form":form, "pedido":pedido})

def Editar_Integrante(request, pedido_pk, integrante_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		integrante = Integrante_Familia.objects.get(pk = integrante_pk)
		if request.method == "POST":
			form = IntegranteForm(request.POST, instance = integrante)
			if form.is_valid():
				registro_integrante = form.save(commit = False)
				registro_integrante.evaluacion_socioeco = pedido
				registro_integrante.validado = True
				registro_integrante.save()
				return redirect("resumen", pedido_pk = pedido.pk)
			else:
				print form.errors
				return render(request, "socioeco_sam/integrante_familia.html", {"form":form, "pedido":pedido})
		else:
			form = IntegranteForm(instance = integrante)
			return render(request, "socioeco_sam/integrante_familia.html", {"form":form, "pedido":pedido})
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Integrante_Familia.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Eliminar_Integrante(request, pedido_pk, integrante_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		integrante = Integrante_Familia.objects.get(pk = integrante_pk)
		integrante.delete()
		return redirect("resumen", pedido_pk = pedido.pk)
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Integrante_Famlia.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Registrar_Situacion_Habitacional(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
	if request.method == "POST":
		form = HabitacionalForm(request.POST)
		if form.is_valid():
			situacion = form.save(commit =False)
			if request.POST["tipo_vivienda"] == "OTRO":
				situacion.tipo_vivienda = request.POST.get("tipo_vivienda_otro")
			situacion.evaluacion_socioeco = pedido
			situacion.save()
			return redirect("resumen", pedido_pk = pedido.pk)
		else:
			print form.errors
			return render(request, "socioeco_sam/habitacional.html", {"form":form, "pedido":pedido})
	else:
		form = HabitacionalForm()
		return render(request, "socioeco_sam/habitacional.html", {"form":form, "pedido":pedido})

def Editar_Situacion_Habitacional(request, pedido_pk, situacion_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		situacion = Situacion_Habitacional.objects.get(pk = situacion_pk)
		if request.method == "POST":
			form = HabitacionalForm(request.POST, instance = situacion)
			if form.is_valid():
				situacion = form.save(commit = False)
				if request.POST["tipo_vivienda"] == "OTRO":
					situacion.tipo_vivienda = request.POST.get("tipo_vivienda_otro")
				situacion.evaluacion_socioeco = pedido
				situacion.save()
				return redirect("resumen", pedido_pk = pedido.pk)
			else:
				print form.errors
				return render(request, "socioeco_sam/habitacional.html", {"form":form, "pedido":pedido})
		else:
			form = HabitacionalForm(instance = situacion)
			return render(request, "socioeco_sam/habitacional.html", {"form":form, "pedido":pedido})
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Situacion_Habitacional.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)

def Registrar_Responsable_Gastos(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
	if request.method == "POST":
		form = ResponsableForm(request.POST)
		if form.is_valid():
			responsable = form.save(commit =False)
			if request.POST["responsable_2"]:
				nombre = request.POST.get("responsable_2")
				relacion = request.POST.get("responsable_3")
				responsable.responsable_gastos = "{0} - {1}".format(nombre, relacion)
			responsable.evaluacion_socioeco = pedido
			responsable.save()
			return redirect("resumen", pedido_pk = pedido.pk)
		else:
			print form.errors
			return render(request, "socioeco_sam/responsable.html", {"form":form, "pedido":pedido})
	else:
		form = ResponsableForm()
		return render(request, "socioeco_sam/responsable.html", {"form":form, "pedido":pedido})

def Editar_Responsable_Gastos(request, pedido_pk, responsable_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		responsable = Responsable_Gastos.objects.get(pk = responsable_pk)
		if request.method == "POST":
			form = ResponsableForm(request.POST, instance = responsable)
			if form.is_valid():
				responsable = form.save(commit = False)
				if request.POST["responsable_2"]:
					nombre = request.POST.get("responsable_2")
					relacion = request.POST.get("responsable_3")
					responsable.responsable_gastos = "{0} - {1}".format(nombre, relacion)				
				responsable.evaluacion_socioeco = pedido
				responsable.save()
				return redirect("resumen", pedido_pk = pedido.pk)
			else:
				print form.errors
				return render(request, "socioeco_sam/responsable.html", {"form":form, "pedido":pedido})
		else:
			form = ResponsableForm(instance = responsable)
			return render(request, "socioeco_sam/responsable.html", {"form":form, "pedido":pedido})
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Responsable_Gastos.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Registrar_Propiedades(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
	if request.method == "POST":
		form = PropiedadesForm(request.POST)
		if form.is_valid():
			propiedades = form.save(commit =False)
			propiedades.evaluacion_socioeco = pedido
			propiedades.save()
			return redirect("resumen", pedido_pk = pedido.pk)
		else:
			print form.errors
			return render(request, "socioeco_sam/propiedades.html", {"form":form, "pedido":pedido})
	else:
		form = PropiedadesForm()
		return render(request, "socioeco_sam/propiedades.html", {"form":form, "pedido":pedido})

def Editar_Propiedades(request, pedido_pk, propiedades_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		propiedades = Propiedades.objects.get(pk = propiedades_pk)
		if request.method == "POST":
			form = PropiedadesForm(request.POST, instance = propiedades)
			if form.is_valid():
				propiedades = form.save(commit = False)
				propiedades.evaluacion_socioeco = pedido
				propiedades.save()
				return redirect("resumen", pedido_pk = pedido.pk)
			else:
				print form.errors
				return render(request, "socioeco_sam/propiedades.html", {"form":form, "pedido":pedido})
		else:
			form = PropiedadesForm(instance = propiedades)
			return render(request, "socioeco_sam/propiedades.html", {"form":form, "pedido":pedido})
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Propiedades.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Registrar_Ingresos(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
	if request.method == "POST":
		form = IngresosForm(request.POST)
		if form.is_valid():
			ingresos = form.save(commit =False)
			ingresos.evaluacion_socioeco = pedido
			ingresos.save()
			return redirect("resumen", pedido_pk = pedido.pk)
		else:
			print form.errors
			return render(request, "socioeco_sam/ingresos.html", {"form":form, "pedido":pedido})
	else:
		form = IngresosForm()
		return render(request, "socioeco_sam/ingresos.html", {"form":form, "pedido":pedido})

def Editar_Ingresos(request, pedido_pk, ingresos_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		ingresos = Ingresos.objects.get(pk = ingresos_pk)
		if request.method == "POST":
			form = IngresosForm(request.POST, instance = ingresos)
			if form.is_valid():
				ingresos = form.save(commit = False)
				ingresos.evaluacion_socioeco = pedido
				ingresos.save()
				return redirect("resumen", pedido_pk = pedido.pk)
			else:
				print form.errors
				return render(request, "socioeco_sam/ingresos.html", {"form":form, "pedido":pedido})
		else:
			form = IngresosForm(instance = ingresos)
			return render(request, "socioeco_sam/ingresos.html", {"form":form, "pedido":pedido})
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Ingresos.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)


def Registrar_Gastos(request, pedido_pk):
	pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
	if request.method == "POST":
		form = GastosForm(request.POST)
		if form.is_valid():
			gastos = form.save(commit =False)
			gastos.evaluacion_socioeco = pedido
			gastos.save()
			return redirect("resumen", pedido_pk = pedido.pk)
		else:
			print form.errors
			return render(request, "socioeco_sam/gastos.html", {"form":form, "pedido":pedido})
	else:
		form = GastosForm()
		return render(request, "socioeco_sam/gastos.html", {"form":form, "pedido":pedido})

def Editar_Gastos(request, pedido_pk, gastos_pk):
	try:
		pedido = Evaluacion_Socioeco.objects.get(pk = pedido_pk)
		gastos = Gastos.objects.get(pk = gastos_pk)
		if request.method == "POST":
			form = GastosForm(request.POST, instance = gastos)
			if form.is_valid():
				gastos = form.save(commit = False)
				gastos.evaluacion_socioeco = pedido
				gastos.save()
				return redirect("resumen", pedido_pk = pedido.pk)
			else:
				print form.errors
				return render(request, "socioeco_sam/gastos.html", {"form":form, "pedido":pedido})
		else:
			form = GastosForm(instance = gastos)
			return render(request, "socioeco_sam/gastos.html", {"form":form, "pedido":pedido})
	except Evaluacion_Socioeco.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
	except Gastos.DoesNotExist:
		return redirect("resumen", pedido_pk = pedido.pk)
# Create your views here.
