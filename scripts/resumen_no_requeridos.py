# -*- coding: utf-8 -*-

import openpyxl
from socioeco_sam.models import Evaluacion_Socioeco, Estudiante

wb = openpyxl.Workbook()
ws = wb.get_active_sheet()
columns = [(u"ID FORM", 100), (u"SOLICITANTE",100), (u"CORREO",100), (u"ACEPTADO", 100) , (u"COMPLETADO", 100)]
row_num = 0
for col_num in xrange(len(columns)):
  c = ws.cell(row = row_num+1, column = col_num +1)
  c.value = columns[col_num][0]

enviados = Evaluacion_Socioeco.objects.exclude(requerido = True)


for enviado in enviados:
  row_num += 1
  row = []
  row.append(u"{}".format(enviado.pk))
  row.append(enviado.solicitante)
  row.append(enviado.email)
  try:
  	if enviado.aceptacion.aceptacion == True:
  	  row.append("X")
  	else:
  	  row.append("-")
  except:
  	row.append("-")
  if enviado.enviado == None:
  	row.append("-")
  else:
  	row.append("X")
  for col_num in xrange(len(row)):
    c = ws.cell(row = row_num +1, column = col_num +1)
    c.value = row[col_num]

wb.save("estado_no_requeridas.xlsx")