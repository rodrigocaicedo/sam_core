# -*- coding: utf-8 -*-

import openpyxl
from socioeco_sam.models import Evaluacion_Socioeco

wb = openpyxl.Workbook()
ws = wb.get_active_sheet()
columns = [(u"ID",100), (u"FAMILIA",100), (u"SOLICITANTE",100), (u"CORREO",100)]
row_num = 0
for col_num in xrange(len(columns)):
  c = ws.cell(row = row_num+1, column = col_num +1)
  c.value = columns[col_num][0]


enviados = Evaluacion_Socioeco.objects.filter(requerido = False)


for enviado in enviados:
  row_num += 1
  row = [u"{0}".format(enviado.pk), u"{0}".format(enviado.familia), u"{0}".format(enviado.solicitante), u"{0}".format(enviado.email)]
  for col_num in xrange(len(row)):
    c = ws.cell(row = row_num +1, column = col_num +1)
    c.value = row[col_num]

wb.save("solicitudes_anteriores.xlsx")