# -*- coding: utf-8 -*-

import openpyxl
from socioeco_sam.models import Evaluacion_Socioeco, Estudiante

wb = openpyxl.Workbook()
ws = wb.get_active_sheet()
columns = [(u"ID ESTUDIANTE",100),(u"ESTUDIANTE",100), (u"NO.FORMS",100), (u"ID FORM", 100), (u"SOLICITANTE",100), (u"CORREO",100), (u"ACEPTADO", 100) , (u"COMPLETADO", 100)]
row_num = 0
for col_num in xrange(len(columns)):
  c = ws.cell(row = row_num+1, column = col_num +1)
  c.value = columns[col_num][0]

enviados = Evaluacion_Socioeco.objects.filter(requerido = True)

estudiantes = Estudiante.objects.filter(evaluacion_socioeco__requerido = True)

for estudiante in estudiantes:
  row_num += 1
  row = [u"{}".format(estudiante.numero_id), u"{0} {1} {2}".format(estudiante.apellido_paterno, estudiante.apellido_materno, estudiante.nombres)]
  forms = Evaluacion_Socioeco.objects.filter(requerido = True, estudiante__numero_id = estudiante.numero_id)
  row.append(forms.count())
  row.append(u"{}".format(estudiante.evaluacion_socioeco.pk))
  row.append(estudiante.evaluacion_socioeco.solicitante)
  row.append(estudiante.evaluacion_socioeco.email)
  try:
  	if estudiante.evaluacion_socioeco.aceptacion.aceptacion == True:
  	  row.append("X")
  	else:
  	  row.append("-")
  except:
  	row.append("-")
  if estudiante.evaluacion_socioeco.enviado == None:
  	row.append("-")
  else:
  	row.append("X")
  for col_num in xrange(len(row)):
    c = ws.cell(row = row_num +1, column = col_num +1)
    c.value = row[col_num]

wb.save("estado_requeridas.xlsx")