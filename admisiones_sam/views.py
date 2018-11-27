#!/usr/bin/env python
# -*- coding: utf-8 -*-
import json,unicodedata, operator,os,csv, openpyxl,datetime, random, string
import sqlite3
from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from django.db.models import Min,Q,Count,Max
from django.db.models.functions import Lower
from django.core.exceptions import ObjectDoesNotExist
from django.forms import modelformset_factory
from django.shortcuts import render_to_response, redirect, render
from django.http import HttpResponseRedirect, HttpResponse
from django.template import RequestContext,loader, Context
from django.contrib.auth.decorators import login_required, user_passes_test
from admisiones_sam.models import PeriodSchool, Applications,Type_Tours,Tours,State_Applications,Type_Tests,Test,Forms,Quotas,Invoices,Mails,Reports_Tests,Det_Reports,Documents_Type,Det_Documents,AccountingDepartment,AcademicSecretary,Type_Grade,AccountingDepartmentTest
from admisiones_sam.forms import Applicationsnewform,TypeToursnewform,Toursnewform,Toursupdateform,TypeToursEditform,PeriodNewform,PeriodEditform,ApplicationsChangeState,TypeTestsnewform,FormsNewform,QuotaNewform,InvoicesNewform,MailsNewform,NewReportsTests,NewDetReports,NewDocuments_Type,NewDetDocuments,NewAccounting,GradeNewform,Testsnewform,CustomUserform
from usuarios_sam.models import CustomUser,Students,People,RelationshipStudent, Student, Relative, Student_Relative, Relative_Details
from usuarios_sam.outlookservice import create_student, create_tutor, create_parent
from comunicaciones_sam.models import MailerMessage
from general_sam.models import Matricula, Nivel, Periodo_Lectivo, Clase, Aptitud_Matricula
from datetime import date
#from models import Prompts
from operator import itemgetter
from collections import OrderedDict

def is_admisiones(user):
    return user.groups.filter(name = "Admisiones").exists()
    
def is_psicologia(user):
    return user.groups.filter(name = "Psicologia").exists()

#def

# Create your views here.

def Home(request):
    return render(request, "admisiones_sam/Home.html")  


def Applications_new(request):
    
    periodo=PeriodSchool.objects.filter(per_state=1)
   
    if request.method == "POST":
    #selecciono el maximo codigo del banco asignado para incrementar en uno
        
        obj=Applications.objects.all().aggregate(Max('id_applications'))
        ultimo = Applications.objects.get(id_applications=obj['id_applications__max'])
        codigo=int(ultimo.bank_code)
        codigo=codigo+1
        codigo=str(codigo).zfill(6)
        
        form = Applicationsnewform(request.POST)
        if form.is_valid():
            new_app = form.save(commit=False)
            new_app.bank_code=codigo
            new_app.save()
            request.session['id_formulario']=new_app.pk  #cookies
            estados = State_Applications(state='Error',final_date=None,observations='Aplicacion Inicial',id_applications=new_app)
            estados.save()
            return redirect("SeleccionaTour")
        else:
            print form.errors
            return render(request, "admisiones_sam/AppCreate.html", {"form":form,"periodo":periodo})
    else:
        form = Applicationsnewform(initial={'birth_country': 'EC','country_home': 'EC'})
        return render(request, "admisiones_sam/AppCreate.html", {"form":form,"periodo":periodo})

def Applications_newCH(request):
    if request.method == "POST":
        mensaje = MailerMessage()
        form = Applicationsnewform(request.POST)
        if form.is_valid():
            new_app = form.save(commit=False)
            new_app.type_student='OLD'
            new_app.save()
            request.session['id_formulario']=new_app.pk  #cookies
            
            estados = State_Applications(state='Error',final_date=None,observations='Aplicacion Inicial',id_applications=new_app)
            estados.save()
           
            mensaje.general_message(new_app.mail_tutor, "admisiones@montebelloacademy.org", 3, new_app)
            #mensaje.application_message(new_app.pk,3) #envio de mail a representantes para confirmar ingreso de solicitud
            return redirect("MensajeSolicitudHermanos")
            
            
        else:
            print form.errors
            return render(request, "admisiones_sam/AppCreateCH.html", {"form":form})
    else:
        form = Applicationsnewform(initial={'birth_country': 'EC','country_home': 'EC'})
        return render(request, "admisiones_sam/AppCreateCH.html", {"form":form})
  

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)


    
def Applications_edit(request, id_applications):
    periodo=PeriodSchool.objects.filter(per_state=1)
    aplicacion = Applications.objects.get(pk = id_applications)
   
    if request.method == "POST":
        form = Applicationsnewform(request.POST, instance = aplicacion)
        if form.is_valid():
            form.save()
           
            return redirect('Administracion')
        else:
            print form.errors
            return render(request,'admisiones_sam/AppEdit.html', {"form":form,"aplicacion":aplicacion,"periodo":periodo})
    else:
       
        form = Applicationsnewform(instance = aplicacion)
        return render(request,'admisiones_sam/AppEdit.html', {"form":form,"aplicacion":aplicacion,"periodo":periodo})


@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)    
def DeleteApplications (request, id_applications): 

    aplicacion = Applications.objects.get(id_applications = id_applications)
    aplicacion.delete()
    return redirect('/admisiones/Administracion')
    
        
        
        
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)        
def Applications_Select(request, id_applications):     

 
    aplicacion = Applications.objects.get(pk = id_applications)
   
    '''if request.method == "POST":
        form = Applicationsnewform(request.POST, instance = aplicacion) 
        if form.is_valid():
            form.save()
            return redirect('Administracion')
        else:
            print form.errors
            return render(request,'admisiones_sam/AppEdit.html', {"form":form,"aplicacion":aplicacion})
    else:'''
       
    form = Applicationsnewform(instance = aplicacion)
    return render(request,'admisiones_sam/AppSelect.html', {"form":form,"aplicacion":aplicacion})           

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def ChangeState(request,id_applications):
    aplicante=Applications.objects.get(id_applications = id_applications)
    estados = State_Applications.objects.get(id_applications = id_applications, final_date__isnull=True)
    debug=1   
    mensajes = MailerMessage()
    
    
    if estados.state=='Error' and  aplicante.type_student=='OLD': # Estudiantes con hermanos que no toman tour debe continuar a documentos
        lista=['Documentos','Terminado']
        mensaje='Debe continuar el proceso a Entrega Documentos o dar por terminado el Trámite'
        
    elif  estados.state=='Error' and  aplicante.type_student=='NEW': # debe cambiar  Terminado o agendar nueva cita para que continue a Pendiente (otra fecha), Ingresado(si escogio fecha)
        lista=['Documentos','Terminado'] # estaba solo terminado 06-04-2018
        mensaje='Revise el Trámite o Puede dar por Terminado este trámite'
        
    elif  estados.state=='Pendiente' : # debe cambiar  Terminado o agendar nueva cita para que continue a Pendiente (otra fecha), Ingresado(si escogio fecha)
        lista=['Terminado','Documentos'] #solo era terminado 06-04-2018
        mensaje='Puede continuar con el proceso a Entrega de Documentos o puede dar por Terminado este trámite, o continuar con el mismo agendando una nueva cita para el Tour'
    elif estados.state=='Ingresado': # debe cambiar a entrega Documentos o Terminar el Tramite
        lista=['Documentos','Terminado']
        mensaje='Puede continuar con el proceso a Entrega de Documentos, o dar por terminado el trámite'
    elif estados.state=='Terminado': #solo puede cambiar a Pendiente para ingresar una nueva fecha de Tour
        lista=['Pendiente']
        mensaje='Debe cambiar a estado Pendiente para lo cual agende una nueva cita para el Tour'
    elif estados.state=='Documentos' : # debe cambiar  Terminado .
        lista=['Terminado']
        mensaje='Puede dar por Terminado este trámite.'
    elif estados.state=='Evaluaciones' : # debe cambiar  Terminado .
        lista=['Terminado']
        mensaje='Puede dar por Terminado este trámite.'  
    elif estados.state=='Contabilidad' : # debe cambiar  Terminado .
        lista=['Terminado']
        mensaje='Puede dar por Terminado este trámite.'
    elif estados.state=='Secretaria' : # debe cambiar  Terminado .
        lista=['Terminado']
        mensaje='Puede dar por Terminado este trámite.'          
    else:
        lista=['']
        
        mensaje='No se puede cambiar los estados de la solicitud'
        return render(request,"admisiones_sam/ChangeStateMessage.html", {"mensaje":mensaje})  
    
    
    
    if request.method == "POST":
        debug=2
        form = ApplicationsChangeState(request.POST)
        if form.is_valid():
            debug=3
            item = form.save(commit=False)
            item.initial_date=date.today()
            #item.observations = "Pedido de Documentación"
            item.id_applications = estados.id_applications
            item.final_date= None
            seleccion_estado=item.state
            
            item.save()
            
            
            estados.final_date = date.today()
            estados.save()
            
            if seleccion_estado=='Documentos':  
               
                # JB 24-04-2018
                #cargo el registro en la tabla de pago de evaluaciones en contabilidad.
                contabilidad = AccountingDepartmentTest(state='False',id_applications=aplicante)
                contabilidad.save()
                
                #envio de mail a representantes para subir la documentacion 
                #mensajes.application_message(aplicante.pk,7)
                mensajes.general_message(aplicante.mail_tutor, "admisiones@montebelloacademy.org", 7, aplicante)
                # Mensaje del sistema  a Contabilidad para que active el codigo del banco para que el representante pueda pagar las evaluaciones.
                mensajes.general_message("jeaneth@montebelloacademy.org",'admisiones@montebelloacademy.org', 33, aplicante)
                
            else:
                pass
            
            return redirect('Administracion')
            
           
        else:
            debug=4
            print form.errors
            return render(request,'admisiones_sam/ChangeState.html', {"form":form,"estados":estados,"lista":lista,"mensaje":mensaje})
    else:
        debug=5   
       
        form = ApplicationsChangeState()
        return render(request,'admisiones_sam/ChangeState.html', {"form":form,"estados":estados,"lista":lista,"mensaje":mensaje})
        
        
def Tours_select(request): 
    dia_inicial=1 #dia martes
    dia_final=3 #dia jueves
    maximo=2 # numero de registros de fechas que deben estar minimo parametrizadas.
    
    
    try:
        id_formulario=request.session['id_formulario']  
    except:
        
        return redirect('NuevaSolicitud')
    
    #valido el numero de cupos utilizados en cada fecha de tour     
    aplicacion=Applications.objects.get(id_applications=id_formulario)
    estado=State_Applications.objects.get(id_applications = aplicacion)
    lista_fechas=[]
    fechatour = Type_Tours.objects.filter(state=True,date_tour__gt=date.today()).order_by('date_tour')
    
    
    for x in fechatour:
        registradas=Tours.objects.filter(id_typetour_id=x.id_typetour, 
                                    id_applications_id__in=State_Applications.objects.filter(~Q(state="Finalizado"),
                                    final_date__isnull=True).
                                    values('id_applications_id') ).count()
           
        if registradas > x.quota or registradas == x.quota:
            
            pass
        else:

            if len(lista_fechas) == 3:
                pass
            else:
                lista_fechas.append(x.id_typetour)
            fechatour=Type_Tours.objects.filter(id_typetour__in=lista_fechas)
    
    # fin validacion
    
    # valido que existan por los menos dos fechas de tour para seleccion del representante, si no hay se debe ingresar las que falten para completar 2.
    
    
    try:
        numero_fechas=Type_Tours.objects.filter(state=True,date_tour__gt=date.today()).order_by('date_tour').count()
    except ObjectDoesNotExist:
        numero_fechas=0
    
        
        
    if numero_fechas < 2 and numero_fechas>=1:     #inserta un registro
        fechas_disponibles=Type_Tours.objects.filter(state=True,date_tour__gt=date.today()).aggregate(Max('date_tour'))
        dia = fechas_disponibles["date_tour__max"].weekday()
        
        if dia==1:
            nueva_fecha = fechas_disponibles["date_tour__max"] + datetime.timedelta(days=2)
         
        elif dia==3:
            nueva_fecha = fechas_disponibles["date_tour__max"] + datetime.timedelta(days=5)
        else:
            pass #nueva_fecha=datetime.datetime.today()
            
        nuevas_fechas = Type_Tours(name_tour=str(nueva_fecha),date_tour=nueva_fecha,quota=20,state=True)
        nuevas_fechas.save()
        
        
    elif numero_fechas==0:  #inserta dos registro
        fechas_disponibles=Type_Tours.objects.filter(state=True,date_tour__lt=date.today()).aggregate(Max('date_tour'))
        dia = fechas_disponibles["date_tour__max"].weekday()
    
        if dia==1: #martes
            nueva_fecha=fechas_disponibles["date_tour__max"]+ datetime.timedelta(days=2)
            nueva_fecha1=fechas_disponibles["date_tour__max"]+ datetime.timedelta(days=7)
            
            nuevas_fechas = Type_Tours(name_tour=str(nueva_fecha),date_tour=nueva_fecha,quota=20,state=True)
            nuevas_fechas.save()
            
            nuevas_fechas = Type_Tours(name_tour=str(nueva_fecha1),date_tour=(nueva_fecha1),quota=20,state=True)
            nuevas_fechas.save()
        
        elif dia==3:   #jueves
            nueva_fecha=fechas_disponibles["date_tour__max"]+ datetime.timedelta(days=5)
            nueva_fecha1=fechas_disponibles["date_tour__max"]+ datetime.timedelta(days=7)
            
            nuevas_fechas = Type_Tours(name_tour=str(nueva_fecha),date_tour=nueva_fecha,quota=20,state=True)
            nuevas_fechas.save()
            
            nuevas_fechas = Type_Tours(name_tour=str(nueva_fecha1),date_tour=(nueva_fecha1),quota=20,state=True)
            nuevas_fechas.save()
        else:
            pass
     
    else:
        pass
        
    
         
    # fin validacion

    
    
    if request.method == "POST":
        mensaje = MailerMessage()
        checkbox_estado = request.POST.get('otra_fecha', '')
        
        if checkbox_estado == "True":
            estado.state='Pendiente'
            estado.save()

            #mensaje.application_message(aplicacion.pk,5)   #envio de mail a representantes para confirmar ingreso de solicitud y solicitar agendar fecha de tour
            mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 5, aplicacion)
            return redirect('MensajeSolicitudEstado')
        else:
            
            
            form = Toursnewform(request.POST)
            
            if form.is_valid():
                item1 = form.save(commit=False)
                #fecha=item1.date_tour
                
                #Valida que no se haya completado el cupo asignado a ese tour(fecha)
                #cupos=Type_Tours.objects.get(date_tour=fecha)
                #utilizado=Tour.objects.filter(id_typetour_id=cupo.id_typetour).count(1)
            
                #if utilizado<cupos:
                
                #control para que no grave dos registros de fecha de tour
        
                if  Tours.objects.filter(id_applications=aplicacion.pk, estado=True).count()>0:
                     pass
                     #return redirect('MensajeSolicitudEstado')
                else:
                    item1.save()
                
                estado.state='Ingresado'
                estado.save()
                
             
                #mensaje.application_message(aplicacion.pk,4) #envio de mail a representantes para confirmar ingreso de solicitud y fecha de tour
                mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 4, aplicacion)
                return redirect('MensajeSolicitud')
                #else:
                #return redirect('MensajeSolicitud')
            else:
                print form.errors
                return render(request,"admisiones_sam/SelectTypeTour.html", {"form":form,"registradas":registradas,"fechatour":lista_fechas})
            
            
    else:
    
        form = Toursnewform(initial={'id_applications':aplicacion,'estado':True})
        form.fields["id_typetour"].queryset = fechatour
        
        return render(request,"admisiones_sam/SelectTypeTour.html", {"form":form,"fechatour":lista_fechas})


@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
        
def DateTourUpdate (request,id_applications):
    mensaje = MailerMessage()
  
    aplicacion=Applications.objects.get(id_applications=id_applications)
    request.session['id_aplicante']=aplicacion.pk  #cookies
    try:
        #if aplicacion.tours_set.get(id_applications = aplicacion , estado = True).exists():
        fecha_inicial = aplicacion.tours_set.get(id_applications = aplicacion , estado = True)
        tipo=Tours.objects.get(id_typetour= fecha_inicial.id_typetour,id_applications=id_applications,estado=True)
        fechatour = Type_Tours.objects.filter(state=True,date_tour__gt=date.today()).exclude(pk = fecha_inicial.id_typetour.pk).order_by('date_tour')
    except ObjectDoesNotExist:
        fecha_inicial=None
        fechatour = Type_Tours.objects.filter(state=True,date_tour__gt=date.today()).order_by('date_tour')
        
    lista_fechas=[]
    
    
    
    for x in fechatour:
        registradas=Tours.objects.filter(id_typetour_id=x.id_typetour, 
                                    id_applications_id__in=State_Applications.objects.filter(~Q(state="Finalizado"),
                                    final_date__isnull=True).
                                    values('id_applications_id') ).count()
           
        if registradas > x.quota or registradas == x.quota:
            
            pass
        else:
            lista_fechas.append(x.id_typetour)
            fechatour=Type_Tours.objects.filter(id_typetour__in=lista_fechas)    


    
   
    estado=State_Applications.objects.get(id_applications = aplicacion, final_date__isnull = True)
    
    if estado.state == 'Error' or estado.state == 'Pendiente' or estado.state == 'Ingresado' or estado.state == 'Terminado':# valida los estados de la solicitud
    
        if request.method == "POST":
            checkbox_estado = request.POST.get('otra_fecha', '')
            if checkbox_estado == "True":
                nuevo_estado = State_Applications(state = "Pendiente", initial_date=date.today(), observations = "Pendiente de Agendar Fecha Tour", id_applications = aplicacion)
                nuevo_estado.save()
                estado.final_date = date.today()
                estado.save()
                #Tours.objects.get(id_applications=id_applications, estado=True).count()
                if  Tours.objects.filter(id_applications=id_applications, estado=True).count()>0:
                    tipo=Tours.objects.get(id_applications=id_applications, estado=True)
                    tipo.estado=False
                    tipo.save()
                
                #mensaje.application_message(aplicacion.pk,9) #envio de mail a representantes cambio fecha de tour a otra fecha.
                mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 9, aplicacion)
                return redirect('Administracion')
                
                
                
            else:
                
                form = Toursnewform(request.POST)
                if form.is_valid():
                    
                    nuevo_estado = State_Applications(state = "Ingresado", initial_date=date.today(), observations = "Se reagendo Fecha Tour", id_applications = aplicacion)
                    nuevo_estado.save()
                    estado.final_date = date.today()
                    estado.save()
                    
                   
                    if Tours.objects.filter(id_applications=id_applications, estado=True).count()>0:
                        tipo=Tours.objects.get(id_applications=id_applications, estado=True)
                        tipo.estado=False
                        tipo.save()
                    form.save()
                    #mensaje.application_message(aplicacion.pk,6) #envio de mail a representantes cambio fecha de tour a otra fecha seleccionada.
                    mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 6, aplicacion)
                    return redirect('Administracion')
                    
                else:
                    print form.errors
                    return render(request,"admisiones_sam/DateTourUpdate.html", {"form":form})
                
                
        else:
            #debug = aplicacion.tours_set.get(id_applications = aplicacion)
            form = Toursnewform(initial={'id_applications':aplicacion,'estado':True})
            form.fields["id_typetour"].queryset = fechatour
            
            return render(request,"admisiones_sam/DateTourUpdate.html", {"form":form, "aplicacion": aplicacion})    
    else:
        mensaje='No se puede cambiar la fecha de Tour a un Tramite que esta en proceso de recopilación de documentos'
        return render(request,"admisiones_sam/ChangeStateMessage.html", {"mensaje":mensaje})  

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)    
def TypeTours(request):
    id_aplicante=request.session['id_aplicante'] 
    if request.method == "POST":
        form = TypeToursnewform(request.POST)
         
       
        if form.is_valid():
            item = form.save(commit=False)
            item.state = True
            item.name_tour = item.date_tour.strftime('%Y-%m-%d')
            if not Type_Tours.objects.filter(date_tour=item.date_tour).count()>0:
                item.save()
                form = TypeToursnewform()
                return redirect('EditarFechaTour',id_aplicante)
            #return HttpResponseRedirect(request.META.get('HTTP_REFERER','/'))
            else:
                return render(request,"admisiones_sam/TypeTourNewMessage2.html")
        else:
            
            print form.errors
            return render(request, "admisiones_sam/TypeTour.html", {"form":form,"id_aplicante":id_aplicante})
    else:
        form = TypeToursnewform()
        return render(request, "admisiones_sam/TypeTour.html", {"form":form,"id_aplicante":id_aplicante})

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)    
def KeepTypeTour(request):
    fechas=Type_Tours.objects.filter(state=True).order_by('date_tour')
    
    if request.method == "POST":
        form = TypeToursnewform(request.POST)
        if form.is_valid():
          
            item = form.save(commit=False)
            item.state = True
            item.name_tour = item.date_tour.strftime('%Y-%m-%d')
            
            if item.date_tour > date.today():
            
                if not Type_Tours.objects.filter(date_tour=item.date_tour).count()>0:
                    item.save()
                    form = TypeToursnewform()
                    return render(request, "admisiones_sam/KeepTypeTour.html", {"form":form,"fechas":fechas})
                else:
                    return render(request,"admisiones_sam/TypeTourNewMessage.html")
            else:
                    return render(request,"admisiones_sam/TypeTourNewMessage3.html")
          
        else:
            
            print form.errors
            return render(request, "admisiones_sam/KeepTypeTour.html", {"form":form,"fechas":fechas })
    else:
        
        form = TypeToursnewform()
        return render(request,"admisiones_sam/KeepTypeTour.html",{"form":form,"fechas":fechas})
    
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)    
def EditTypeTours(request, id_typetour):
  
    Tipos = Type_Tours.objects.get(pk = id_typetour)
    
    if request.method == "POST":
        form = TypeToursEditform(request.POST, instance = Tipos)    
        if form.is_valid():
            item = form.save(commit=False)
            item.name_tour = item.date_tour.strftime('%Y-%m-%d')
            
            #if not Type_Tours.objects.filter(date_tour=item.date_tour).count()>0: # si no existe ya creada
            if item.date_tour > date.today():
                item.save()
                return redirect('/admisiones/MantenerTipoTour')
            else:
                return render(request,"admisiones_sam/TypeTourNewMessage3.html") # no se puede actualizar
            #else:
                #return render(request,"admisiones_sam/TypeTourNewMessage.html")
        else:
            print form.errors
            return render(request,'admisiones_sam/EditTypeTour.html', {"form":form,"Tipos":Tipos})
    else:
        form = TypeToursEditform(instance = Tipos)
        return render(request,'admisiones_sam/EditTypeTour.html', {"form":form,"Tipos":Tipos})

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def DeleteTypeTours(request, id_typetour): 
    
    if not Tours.objects.filter(id_typetour=id_typetour,estado=True).count()>0:
    
        type_tour = Type_Tours.objects.get(pk = id_typetour)
        type_tour.delete()
        return redirect('/admisiones/MantenerTipoTour')
    else:
        return render(request,"admisiones_sam/TypeTourDeleteMessage.html")

def ToursMassiveEmails(request):
    mensaje = MailerMessage()
   
    consultafechastour=Type_Tours.objects.filter(date_tour__gte = date.today(),state=True).order_by('date_tour')
    fechas=[]
    
    for y in consultafechastour:
        if y.date_tour in fechas:
            pass
        else:
            fechas.append(y.date_tour.strftime("%Y-%m-%d"))
            
    if request.method == "GET" and "q_fecha" in request.GET:
        fecha=request.GET["q_fecha"]
        if fecha:
            aplicaciones=Applications.objects.filter(tours__id_typetour__date_tour=fecha,tours__id_typetour__state=True,tours__estado=True)
            if aplicaciones:
                for x in aplicaciones:
                    #mensaje.application_message(x.pk,19) #envio de mail a representantes para recordar la fecha de tour
                    mensaje.general_message(x.mail_tutor, "admisiones@montebelloacademy.org", 19, x)
                return render(request,'admisiones_sam/MassiveTourMessage.html',{"aplicaciones":aplicaciones,"fecha":fecha})  

            else:
                return redirect('/admisiones/ListaTour')
        else:
            return render(request,'admisiones_sam/ReportListTours.html',{"fechas":fechas})         
    else:
        return render2(request,'admisiones_sam/ReportListTours.html',{"fechas":fechas})     
        
        
        
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def KeepTypeTest(request):
    
    fechas=Type_Tests.objects.all().order_by('date_test')
    
    consultagrados=Type_Grade.objects.all().order_by('id_typegrade')
    lista=[]
    
    for y in consultagrados:
        if y.typegrade in lista:
            pass
        else:
            lista.append(y.typegrade)
            
               
    tipos=['General','PreEscolar','Individual']
    if request.method == "POST":
        
        form = TypeTestsnewform(request.POST)
        if form.is_valid():
            debug='3'
            form.save()
            return redirect('/admisiones/MantenerTipoTest')
        else:
            
            print form.errors
            return render(request, "admisiones_sam/KeepTypeTest.html", {"form":form,"fechas":fechas,"lista":lista,"tipos":tipos })
    else:
        debug='5'
        form = TypeTestsnewform()
        return render(request,"admisiones_sam/KeepTypeTest.html",{"form":form,"fechas":fechas,"lista":lista,"tipos":tipos})

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def DeleteTypeTest(request, id_typetest): 
    
    if not Test.objects.filter(id_typetest=id_typetest).count()>0:
    
        type_test = Type_Tests.objects.get(pk = id_typetest)
        type_test.delete()
        return redirect('/admisiones/MantenerTipoTest')
    else:
        return render(request,"admisiones_sam/TypeTestDeleteMessage.html")    

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def EditTypeTest(request, id_typetest):
  
    Tipos = Type_Tests.objects.get(pk = id_typetest)
    consultagrados=Type_Grade.objects.all().order_by('id_typegrade')
    lista=[]
    
    for y in consultagrados:
        if y.typegrade in lista:
            pass
        else:
            lista.append(y.typegrade)
   
    tipo=['General','PreEscolar','Individual']
    
    if request.method == "POST":
        form = TypeTestsnewform(request.POST, instance = Tipos) 
        if form.is_valid():
            form.save()
            return redirect('/admisiones/MantenerTipoTest')
        else:
            print form.errors
            return render(request,'admisiones_sam/EditTypeTest.html', {"form":form,"Tipos":Tipos,"lista":lista,"tipo":tipo})
    else:
        form = TypeTestsnewform(instance = Tipos)
        return render(request,'admisiones_sam/EditTypeTest.html', {"form":form,"Tipos":Tipos,"lista":lista,"tipo":tipo})        

def TestMassiveEmails(request):
    mensaje = MailerMessage()
    
    consultafechastest=Type_Tests.objects.filter(date_test__gte = date.today(),state=True).order_by('date_test')
    
    fechas=[]
    
    for y in consultafechastest:
        if y.date_test.strftime("%Y-%m-%d") in fechas:
            pass
        else:
            fechas.append(y.date_test.strftime("%Y-%m-%d"))
            
            
            
    if request.method == "GET" and "q_fecha" in request.GET:
        fecha=request.GET["q_fecha"]
        if fecha:
            aplicaciones=Applications.objects.filter(test__id_typetest__date_test=fecha,test__id_typetest__state=True,test__state=True)
            if aplicaciones:
                for x in aplicaciones:
                
                    if x.applied_grade==u'Inicial 1 (de 2 a 3 años)' or x.applied_grade==u'Inicial 2 (de 3 a 4 años)' or x.applied_grade==u'Inicial 2 (de 4 a 5 años) / Prekinder':
                        #mensaje.application_message(x.pk,31) #envio de mail a representantes de preescolar, para recordar evaluaciones
                        mensaje.general_message(x.mail_tutor, "admisiones@montebelloacademy.org", 31, x)
                    elif x.applied_grade==u'1ro de Básica / Kinder':
                        #mensaje.application_message(x.pk,32) #envio de mail a representantes de kinder para recordar la fecha de evaluaciones.
                        mensaje.general_message(x.mail_tutor, "admisiones@montebelloacademy.org", 32, x)
                    else:
                        #mensaje.application_message(x.pk,20) #envio de mail a representantes para recordar la fecha de evaluaciones
                        mensaje.general_message(x.mail_tutor, "admisiones@montebelloacademy.org", 20, x)
                
                return render(request,'admisiones_sam/MassiveTestMessage.html',{"aplicaciones":aplicaciones,"fecha":fecha})  

            else:
                return redirect('ListaEvaluacionesxlsx', fecha)   
        else:
            return render(request,'admisiones_sam/ReportListTest.html',{"fechas":fechas})         
    else:
        return render(request,'admisiones_sam/ReportListTest.html',{"fechas":fechas})      
         
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones) 

def SelectDateTest(request,id_applications): 
    aplicante=Applications.objects.get(id_applications=id_applications)
    mensaje = MailerMessage()
 
    fechas = Type_Tests.objects.filter(grade=aplicante.applied_grade,state=True,date_test__gt= date.today())
    #filter(grade=aplicante.applied_grade,state=True).order_by('date_test','grade')
    
    debug1='1'
    radio_fecha = None
    debug=radio_fecha  
    n=0
    c=0
    
    #validacion del estado de la solicitud
    
    estado=State_Applications.objects.get(id_applications=id_applications,final_date__isnull=True)
    documentos=Det_Documents.objects.filter(id_applications=id_applications)
    aplicacion=Applications.objects.get(id_applications=id_applications)
    pago_evaluacion=AccountingDepartmentTest.objects.get(id_applications_id=id_applications,state=1)
    #evaluacion=Test.objects.filter(id_applications_id=id_applications,state=True).count()
    if (estado.state =='Documentos' or estado.state =='Evaluaciones' or estado.state =='Finalizado'  or estado.state =='Error'  or estado.state =='Ingresado'):
        for x in documentos:
            n=n+1 #numero de documentos cargados
            if x.state == True:
                c=c+1  #numero de documentos validados
        if pago_evaluacion:
            if n==c:   # si los dos son iguales podra agendar fecha de evaluacion.
                if (estado.state =='Documentos' or estado.state =='Evaluaciones') and n!=0 :
                
                    if request.method == "POST":
                        radio_fecha =request.POST.get('seleccion')
                        if radio_fecha:
                            prueba=Type_Tests.objects.get(id_typetest=radio_fecha,state=True) 
                            debug1=3
                            if not Test.objects.filter(id_applications = id_applications).count()>0:
                                oportunidad=1
                                test= Test(opportunity=oportunidad,id_typetest = prueba,id_applications_id = id_applications,state=True)
                                test.save()
                              
                                nuevo_estado = State_Applications(state = "Evaluaciones", initial_date=date.today(), observations = "En Proceso de Evaluaciones", id_applications = aplicacion)
                                nuevo_estado.save()
                                estado.final_date = date.today()
                                estado.save()
                                
                                if aplicacion.applied_grade==u'Inicial 1 (de 2 a 3 años)' or aplicacion.applied_grade==u'Inicial 2 (de 3 a 4 años)' or aplicacion.applied_grade==u'Inicial 2 (de 4 a 5 años) / Prekinder':
                                    #mensaje.application_message(aplicacion.pk,12) #envio de mail a representantes informando la fecha de evaluaciones individual y grupal.
                                    #mensaje.application_message(aplicacion.pk,12) #envio de mail a representantes de preescolar, se envia las dos fechas
                                    mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 12, aplicacion)
                                
                                elif aplicacion.applied_grade==u'1ro de Básica / Kinder':
                                
                                    #mensaje.application_message(aplicacion.pk,16) #envio de mail a representantes de kinder informando la fecha de evaluaciones.
                                    mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 16, aplicacion)
                                else:
                                    #mensaje.application_message(aplicacion.pk,10) #envio de mail a representantes informando la fecha de evaluaciones.
                                    mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 10, aplicacion)
                                    
                                return redirect('Administracion')
                                
                            else:
                                if Test.objects.filter(id_applications = id_applications,state=True).count()>0:
                                    
                                    obj = Test.objects.get(id_applications=id_applications,state=True)
                                    
                                    
                                    oportunidad=obj.opportunity+1
                                    obj.state=False
                                    obj.save()
                                    
                                    test= Test(opportunity=oportunidad,id_typetest = prueba,id_applications_id = id_applications,state=True)
                                    test.save()
                                    
                                    nuevo_estado = State_Applications(state = "Evaluaciones", initial_date=date.today(), observations = "Se modifica la Fecha de Evaluaciones", id_applications = aplicacion)
                                    nuevo_estado.save()
                                    estado.final_date = date.today()
                                    estado.save()
                                    
                                    if aplicacion.applied_grade==u'Inicial 1 (de 2 a 3 años)' or aplicacion.applied_grade==u'Inicial 2 (de 3 a 4 años)' or aplicacion.applied_grade==u'Inicial 2 (de 4 a 5 años) / Prekinder':
                                        #mensaje.application_message(aplicacion.pk,27) #envio de mail a representantes informando la fecha de evaluaciones individual y grupal.
                                        mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 27, aplicacion)
                                    elif aplicacion.applied_grade==u'1ro de Básica / Kinder':
                                        #mensaje.application_message(aplicacion.pk,26) #envio de mail a representantes informando la nueva fecha de evaluaciones 
                                        mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 26, aplicacion)
                                    else:
                                        #mensaje.application_message(aplicacion.pk,11) #envio de mail a representantes informando la Nueva fecha de evaluaciones.
                                        mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 11, aplicacion)
                                    return redirect('Administracion')
                                else:
                                    ultimo = Test.objects.filter(id_applications=id_applications).aggregate(Max('id_test'))
                                    
                                    obj=Test.objects.get(id_test=ultimo['id_test__max'])
                                    
                                    oportunidad=obj.opportunity+1
                                    test= Test(opportunity=oportunidad,id_typetest = prueba,id_applications_id = id_applications,state=True)
                                    test.save()
                                    
                                    nuevo_estado = State_Applications(state = "Evaluaciones", initial_date=date.today(), observations = "Se modifica la Fecha de Evaluaciones", id_applications = aplicacion)
                                    nuevo_estado.save()
                                    estado.final_date = date.today()
                                    estado.save()
                                    
                                    if aplicacion.applied_grade==u'Inicial 1 (de 2 a 3 años)' or aplicacion.applied_grade==u'Inicial 2 (de 3 a 4 años)' or aplicacion.applied_grade==u'Inicial 2 (de 4 a 5 años) / Prekinder':
                                         #mensaje.application_message(aplicacion.pk,12) #envio de mail a representantes informando la fecha de evaluaciones individual y grupal.
                                         mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 12, aplicacion)
                                    else:
                                        #mensaje.application_message(aplicacion.pk,10) #envio de mail a representantes informando la fecha de evaluaciones.
                                        mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 10, aplicacion)
                                    return redirect('Administracion')
                        else:   
                            debug1='6'
                            debug=radio_fecha
                            return render(request,"admisiones_sam/SelectTypeTest.html", {"fechas":fechas,"id_applications":id_applications,"debug":n,"debug1":documentos})
                    
                    else:
                        debug1='2'
                        debug=radio_fecha
                        return render(request,"admisiones_sam/SelectTypeTest.html", {"fechas":fechas,"id_applications":id_applications,"debug":n,"debug1":documentos})
                else:
                    mensaje='No se puede asignar nueva fecha de evaluación, Revise el estado de la Solicitud o la Validación de la documentacion Requerida'
                    return render(request,"admisiones_sam/CreateTestMessage.html",{"mensaje":mensaje})
            else:
                mensaje='No se ha validado toda la documentación del Aplicante'
                return render(request,"admisiones_sam/CreateTestMessage.html",{"mensaje":mensaje})           
        else:
            mensaje='No se ha registrado el pago de evaluaciones para el Aplicante'
            return render(request,"admisiones_sam/CreateTestMessage.html",{"mensaje":mensaje}) 
        
    else:
        mensaje='No se puede asignar nueva fecha de evaluación, Revise el estado de la Solicitud'
        return render(request,"admisiones_sam/CreateTestMessage.html",{"mensaje":mensaje}) 

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def TypeTest(request,id_applications):
    id_aplicante=id_applications
    fecha_actual=date.today()
    #id_aplicante=request.session['id_aplicante'] 
    consultagrados=Type_Grade.objects.all().order_by('id_typegrade')
    lista=[]
    
    for y in consultagrados:
        if y.typegrade in lista:
            pass
        else:
            lista.append(y.typegrade)
    
    tipo=['General','PreEscolar','Individual']
    
    if request.method == "POST":
        form = TypeTestsnewform(request.POST)
       
        if form.is_valid():
            item = form.save(commit=False)
            item.state = True
            if item.date_test > fecha_actual:
                item.save()
                return redirect('SeleccionaFecha',id_aplicante)
            else:
                return render(request, "admisiones_sam/CreateTestMessage1.html",{"id_aplicante":id_aplicante})
        else:
            print form.errors
            return render(request, "admisiones_sam/TypeTest.html", {"form":form,"id_aplicante":id_aplicante,"lista":lista,"tipo":tipo})
    else:
        form = TypeTestsnewform()
        return render(request, "admisiones_sam/TypeTest.html", {"form":form,"id_aplicante":id_aplicante,"lista":lista,"tipo":tipo})

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def KeepPeriod(request):
    periodos=PeriodSchool.objects.all().order_by('per_startdate')
    
    if request.method == "POST":
        form = PeriodNewform(request.POST)
        if form.is_valid():
          
            item = form.save(commit=False)
            item.per_state = True
            item.per_name = item.per_startdate.strftime('%Y')+'-'+item.per_enddate.strftime('%Y')
            if not PeriodSchool.objects.filter(per_name=item.per_name).count()>0:
                item.save()
                form = PeriodNewform()
                return render(request, "admisiones_sam/KeepPeriod.html", {"form":form,"periodos":periodos})
            else:
                return render(request,"admisiones_sam/PeriodNewMessage.html")
            
        else:
            
            print form.errors
            return render(request, "admisiones_sam/KeepPeriod.html", {"form":form,"periodos":periodos})
    else:
        
        form = PeriodNewform()
        return render(request, "admisiones_sam/KeepPeriod.html", {"form":form,"periodos":periodos})
    
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def DeletePeriod(request, per_id): 
    
    if not Applications.objects.filter(school_period=per_id).count()>0:
    
        periodo = PeriodSchool.objects.get(pk = per_id)
        periodo.delete()
        return redirect('/admisiones/MantenerPeriodos')
    else:
        return render(request,"admisiones_sam/PeriodDeleteMessage.html")

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def EditPeriod(request, per_id):
  
    periodos = PeriodSchool.objects.get(pk = per_id)
    
    if request.method == "POST":
        form = PeriodEditform(request.POST, instance = periodos)    
        if form.is_valid():
            item = form.save(commit=False)
            item.per_name = item.per_startdate.strftime('%Y')+'-'+item.per_enddate.strftime('%Y')
            item.save()
            return redirect('/admisiones/MantenerPeriodos')
        else:
            print form.errors
            return render(request,'admisiones_sam/EditPeriod.html', {"form":form,"periodos":periodos})
    else:
        form = PeriodEditform(instance = periodos)
        return render(request,'admisiones_sam/EditPeriod.html', {"form":form,"periodos":periodos})        


def NewFormsStudents(request,id_applications):
    
    #valida si ya se ha ingresado informacion en el formulario para ese estudiante
    valida=State_Applications.objects.get(id_applications_id=id_applications,final_date__isnull=True)
    secretaria_state=AcademicSecretary.objects.get(id_applications_id=id_applications)
    if valida.state=='Secretaria' and secretaria_state.state=='0':
        
        listap={'---------':'',
         u'Afghanistan': 'AF',
         u'Albania': 'AL',
         u'Algeria': 'DZ',
         u'Aland Islands': 'AX',
         u'American Samoa': 'AS',
         u'Andorra': 'AD',
         u'Angola': 'AO',
         u'Anguilla': 'AI',
         u'Antarctica': 'AQ',
         u'Antigua and Barbuda': 'AG',
         u'Argentina': 'AR',
         u'Armenia': 'AM',
         u'Aruba': 'AW',
         u'Australia': 'AU',
         u'Austria': 'AT',
         u'Azerbaijan': 'AZ',
         u'Bahamas': 'BS',
         u'Bahrain': 'BH',
         u'Bangladesh': 'BD',
         u'Barbados': 'BB',
         u'Belarus': 'BY',
         u'Belgium': 'BE',
         u'Belize': 'BZ',
         u'Benin': 'BJ',
         u'Bermuda': 'BM',
         u'Bhutan': 'BT',
         u'Bolivia, Plurinational State of': 'BO',
         u'Bonaire, Sint Eustatius and Saba': 'BQ',
         u'Bosnia and Herzegovina': 'BA',
         u'Botswana': 'BW',
         u'Bouvet Island': 'BV',
         u'Brazil': 'BR',
         u'British Indian Ocean Territory': 'IO',
         u'Brunei Darussalam': 'BN',
         u'Bulgaria': 'BG',
         u'Burkina Faso': 'BF',
         u'Burundi': 'BI',
         u'Cambodia': 'KH',
         u'Cameroon': 'CM',
         u'Canada': 'CA',
         u'Cape Verde': 'CV',
         u'Cayman Islands': 'KY',
         u'Central African Republic': 'CF',
         u'Chad': 'TD',
         u'Chile': 'CL',
         u'China': 'CN',
         u'Christmas Island': 'CX',
         u'Cocos (Keeling) Islands': 'CC',
         u'Colombia': 'CO',
         u'Comoros': 'KM',
         u'Congo': 'CG',
         u'Congo, the Democratic Republic of the': 'CD',
         u'Cook Islands': 'CK',
         u'Costa Rica': 'CR',
         u'Country name': 'Code',
         u'Croatia': 'HR',
         u'Cuba': 'CU',
         u'Curaçao': 'CW',
         u'Cyprus': 'CY',
         u'Czech Republic': 'CZ',
         u"Côte d'Ivoire": 'CI',
         u'Denmark': 'DK',
         u'Djibouti': 'DJ',
         u'Dominica': 'DM',
         u'Dominican Republic': 'DO',
         u'Ecuador': 'EC',
         u'Egypt': 'EG',
         u'El Salvador': 'SV',
         u'Equatorial Guinea': 'GQ',
         u'Eritrea': 'ER',
         u'Estonia': 'EE',
         u'Ethiopia': 'ET',
         u'Falkland Islands (Malvinas)': 'FK',
         u'Faroe Islands': 'FO',
         u'Fiji': 'FJ',
         u'Finland': 'FI',
         u'France': 'FR',
         u'French Guiana': 'GF',
         u'French Polynesia': 'PF',
         u'French Southern Territories': 'TF',
         u'Gabon': 'GA',
         u'Gambia': 'GM',
         u'Georgia': 'GE',
         u'Germany': 'DE',
         u'Ghana': 'GH',
         u'Gibraltar': 'GI',
         u'Greece': 'GR',
         u'Greenland': 'GL',
         u'Grenada': 'GD',
         u'Guadeloupe': 'GP',
         u'Guam': 'GU',
         u'Guatemala': 'GT',
         u'Guernsey': 'GG',
         u'Guinea': 'GN',
         u'Guinea-Bissau': 'GW',
         u'Guyana': 'GY',
         u'Haiti': 'HT',
         u'Heard Island and McDonald Islands': 'HM',
         u'Holy See (Vatican City State)': 'VA',
         u'Honduras': 'HN',
         u'Hong Kong': 'HK',
         u'Hungary': 'HU',
         u'ISO 3166-2:GB': '(.uk)',
         u'Iceland': 'IS',
         u'India': 'IN',
         u'Indonesia': 'ID',
         u'Iran, Islamic Republic of': 'IR',
         u'Iraq': 'IQ',
         u'Ireland': 'IE',
         u'Isle of Man': 'IM',
         u'Israel': 'IL',
         u'Italy': 'IT',
         u'Jamaica': 'JM',
         u'Japan': 'JP',
         u'Jersey': 'JE',
         u'Jordan': 'JO',
         u'Kazakhstan': 'KZ',
         u'Kenya': 'KE',
         u'Kiribati': 'KI',
         u"Korea, Democratic People's Republic of": 'KP',
         u'Korea, Republic of': 'KR',
         u'Kuwait': 'KW',
         u'Kyrgyzstan': 'KG',
         u"Lao People's Democratic Republic": 'LA',
         u'Latvia': 'LV',
         u'Lebanon': 'LB',
         u'Lesotho': 'LS',
         u'Liberia': 'LR',
         u'Libya': 'LY',
         u'Liechtenstein': 'LI',
         u'Lithuania': 'LT',
         u'Luxembourg': 'LU',
         u'Macao': 'MO',
         u'Macedonia, the former Yugoslav Republic of': 'MK',
         u'Madagascar': 'MG',
         u'Malawi': 'MW',
         u'Malaysia': 'MY',
         u'Maldives': 'MV',
         u'Mali': 'ML',
         u'Malta': 'MT',
         u'Marshall Islands': 'MH',
         u'Martinique': 'MQ',
         u'Mauritania': 'MR',
         u'Mauritius': 'MU',
         u'Mayotte': 'YT',
         u'Mexico': 'MX',
         u'Micronesia, Federated States of': 'FM',
         u'Moldova, Republic of': 'MD',
         u'Monaco': 'MC',
         u'Mongolia': 'MN',
         u'Montenegro': 'ME',
         u'Montserrat': 'MS',
         u'Morocco': 'MA',
         u'Mozambique': 'MZ',
         u'Myanmar': 'MM',
         u'Namibia': 'NA',
         u'Nauru': 'NR',
         u'Nepal': 'NP',
         u'Netherlands': 'NL',
         u'New Caledonia': 'NC',
         u'New Zealand': 'NZ',
         u'Nicaragua': 'NI',
         u'Niger': 'NE',
         u'Nigeria': 'NG',
         u'Niue': 'NU',
         u'Norfolk Island': 'NF',
         u'Northern Mariana Islands': 'MP',
         u'Norway': 'NO',
         u'Oman': 'OM',
         u'Pakistan': 'PK',
         u'Palau': 'PW',
         u'Palestine, State of': 'PS',
         u'Panama': 'PA',
         u'Papua New Guinea': 'PG',
         u'Paraguay': 'PY',
         u'Peru': 'PE',
         u'Philippines': 'PH',
         u'Pitcairn': 'PN',
         u'Poland': 'PL',
         u'Portugal': 'PT',
         u'Puerto Rico': 'PR',
         u'Qatar': 'QA',
         u'Romania': 'RO',
         u'Russian Federation': 'RU',
         u'Rwanda': 'RW',
         u'Réunion': 'RE',
         u'Saint Barthélemy': 'BL',
         u'Saint Helena, Ascension and Tristan da Cunha': 'SH',
         u'Saint Kitts and Nevis': 'KN',
         u'Saint Lucia': 'LC',
         u'Saint Martin (French part)': 'MF',
         u'Saint Pierre and Miquelon': 'PM',
         u'Saint Vincent and the Grenadines': 'VC',
         u'Samoa': 'WS',
         u'San Marino': 'SM',
         u'Sao Tome and Principe': 'ST',
         u'Saudi Arabia': 'SA',
         u'Senegal': 'SN',
         u'Serbia': 'RS',
         u'Seychelles': 'SC',
         u'Sierra Leone': 'SL',
         u'Singapore': 'SG',
         u'Sint Maarten (Dutch part)': 'SX',
         u'Slovakia': 'SK',
         u'Slovenia': 'SI',
         u'Solomon Islands': 'SB',
         u'Somalia': 'SO',
         u'South Africa': 'ZA',
         u'South Georgia and the South Sandwich Islands': 'GS',
         u'South Sudan': 'SS',
         u'Spain': 'ES',
         u'Sri Lanka': 'LK',
         u'Sudan': 'SD',
         u'Suriname': 'SR',
         u'Svalbard and Jan Mayen': 'SJ',
         u'Swaziland': 'SZ',
         u'Sweden': 'SE',
         u'Switzerland': 'CH',
         u'Syrian Arab Republic': 'SY',
         u'Taiwan, Province of China': 'TW',
         u'Tajikistan': 'TJ',
         u'Tanzania, United Republic of': 'TZ',
         u'Thailand': 'TH',
         u'Timor-Leste': 'TL',
         u'Togo': 'TG',
         u'Tokelau': 'TK',
         u'Tonga': 'TO',
         u'Trinidad and Tobago': 'TT',
         u'Tunisia': 'TN',
         u'Turkey': 'TR',
         u'Turkmenistan': 'TM',
         u'Turks and Caicos Islands': 'TC',
         u'Tuvalu': 'TV',
         u'Uganda': 'UG',
         u'Ukraine': 'UA',
         u'United Arab Emirates': 'AE',
         u'United Kingdom': 'GB',
         u'United States': 'US',
         u'United States Minor Outlying Islands': 'UM',
         u'Uruguay': 'UY',
         u'Uzbekistan': 'UZ',
         u'Vanuatu': 'VU',
         u'Venezuela, Bolivarian Republic of': 'VE',
         u'Viet Nam': 'VN',
         u'Virgin Islands, British': 'VG',
         u'Virgin Islands, U.S.': 'VI',
         u'Wallis and Futuna': 'WF',
         u'Western Sahara': 'EH',
         u'Yemen': 'YE',
         u'Zambia': 'ZM',
         u'Zimbabwe': 'ZW'
        }

        listap_o = OrderedDict(sorted(listap.items(), key=operator.itemgetter(0)))
        
        consultagrados=Type_Grade.objects.all().order_by('id_typegrade')
        listag=[]
        
        for y in consultagrados:
            if y.typegrade in listag:
                pass
            else:
                listag.append(y.typegrade)
        
                
        aplicacion=Applications.objects.get(pk=id_applications)
        request.session['id_aplicacion']=aplicacion.pk  #cookies
        debug='0'
        if request.method == "POST":

            # recolectar información del formulario
            
            s_identificacion=request.POST.get('s_identificacion', "")
            s_nombres=request.POST.get('s_nombres', "")
            s_apellidos=request.POST.get('s_apellidos', "")
            s_apellidos_2=request.POST.get('s_apellidos_2', "")
            s_genero=request.POST.get('s_genero')
            s_fecha_nacimiento=request.POST.get('s_fecha_nacimiento', "")
            s_direccion=request.POST.get('s_direccion', "")
            s_pais_nacimiento=request.POST.get('s_pais_nacimiento', "")
            s_pais_residencia=request.POST.get('s_pais_residencia', "")
            s_grado_aplica=aplicacion.applied_grade
            s_informacion_medica=u"{}".format(request.POST.get('s_informacion_medica', ""))

            #crear usuario office365

            o365_estudiante = create_student(s_nombres, s_apellidos, s_apellidos_2)

            #crear usuario sistema con mail de office 365

            usuario_estudiante = CustomUser.objects.create_user(s_nombres, s_apellidos, s_apellidos_2, o365_estudiante["userPrincipalName"])
            
            usuario_estudiante.identity = s_identificacion
            usuario_estudiante.gender = s_genero
            usuario_estudiante.birthdate = s_fecha_nacimiento
            usuario_estudiante.address = s_direccion
            usuario_estudiante.birth_country = s_pais_nacimiento
            usuario_estudiante.country_home = s_pais_residencia
            usuario_estudiante.save()

          
            #validar que exista informacion del papa y mama para ingresar 
            checkbox_papa = request.POST.get('valida_papa', '')

            if checkbox_papa == '':

                p_identificacion=request.POST.get('p_identificacion')
                p_nombres=request.POST.get('p_nombres')
                p_apellidos=request.POST.get('p_apellidos')
                p_apellidos_2=request.POST.get('p_apellidos_2')
                p_genero="M"
                p_fecha_nacimiento=request.POST.get('p_fecha_nacimiento')
                p_pais_nacimiento=request.POST.get('p_pais_nacimiento')
                p_pais_residencia=request.POST.get('p_pais_residencia')
                p_telefono=request.POST.get('p_telefono')
                p_cell=request.POST.get('p_cell')
                p_mail=request.POST.get('p_mail')
                p_direccion=request.POST.get('p_direccion')
                
                p_vive_con=request.POST.get('p_vive_con')
                p_retirar=request.POST.get('p_retirar')
                p_difunto = request.POST.get("p_difunto", "")
                p_instruction=request.POST.get('p_instruccion')
                p_profession=request.POST.get('p_profesion')
                p_address_work=request.POST.get('p_trabajo')
                p_telofi=request.POST.get('p_telofi')

                o365_papa = create_parent(p_nombres, p_apellidos, p_apellidos_2)
                

                usuario_papa = CustomUser.objects.create_user(p_nombres, p_apellidos, p_apellidos_2, o365_papa["userPrincipalName"])
                usuario_papa.identity = p_identificacion
                usuario_papa.gender = p_genero
                usuario_papa.birthdate = p_fecha_nacimiento
                usuario_papa.address = p_direccion
                usuario_papa.birth_country = p_pais_nacimiento
                usuario_papa.country_home = p_pais_residencia
                usuario_papa.phone = p_telefono
                usuario_papa.cell = p_cell
         

                if p_difunto == "true":
                    usuario_papa.is_active = False
                    p_difunto=True
                else:
                    p_difunto=False
                    
                if p_vive_con=="true" :
                    p_vive_con=True
                else:
                    p_vive_con=False
                
                if p_retirar=="true" :
                    p_retirar=True
                else:
                    p_retirar=False
                
                usuario_papa.save()
                
                persona_papa= People(id_people = usuario_papa,
                                    is_dead = p_difunto,
                                    instruction=p_instruction,
                                    profession=p_profession,
                                    address_work=p_address_work,
                                    office = p_telofi,
                                    e_mail=p_mail)
                persona_papa.save()
                
            else:
                usuario_papa = None
                persona_papa = None


            checkbox_mama = request.POST.get('valida_mama', '')

            if checkbox_mama == '':

                m_identificacion=request.POST.get('m_identificacion', "")
                m_nombres=request.POST.get('m_nombres', "")
                m_apellidos=request.POST.get('m_apellidos', "")
                m_apellidos_2=request.POST.get('m_apellidos_2', "")
                m_genero="F"
                m_fecha_nacimiento=request.POST.get('m_fecha_nacimiento', "")
                m_direccion=request.POST.get('m_direccion', "")
                m_pais_nacimiento=request.POST.get('m_pais_nacimiento', "")
                m_pais_residencia=request.POST.get('m_pais_residencia', "")
                m_telefono=request.POST.get('m_telefono', "")
                m_cell=request.POST.get('m_cell', "")
                m_mail=request.POST.get('m_mail', "")
                
                m_difunto = request.POST.get("m_difunto", "")
                m_vive_con=request.POST.get('m_vive_con')
                m_retirar=request.POST.get('m_retirar')
                m_instruction=request.POST.get('m_instruccion')
                m_profession=request.POST.get('m_profesion')
                m_address_work=request.POST.get('m_trabajo')
                m_telofi=request.POST.get('m_telofi', "")

                o365_mama = create_parent(m_nombres, m_apellidos, m_apellidos_2)
                usuario_mama = CustomUser.objects.create_user(m_nombres, m_apellidos, m_apellidos_2, o365_mama["userPrincipalName"])
                usuario_mama.identity = m_identificacion
                usuario_mama.gender = m_genero
                usuario_mama.birthdate = m_fecha_nacimiento
                usuario_mama.address = m_direccion
                usuario_mama.birth_country = m_pais_nacimiento
                usuario_mama.country_home = m_pais_residencia
                usuario_mama.phone = m_telefono
                usuario_mama.cell = m_cell
             

                if m_difunto == "true":
                    usuario_mama.is_active = False
                    m_difunto=True
                else:
                    m_difunto=False
                 
                if m_vive_con=="true" :
                    m_vive_con=True
                else:
                    m_vive_con=False
                
                if m_retirar=="true" :
                    m_retirar=True
                else:
                    m_retirar=False             
                    
                usuario_mama.save()
                persona_mama= People(id_people = usuario_mama,
                                        is_dead = m_difunto,
                                        instruction=m_instruction,
                                        profession=m_profession,
                                        address_work=m_address_work,
                                        office = m_telofi,
                                        e_mail=m_mail)
                persona_mama.save()
                
              
            else:
                usuario_mama = None
                persona_mama=None


            representante = request.POST.get('representante')

            if representante=='P':
                usuario_representante=usuario_papa
                instruction_p=persona_papa.instruction
                profession_p=persona_papa.profession
                address_work_p=persona_papa.address_work
                office_p = persona_papa.office
                e_mail_p=persona_papa.e_mail
                vive_con_p=p_vive_con
                retira_p=p_retirar
                
                
            elif representante=='M':
                usuario_representante=usuario_mama
                instruction_p=persona_mama.instruction
                profession_p=persona_mama.profession
                address_work_p=persona_mama.address_work
                office_p = persona_mama.office
                e_mail_p=persona_mama.e_mail
                vive_con_p=m_vive_con
                retira_p=m_retirar
                
            elif representante=='O':
                
                r_identificacion=request.POST.get('r_identificacion', "")
                r_nombres=request.POST.get('r_nombres', "")
                r_apellidos=request.POST.get('r_apellidos')
                r_apellidos_2=request.POST.get('r_apellidos_2', "")
                r_genero=request.POST.get('r_genero')
                r_fecha_nacimiento=request.POST.get('r_fecha_nacimiento', "")
                r_direccion=request.POST.get('r_direccion', "")
                r_pais_nacimiento=request.POST.get('r_pais_nacimiento', "")
                r_pais_residencia=request.POST.get('r_pais_residencia', "")
                r_grado_aplica=request.POST.get('r_grado_aplica', "")
                r_informacion_medica=request.POST.get('r_informacion_medica', "")
                r_telefono=request.POST.get('r_telefono', "")
                r_cell=request.POST.get('r_cell', "")
                r_mail=request.POST.get('r_mail', "")
                
                r_vive_con=request.POST.get('r_vive_con')
                r_retirar=request.POST.get('r_retirar')
                r_instruction=request.POST.get('r_instruccion')
                r_profession=request.POST.get('r_profesion')
                r_address_work=request.POST.get('r_trabajo')
                r_telofi=request.POST.get('r_telofi', "")
               
                o365_tutor = create_tutor(r_nombres, r_apellidos, r_apellidos_2)
                usuario_representante = CustomUser.objects.create_user(r_nombres, r_apellidos, r_apellidos_2, o365_tutor["userPrincipalName"])
                usuario_representante.identity = r_identificacion
                usuario_representante.gender = r_genero
                usuario_representante.birthdate = r_fecha_nacimiento
                usuario_representante.address = r_direccion
                usuario_representante.birth_country = r_pais_nacimiento
                usuario_representante.country_home = r_pais_residencia
                usuario_representante.phone = r_telefono
                usuario_representante.cell = r_cell
               

                usuario_representante.save()
                
                if r_vive_con=="true" :
                    r_vive_con=True
                else:
                    r_vive_con=False
                
                if r_retirar=="true" :
                    r_retirar=True
                else:
                    r_retirar=False
                    
            
                instruction_p=r.instruction
                profession_p=r.profession
                address_work_p=r.address_work
                office_p = r.telofi
                e_mail_p=r.mail
                vive_con_p=r_vive_con
                retira_p=r_retirar
                
         #ojo si existe el usuario creado ya no lo crea.
         
            #if People.objects.filter(id_people=usuario_representante).count()==0:   
            if People.objects.get(id_people=usuario_representante):         
                #persona_representante=People.objects.get(id_people=usuario_representante)
                persona_representante=None
            else:            
                persona_representante= People(id_people = usuario_representante,
                                        instruction=instruction_p,
                                        profession=profession_p,
                                        address_work=address_work_p,
                                        office = office_p,
                                        e_mail=e_mail_p)
                persona_representante.save()
               
            estudiante= Students(id_students =usuario_estudiante,
                                id_tutor=usuario_representante,
                                id_father=usuario_papa,
                                id_mother=usuario_mama,
                                applied_grade = s_grado_aplica,
                                medical_information =s_informacion_medica,
                                id_applications=aplicacion)
            estudiante.save()
            
            if  persona_papa != None:
                relacion_papa=RelationshipStudent(id_people =persona_papa,
                                                      id_students =estudiante,
                                                      live_student = p_vive_con ,
                                                      withdraw = p_retirar)
                relacion_papa.save()
            
            if persona_mama !=None:
                relacion_mama=RelationshipStudent(id_people =persona_mama,
                                                      id_students =estudiante,
                                                      live_student = m_vive_con ,
                                                      withdraw = m_retirar)
                relacion_mama.save()
            
            if persona_representante !=None:
                relacion_representante=RelationshipStudent(id_people =persona_representante,
                                                      id_students =estudiante,
                                                      live_student = vive_con_p ,
                                                      withdraw = retira_p)
                relacion_representante.save()
                
            #Actualizo el registro en secretaria
                    
            secretaria_old=AcademicSecretary.objects.get(state='0',id_applications=aplicacion) 
            secretaria_old.state='1'
            secretaria_old.save()            
            #secretaria = AcademicSecretary(state='False',id_applications=aplicacion)
            #secretaria.save()
            #old_estado=State_Applications.objects.get(id_applications = aplicacion.id_applications, final_date__isnull=True)
            #old_estado.final_date=date.today()
            #old_estado.save()
            #new_estado = State_Applications(state='Secretaria',final_date=None,observations='Registro en Secretaria',id_applications=aplicacion)
            #new_estado.save()
                        
            debug='3'

            """Cambio introducido por reestructuración de Estudiantes y Representantes en el sistema. 
            No se elimina la creación de usuarios antiguos (Students, People, Relationship) por temas de compatibilidad.
            Una vez determinado la correcta actualización de todos los módulos en relación a estos objetos antiguos, revisar las
            definiciones cruzadas entre el código de esta nueva sección y el código que antecede, que debería eliminarse."""

            estudiante = Student.objects.create(user = usuario_estudiante, through = "Admisiones")
            
            if usuario_papa != None:
                if p_difunto == True:
                    pass
                else:
                    usuario_papa, creado = CustomUser.objects.get_or_create(email = p_mail, defaults = {
                        "identity": p_identificacion,
                        "first_name": p_nombres,
                        "father_last_name": p_apellidos,
                        "mother_last_name": p_apellidos_2,
                        "gender":"M", "birthdate":p_fecha_nacimiento,
                        "phone":p_telefono,"cell":p_cell,
                        "address":p_direccion, 
                        "birth_country":p_pais_nacimiento, 
                        "country_home":p_pais_residencia
                        })


                papa, creado = Relative.objects.get_or_create(user = usuario_papa, defaults = {
                    "through": "Admisiones"
                    })

                detalle_papa, creado = Relative_Details.objects.get_or_create(relative = papa, defaults = {
                    "alive": not(p_difunto),
                    "instruction":p_instruction,
                    "profession": p_profession,
                    "address_work":p_address_work,
                    "office":p_telofi
                    })

                relacion_papa = Student_Relative.objects.create(
                    student = estudiante,
                    relative = papa,
                    relationship = "PADRE",
                    live_together = p_vive_con,
                    withdraw = p_retirar)
         

            if usuario_mama != None:
                if m_difunto == True:
                    pass
                else:
                    usuario_mama, creado = CustomUser.objects.get_or_create(email = m_mail, defaults = {
                        "identity": m_identificacion,
                        "first_name": m_nombres,
                        "father_last_name": m_apellidos,
                        "mother_last_name": m_apellidos_2,
                        "gender":"M", "birthdate":m_fecha_nacimiento,
                        "phone":m_telefono,"cell":m_cell,
                        "address":m_direccion, 
                        "birth_country":m_pais_nacimiento, 
                        "country_home":m_pais_residencia
                        })


                mama, creado = Relative.objects.get_or_create(user = usuario_mama, defaults = {
                    "through": "Admisiones"
                    })

                detalle_mama, creado = Relative_Details.objects.get_or_create(relative = mama, defaults = {
                    "alive" : not(m_difunto),
                    "instruction":m_instruction,
                    "profession": m_profession,
                    "address_work":m_address_work,
                    "office":m_telofi
                    })

                relacion_mama = Student_Relative.objects.create(
                    student = estudiante,
                    relative = mama,
                    relationship = "MADRE",
                    live_together = m_vive_con,
                    withdraw = m_retirar)

            if representante == "P":
                relacion_papa.legal_representantive = True
                relacion_papa.notifications = True
                relacion_papa.save()

            elif representante == "M":
                relacion_mama.legal_representantive = True
                relacion_mama.notifications = True
                relacion_mama.save()

            elif representante == "O":

                    usuario_representante, creado = CustomUser.objects.get_or_create(email = r_mail, defaults = {
                        "identity": r_identificacion,
                        "first_name": r_nombres,
                        "father_last_name": r_apellidos,
                        "mother_last_name": r_apellidos_2,
                        "gender":"M", "birthdate":r_fecha_nacimiento,
                        "phone":r_telefono,"cell":r_cell,
                        "address":r_direccion, 
                        "birth_country":r_pais_nacimiento, 
                        "country_home":r_pais_residencia
                        })


                    representante, creado = Relative.objects.get_or_create(user = usuario_representante, defaults = {
                        "through": "Admisiones"
                        })

                    detalle_representante, creado = Relative_Details.objects.get_or_create(relative = representante, defaults = {
                        "alive" :True,
                        "instruction":r_instruction,
                        "profession": r_profession,
                        "address_work":r_address_work,
                        "office":r_telofi
                        })

                    relacion_representante = Student_Relative.objects.create(
                        student = estudiante,
                        relative = representante,
                        relationship = "OTRA",
                        live_together = r_vive_con,
                        withdraw = r_retirar,
                        legal_representantive = True,
                        notifications = True)


            return render(request, "admisiones_sam/FormsStudentsMessageOk.html", {"Aplicacion":aplicacion,"debug":debug,"listap":listap_o,"listag":listag})
        else:
            debug='99'
            
            return render(request, "admisiones_sam/NewFormsStudents.html", {"aplicacion":aplicacion,"debug":debug,"listap":listap_o,"listag":listag})
         
    else:    
        return render(request,'admisiones_sam/FormsExistMessage.html', {"valida":valida})   


@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def EditFormsStudents(request):

    consultagrados=Type_Grade.objects.all().order_by('id_typegrade')
    listag=[]
    
    for y in consultagrados:
        if y.typegrade in listag:
            pass
        else:
            listag.append(y.typegrade)
       
    listap={'---------':'NN',
     u'Afghanistan': 'AF',
     u'Albania': 'AL',
     u'Algeria': 'DZ',
     u'Aland Islands': 'AX',
     u'American Samoa': 'AS',
     u'Andorra': 'AD',
     u'Angola': 'AO',
     u'Anguilla': 'AI',
     u'Antarctica': 'AQ',
     u'Antigua and Barbuda': 'AG',
     u'Argentina': 'AR',
     u'Armenia': 'AM',
     u'Aruba': 'AW',
     u'Australia': 'AU',
     u'Austria': 'AT',
     u'Azerbaijan': 'AZ',
     u'Bahamas': 'BS',
     u'Bahrain': 'BH',
     u'Bangladesh': 'BD',
     u'Barbados': 'BB',
     u'Belarus': 'BY',
     u'Belgium': 'BE',
     u'Belize': 'BZ',
     u'Benin': 'BJ',
     u'Bermuda': 'BM',
     u'Bhutan': 'BT',
     u'Bolivia, Plurinational State of': 'BO',
     u'Bonaire, Sint Eustatius and Saba': 'BQ',
     u'Bosnia and Herzegovina': 'BA',
     u'Botswana': 'BW',
     u'Bouvet Island': 'BV',
     u'Brazil': 'BR',
     u'British Indian Ocean Territory': 'IO',
     u'Brunei Darussalam': 'BN',
     u'Bulgaria': 'BG',
     u'Burkina Faso': 'BF',
     u'Burundi': 'BI',
     u'Cambodia': 'KH',
     u'Cameroon': 'CM',
     u'Canada': 'CA',
     u'Cape Verde': 'CV',
     u'Cayman Islands': 'KY',
     u'Central African Republic': 'CF',
     u'Chad': 'TD',
     u'Chile': 'CL',
     u'China': 'CN',
     u'Christmas Island': 'CX',
     u'Cocos (Keeling) Islands': 'CC',
     u'Colombia': 'CO',
     u'Comoros': 'KM',
     u'Congo': 'CG',
     u'Congo, the Democratic Republic of the': 'CD',
     u'Cook Islands': 'CK',
     u'Costa Rica': 'CR',
     u'Country name': 'Code',
     u'Croatia': 'HR',
     u'Cuba': 'CU',
     u'Curaçao': 'CW',
     u'Cyprus': 'CY',
     u'Czech Republic': 'CZ',
     u"Côte d'Ivoire": 'CI',
     u'Denmark': 'DK',
     u'Djibouti': 'DJ',
     u'Dominica': 'DM',
     u'Dominican Republic': 'DO',
     u'Ecuador': 'EC',
     u'Egypt': 'EG',
     u'El Salvador': 'SV',
     u'Equatorial Guinea': 'GQ',
     u'Eritrea': 'ER',
     u'Estonia': 'EE',
     u'Ethiopia': 'ET',
     u'Falkland Islands (Malvinas)': 'FK',
     u'Faroe Islands': 'FO',
     u'Fiji': 'FJ',
     u'Finland': 'FI',
     u'France': 'FR',
     u'French Guiana': 'GF',
     u'French Polynesia': 'PF',
     u'French Southern Territories': 'TF',
     u'Gabon': 'GA',
     u'Gambia': 'GM',
     u'Georgia': 'GE',
     u'Germany': 'DE',
     u'Ghana': 'GH',
     u'Gibraltar': 'GI',
     u'Greece': 'GR',
     u'Greenland': 'GL',
     u'Grenada': 'GD',
     u'Guadeloupe': 'GP',
     u'Guam': 'GU',
     u'Guatemala': 'GT',
     u'Guernsey': 'GG',
     u'Guinea': 'GN',
     u'Guinea-Bissau': 'GW',
     u'Guyana': 'GY',
     u'Haiti': 'HT',
     u'Heard Island and McDonald Islands': 'HM',
     u'Holy See (Vatican City State)': 'VA',
     u'Honduras': 'HN',
     u'Hong Kong': 'HK',
     u'Hungary': 'HU',
     u'ISO 3166-2:GB': '(.uk)',
     u'Iceland': 'IS',
     u'India': 'IN',
     u'Indonesia': 'ID',
     u'Iran, Islamic Republic of': 'IR',
     u'Iraq': 'IQ',
     u'Ireland': 'IE',
     u'Isle of Man': 'IM',
     u'Israel': 'IL',
     u'Italy': 'IT',
     u'Jamaica': 'JM',
     u'Japan': 'JP',
     u'Jersey': 'JE',
     u'Jordan': 'JO',
     u'Kazakhstan': 'KZ',
     u'Kenya': 'KE',
     u'Kiribati': 'KI',
     u"Korea, Democratic People's Republic of": 'KP',
     u'Korea, Republic of': 'KR',
     u'Kuwait': 'KW',
     u'Kyrgyzstan': 'KG',
     u"Lao People's Democratic Republic": 'LA',
     u'Latvia': 'LV',
     u'Lebanon': 'LB',
     u'Lesotho': 'LS',
     u'Liberia': 'LR',
     u'Libya': 'LY',
     u'Liechtenstein': 'LI',
     u'Lithuania': 'LT',
     u'Luxembourg': 'LU',
     u'Macao': 'MO',
     u'Macedonia, the former Yugoslav Republic of': 'MK',
     u'Madagascar': 'MG',
     u'Malawi': 'MW',
     u'Malaysia': 'MY',
     u'Maldives': 'MV',
     u'Mali': 'ML',
     u'Malta': 'MT',
     u'Marshall Islands': 'MH',
     u'Martinique': 'MQ',
     u'Mauritania': 'MR',
     u'Mauritius': 'MU',
     u'Mayotte': 'YT',
     u'Mexico': 'MX',
     u'Micronesia, Federated States of': 'FM',
     u'Moldova, Republic of': 'MD',
     u'Monaco': 'MC',
     u'Mongolia': 'MN',
     u'Montenegro': 'ME',
     u'Montserrat': 'MS',
     u'Morocco': 'MA',
     u'Mozambique': 'MZ',
     u'Myanmar': 'MM',
     u'Namibia': 'NA',
     u'Nauru': 'NR',
     u'Nepal': 'NP',
     u'Netherlands': 'NL',
     u'New Caledonia': 'NC',
     u'New Zealand': 'NZ',
     u'Nicaragua': 'NI',
     u'Niger': 'NE',
     u'Nigeria': 'NG',
     u'Niue': 'NU',
     u'Norfolk Island': 'NF',
     u'Northern Mariana Islands': 'MP',
     u'Norway': 'NO',
     u'Oman': 'OM',
     u'Pakistan': 'PK',
     u'Palau': 'PW',
     u'Palestine, State of': 'PS',
     u'Panama': 'PA',
     u'Papua New Guinea': 'PG',
     u'Paraguay': 'PY',
     u'Peru': 'PE',
     u'Philippines': 'PH',
     u'Pitcairn': 'PN',
     u'Poland': 'PL',
     u'Portugal': 'PT',
     u'Puerto Rico': 'PR',
     u'Qatar': 'QA',
     u'Romania': 'RO',
     u'Russian Federation': 'RU',
     u'Rwanda': 'RW',
     u'Réunion': 'RE',
     u'Saint Barthélemy': 'BL',
     u'Saint Helena, Ascension and Tristan da Cunha': 'SH',
     u'Saint Kitts and Nevis': 'KN',
     u'Saint Lucia': 'LC',
     u'Saint Martin (French part)': 'MF',
     u'Saint Pierre and Miquelon': 'PM',
     u'Saint Vincent and the Grenadines': 'VC',
     u'Samoa': 'WS',
     u'San Marino': 'SM',
     u'Sao Tome and Principe': 'ST',
     u'Saudi Arabia': 'SA',
     u'Senegal': 'SN',
     u'Serbia': 'RS',
     u'Seychelles': 'SC',
     u'Sierra Leone': 'SL',
     u'Singapore': 'SG',
     u'Sint Maarten (Dutch part)': 'SX',
     u'Slovakia': 'SK',
     u'Slovenia': 'SI',
     u'Solomon Islands': 'SB',
     u'Somalia': 'SO',
     u'South Africa': 'ZA',
     u'South Georgia and the South Sandwich Islands': 'GS',
     u'South Sudan': 'SS',
     u'Spain': 'ES',
     u'Sri Lanka': 'LK',
     u'Sudan': 'SD',
     u'Suriname': 'SR',
     u'Svalbard and Jan Mayen': 'SJ',
     u'Swaziland': 'SZ',
     u'Sweden': 'SE',
     u'Switzerland': 'CH',
     u'Syrian Arab Republic': 'SY',
     u'Taiwan, Province of China': 'TW',
     u'Tajikistan': 'TJ',
     u'Tanzania, United Republic of': 'TZ',
     u'Thailand': 'TH',
     u'Timor-Leste': 'TL',
     u'Togo': 'TG',
     u'Tokelau': 'TK',
     u'Tonga': 'TO',
     u'Trinidad and Tobago': 'TT',
     u'Tunisia': 'TN',
     u'Turkey': 'TR',
     u'Turkmenistan': 'TM',
     u'Turks and Caicos Islands': 'TC',
     u'Tuvalu': 'TV',
     u'Uganda': 'UG',
     u'Ukraine': 'UA',
     u'United Arab Emirates': 'AE',
     u'United Kingdom': 'GB',
     u'United States': 'US',
     u'United States Minor Outlying Islands': 'UM',
     u'Uruguay': 'UY',
     u'Uzbekistan': 'UZ',
     u'Vanuatu': 'VU',
     u'Venezuela, Bolivarian Republic of': 'VE',
     u'Viet Nam': 'VN',
     u'Virgin Islands, British': 'VG',
     u'Virgin Islands, U.S.': 'VI',
     u'Wallis and Futuna': 'WF',
     u'Western Sahara': 'EH',
     u'Yemen': 'YE',
     u'Zambia': 'ZM',
     u'Zimbabwe': 'ZW'
    }
    listap_o = OrderedDict(sorted(listap.items(), key=operator.itemgetter(0)))
    #sorted(listap.items(), key=operator.itemgetter(0))
    #print(listap[0])    
    disabled=''
    debug=1
    check_informacion=[]
    if request.method == "GET" and "q_estudiante" in request.GET:
        estudiante=request.GET["q_estudiante"]
        debug=estudiante
        if estudiante:
            debug=2
            palabras = estudiante.split()
            
            try:
                datos_es= Students.objects.get(reduce(operator.and_, (Q(id_students__first_name__icontains=x) | Q(id_students__father_last_name__icontains=x) | Q(id_students__mother_last_name__icontains=x)| Q(id_students__identity__icontains=x)for x in palabras)))
                request.session['id_students']=datos_es.id_students_id #cookies
            except Students.DoesNotExist:
                datos_es = None
                return render(request,'admisiones_sam/EditFormsStudents.html')
            
            #datose=Students.objects.get(id_students=datos_e)   
            try:
                datos_e=CustomUser.objects.get(id=datos_es.id_students_id)
            except CustomUser.DoesNotExist:
                datos_e=None
            
            # datos papa
            try:
                datos_p=CustomUser.objects.get(id=datos_es.id_father_id)
            except CustomUser.DoesNotExist:
                datos_p=None
                check_informacion.append("P")
            try:
                datos_pd=People.objects.get(id_people_id=datos_es.id_father_id)
                datos_pr=RelationshipStudent.objects.update_or_create(id_people_id=datos_pd.id,id_students_id=datos_es.id, defaults={
                    "live_student":True,
                    "withdraw":True                    
                    })
            except People.DoesNotExist:
                datos_pd=None
                datos_pr=None
            
            #datos mama
            try:
                datos_m=CustomUser.objects.get(id=datos_es.id_mother_id)
            except CustomUser.DoesNotExist:
                datos_m=None
                check_informacion.append("M")
            try:
                datos_md=People.objects.get(id_people_id=datos_es.id_mother_id)
                datos_mr=RelationshipStudent.objects.update_or_create(id_people_id=datos_md.id,id_students_id=datos_es.id, defaults={
                    "live_student":True,
                    "withdraw":True                    
                    })
            except People.DoesNotExist:
                datos_md=None
                datos_mr=None
            
            #datos representante
            try:
                datos_t=CustomUser.objects.get(id=datos_es.id_tutor_id)
            except CustomUser.DoesNotExist:
                datos_t=None
            try:
                datos_td=People.objects.get(id_people_id=datos_es.id_tutor_id)
                datos_tr=RelationshipStudent.objects.update_or_create(id_people_id=datos_td.id,id_students_id=datos_es.id, defaults={
                    "live_student":True,
                    "withdraw":True
                    })
            except People.DoesNotExist:
                datos_td=None   
                datos_tr=None
                
                
            if datos_es.id_father_id==datos_es.id_tutor_id :
                representante_check='P'
            elif datos_es.id_mother_id==datos_es.id_tutor_id:
                representante_check='M'
            else:
                representante_check='O'
            
            
            return render(request,'admisiones_sam/EditFormsStudents.html',{"estudiante":datos_es,"papa":datos_p,"papa_det":datos_pd,"papa_rel":datos_pr,"mama":datos_m,"mama_det":datos_md,"mama_rel":datos_mr,"tutor":datos_t,"tutor_det":datos_td,"tutor_rel":datos_tr,"listag":listag,"listap":listap_o,"deshabilita":disabled,"debug":debug,"representante_check":representante_check,"check_informacion":check_informacion})
        
        else:
            return render(request,'admisiones_sam/EditFormsStudents.html',{"debug":debug})
        
        
    elif request.method == "POST":
        checkbox_papa = request.POST.get('valida_papa', '')
        checkbox_mama = request.POST.get('valida_mama', '')
        check_informacion=[]
        
        id_st=request.session['id_students']
        
        try:
            datos_es= Students.objects.get(id_students_id=id_st)
             #datos_es.applied_grade=request.POST.get('s_grado_aplica')
            datos_es.medical_information=request.POST.get('s_informacion_medica')
            datos_es.save()
            datos_e=CustomUser.objects.get(id=datos_es.id_students_id)
            datos_es.id_students.identity=request.POST.get('s_identificacion')
            datos_es.id_students.first_name=request.POST.get('s_nombres')
            datos_es.id_students.father_last_name=request.POST.get('s_apellidos')
            datos_es.id_students.mother_last_name=request.POST.get('s_apellidos_2')
            datos_es.id_students.gender=request.POST.get('s_genero')
            datos_es.id_students.birthdate=request.POST.get('s_fecha_nacimiento')
            datos_es.id_students.address=request.POST.get('s_direccion')
            datos_es.id_students.birth_country=request.POST.get('s_pais_nacimiento')
            datos_es.id_students.country_home=request.POST.get('s_pais_residencia')
            datos_es.id_students.save()

            estudiante = Student.objects.get(user = datos_es.id_students)


        except Students.DoesNotExist:
            datos_es=None

        except Student.DoesNotExist:
            estudiante = Student.objects.create(user = datos_es.id_students)


        #chequear si formulario tenia espacio del tutor

        if datos_es.id_tutor == datos_es.id_father or datos_es.id_tutor == datos_es.id_mother:
            
            if datos_es.id_tutor == datos_es.id_father:
                representante = "P"
            else:
                representante = "M"
        else:
            representante = "O"
      
        # datos papa
        if not checkbox_papa: #SI EXISTE INFORMACION DEL PADRE

            #OBTENER LA INFORMACION DEL POST
            p_identificacion=request.POST.get('p_identificacion')
            p_nombres=request.POST.get('p_nombres')
            p_apellidos=request.POST.get('p_apellidos')
            p_apellidos_2=request.POST.get('p_apellidos_2')
            p_genero="M"
            p_fecha_nacimiento=request.POST.get('p_fecha_nacimiento')
            p_pais_nacimiento=request.POST.get('p_pais_nacimiento')
            p_pais_residencia=request.POST.get('p_pais_residencia')
            p_telefono=request.POST.get('p_telefono')
            p_cell=request.POST.get('p_cell')
            p_mail=request.POST.get('p_mail')
            p_direccion=request.POST.get('p_direccion')
            p_vive_con=request.POST.get('p_vive_con')
            p_retirar=request.POST.get('p_retirar')
            p_difunto = request.POST.get("p_difunto", "")
            p_instruction=request.POST.get('p_instruccion')
            p_profession=request.POST.get('p_profesion')
            p_address_work=request.POST.get('p_trabajo')
            p_telofi=request.POST.get('p_telofi')

            #CREAR O ACTUALIZAR EL USUARIO DEL PADRE CON LOS DATOS DE POST
            usuario_papa, creado = CustomUser.objects.update_or_create(email = p_mail, defaults = {
                "identity": p_identificacion,
                "first_name": p_nombres,
                "father_last_name": p_apellidos,
                "mother_last_name": p_apellidos_2,
                "gender":"M",
                "birthdate":p_fecha_nacimiento,
                "phone":p_telefono,
                "cell":p_cell,
                "address":p_direccion, 
                "birth_country":p_pais_nacimiento, 
                "country_home":p_pais_residencia
                })

            if p_difunto == "true":
                usuario_papa.is_active = False
                p_difunto=True
            else:
                p_difunto=False

            if p_vive_con=="true" :
                p_vive_con=True
            else:
                p_vive_con=False
            
            if p_retirar=="true" :
                p_retirar=True
            else:
                p_retirar=False                    

            usuario_papa.save()

            datos_p = usuario_papa

            #crear o actualizar la persona papa con el usuario papa creado

            persona_papa, creado = People.objects.update_or_create(id_people = usuario_papa, defaults={
                "is_dead": p_difunto,
                "instruction":p_instruction,
                "profession":p_profession,
                "address_work":p_address_work,
                "office": p_telofi,
                "e_mail":p_mail                    
                })

            datos_pd = persona_papa

            #crea o actualiza la relacion

            datos_pr, creado = RelationshipStudent.objects.update_or_create(id_people = datos_pd, id_students = datos_es, defaults ={
                "live_student":p_vive_con,
                "withdraw": p_retirar
                })
          

            #para la nueva estructura, crear o actualizar los objetos relative, relative details y student_relative

            papa, creado = Relative.objects.update_or_create(user = usuario_papa, defaults = {
                "through": "Admisiones"
                })

            detalle_papa, creado = Relative_Details.objects.update_or_create(relative = papa, defaults = {
                "alive": not(p_difunto),
                "instruction":p_instruction,
                "profession": p_profession,
                "address_work":p_address_work,
                "office":p_telofi
                })

            relacion_papa, creado = Student_Relative.objects.update_or_create(student = estudiante, relative = papa, defaults={
                "relationship": "PADRE",
                "live_together": p_vive_con,
                "withdraw": p_retirar                
                })



            datos_es.id_father = usuario_papa
            datos_es.save()

            if representante == "P":
                relacion_papa.legal_representantive = True
                relacion_papa.notifications = True
                relacion_papa.save()                
                datos_es.id_tutor = usuario_papa
                datos_es.save()                


        else: #si no existe informacion del padre

            datos_p= None
            usuario_papa = None
            persona_papa = None
            datos_pd=None
            datos_pr=None            
            check_informacion.append("P")


        #datos mama
        if not checkbox_mama: #SI EXISTE INFORMACION DE MADRE

            #OBTENER LA INFORMACION DEL POST
            m_identificacion=request.POST.get('m_identificacion')
            m_nombres=request.POST.get('m_nombres')
            m_apellidos=request.POST.get('m_apellidos')
            m_apellidos_2=request.POST.get('m_apellidos_2')
            m_genero="F"
            m_fecha_nacimiento=request.POST.get('m_fecha_nacimiento')
            m_pais_nacimiento=request.POST.get('m_pais_nacimiento')
            m_pais_residencia=request.POST.get('m_pais_residencia')
            m_telefono=request.POST.get('m_telefono')
            m_cell=request.POST.get('m_cell')
            m_mail=request.POST.get('m_mail')
            m_direccion=request.POST.get('m_direccion')
            m_vive_con=request.POST.get('m_vive_con')
            m_retirar=request.POST.get('m_retirar')
            m_difunto = request.POST.get("m_difunto", "")
            m_instruction=request.POST.get('m_instruccion')
            m_profession=request.POST.get('m_profesion')
            m_address_work=request.POST.get('m_trabajo')
            m_telofi=request.POST.get('m_telofi')

            #CREAR O ACTUALIZAR EL USUARIO DEL PADRE CON LOS DATOS DE POST
            usuario_mama, creado = CustomUser.objects.update_or_create(email = m_mail, defaults = {
                "identity": m_identificacion,
                "first_name": m_nombres,
                "father_last_name": m_apellidos,
                "mother_last_name": m_apellidos_2,
                "gender":"M", "birthdate":m_fecha_nacimiento,
                "phone":m_telefono,
                "cell":m_cell,
                "address":m_direccion, 
                "birth_country":m_pais_nacimiento, 
                "country_home":m_pais_residencia
                })

            if m_difunto == "true":
                usuario_mama.is_active = False
                m_difunto=True
            else:
                m_difunto=False

            if m_vive_con=="true" :
                m_vive_con=True
            else:
                m_vive_con=False
            
            if m_retirar=="true" :
                m_retirar=True
            else:
                m_retirar=False                    

            usuario_mama.save()

            datos_m = usuario_mama

            #crear o actualizar la persona mama con el usuario mama creado

            persona_mama, creado = People.objects.update_or_create(id_people = usuario_mama, defaults={
                "is_dead": m_difunto,
                "instruction":m_instruction,
                "profession":m_profession,
                "address_work":m_address_work,
                "office": m_telofi,
                "e_mail":m_mail                    
                })

            datos_md = persona_mama

            #crea o actualiza la relacion

            datos_mr, creado = RelationshipStudent.objects.update_or_create(id_people = datos_md, id_students = datos_es, defaults ={
                "live_student":m_vive_con,
                "withdraw": m_retirar
                })
          

            #para la nueva estructura, crear o actualizar los objetos relative, relative details y student_relative

            mama, creado = Relative.objects.update_or_create(user = usuario_mama, defaults = {
                "through": "Admisiones"
                })

            detalle_mama, creado = Relative_Details.objects.update_or_create(relative = mama, defaults = {
                "alive": not(m_difunto),
                "instruction":m_instruction,
                "profession": m_profession,
                "address_work":m_address_work,
                "office":m_telofi
                })

            relacion_mama, creado = Student_Relative.objects.update_or_create(student = estudiante, relative = mama, defaults={
                "relationship": "MADRE",
                "live_together": m_vive_con,
                "withdraw": m_retirar                
                })



            datos_es.id_mother = usuario_mama
            datos_es.save()


            if representante == "M":
                relacion_mama.legal_representantive = True
                relacion_mama.notifications = True
                relacion_mama.save()
                datos_es.id_tutor = usuario_mama
                datos_es.save() 


        else: #si no existe informacion del madre

            datos_m= None
            usuario_mama = None
            persona_mama = None
            datos_md=None
            datos_mr=None            
            check_informacion.append("M")

        #si el representante no es ni mamá ni papá

        if representante == "O":
            if usuario_papa is not None and request.POST["r_mail"] == usuario_papa.email:
                datos_es.id_tutor = usuario_papa
                relacion_papa.legal_representantive = True
                relacion_papa.notifications = True
                relacion_papa.save()

            elif usuario_mama is not None and request.POST["r_mail"] == usuario_mama.email:
                datos_es.id_tutor = usuario_mama
                relacion_mama.legal_representantive = True
                relacion_mama.notifications = True
                relacion_mama.save()

            else:

                t_identificacion=request.POST.get('r_identificacion')
                t_nombres=request.POST.get('r_nombres')
                t_apellidos=request.POST.get('r_apellidos')
                t_apellidos_2=request.POST.get('r_apellidos_2')
                t_genero=request.POST.get("r_genero")
                t_fecha_nacimiento=request.POST.get('r_fecha_nacimiento')
                t_pais_nacimiento=request.POST.get('r_pais_nacimiento')
                t_pais_residencia=request.POST.get('r_pais_residencia')
                t_telefono=request.POST.get('r_telefono')
                t_cell=request.POST.get('r_cell')
                t_mail=request.POST.get('r_mail')
                t_direccion=request.POST.get('r_direccion')
                t_vive_con=request.POST.get('r_vive_con')
                t_retirar=request.POST.get('r_retirar')
                t_difunto = request.POST.get("r_difunto", "")
                t_instruction=request.POST.get('r_instruccion')
                t_profession=request.POST.get('r_profesion')
                t_address_work=request.POST.get('r_trabajo')
                t_telofi=request.POST.get('r_telofi')

                usuario_tutor, creado = CustomUser.objects.update_or_create(email = t_mail, defaults = {
                    "identity": t_identificacion,
                    "first_name": t_nombres,
                    "father_last_name": t_apellidos,
                    "mother_last_name": t_apellidos_2,
                    "gender":t_genero, 
                    "birthdate":t_fecha_nacimiento,
                    "phone":t_telefono,
                    "cell":t_cell,
                    "address":t_direccion, 
                    "birth_country":t_pais_nacimiento, 
                    "country_home":t_pais_residencia
                    })

                if t_vive_con=="true" :
                    t_vive_con=True
                else:
                    t_vive_con=False
                
                if t_retirar=="true" :
                    t_retirar=True
                else:
                    t_retirar=False                    

                datos_t = usuario_tutor               


                persona_tutor, creado = People.objects.update_or_create(id_people = usuario_tutor, defaults={
                    "is_dead": False,
                    "instruction":t_instruction,
                    "profession":t_profession,
                    "address_work":t_address_work,
                    "office": t_telofi,
                    "e_mail":t_mail                    
                    })

                datos_td = persona_tutor

                #crea o actualiza la relacion

                datos_tr, creado = RelationshipStudent.objects.update_or_create(id_people = datos_td, id_students = datos_es, defaults ={
                    "live_student":t_vive_con,
                    "withdraw": t_retirar
                    })
              
                #para la nueva estructura, crear o actualizar los objetos relative, relative details y student_relative

                tutor, creado = Relative.objects.update_or_create(user = usuario_tutor, defaults = {
                    "through": "Admisiones"
                    })

                detalle_tutor, creado = Relative_Details.objects.update_or_create(relative = tutor, defaults = {
                    "alive": not(t_difunto),
                    "instruction":t_instruction,
                    "profession": t_profession,
                    "address_work":t_address_work,
                    "office":t_telofi
                    })

                relacion_tutor, creado = Student_Relative.objects.update_or_create(student = estudiante, relative = tutor, defaults={
                    "relationship": "OTRO",
                    "live_together": t_vive_con,
                    "withdraw": t_retirar                
                    })

                datos_es.id_tutor = usuario_tutor

            datos_es.save()
 
        return render(request,'admisiones_sam/EditFormsStudents.html',{"debug":debug}) #,{"estudiante":datos_es,"papa":datos_p,"mama":datos_m,"tutor":datos_t,"listag":listag,"listap":listap_o})
    

    else:
        return render(request,'admisiones_sam/EditFormsStudents.html',{"debug":debug})
       
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)    
def NewForms(request,id_applications):
    
    estudiante=Applications.objects.get(pk=id_applications)
    request.session['id_aplicacion']=estudiante.pk  #cookies
    if request.method == "POST":
      
        form = FormsNewform(request.POST)
        if form.is_valid():
            
            item = form.save(commit=False)
            item.date_form = date.today()
            item.id_applications=estudiante
            item.save()
            id_forms=item.pk
            
            return render(request, "admisiones_sam/FormsMessage.html", {"estudiante":estudiante})
        else:
            
            print form.errors
            return render(request, "admisiones_sam/NewForms.html", {"form":form,"estudiante":estudiante})
    else:
       
        form = FormsNewform(initial={'birth_country': 'EC','country_home': 'EC'})
        return render(request, "admisiones_sam/NewForms.html", {"form":form,"estudiante":estudiante})

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def EditForms(request):

    if request.method == "GET" and "q_estudiante" in request.GET:
        estudiante=request.GET["q_estudiante"]
        
        
        if estudiante:
            palabras = estudiante.split()
            try: 
                datos= Forms.objects.get(reduce(operator.or_, (Q(name_student__icontains=x) | Q(surname_student__icontains=x) | Q(identity_student__icontains=x)for x in palabras)))
                request.session['id_forms']=datos.pk  #cookies
            except Forms.DoesNotExist:
                datos = None
                form = FormsNewform()
                return render(request,'admisiones_sam/EditForms.html', {"form":form})
            
            form = FormsNewform(instance = datos)   
            return render(request,'admisiones_sam/EditForms.html',{"form":form,"aplicante":estudiante})
        else:
            
            form = FormsNewform()
            return render(request,'admisiones_sam/EditForms.html',{"form":form})
    else:
        
        form = FormsNewform()
        return render(request,'admisiones_sam/EditForms.html', {"form":form})
    
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def EditFormsPost(request):
    id_forms=request.session['id_forms']
    datos=Forms.objects.get(pk=id_forms)
    
    debug='1'
    if request.method == "POST":
        debug='2'
        form = FormsNewform(request.POST, instance = datos)
        if form.is_valid():
            debug='3'
            form.save()
            return redirect('MensajeFormularioActualiza')
        else:
            debug='4'
            print form.errors
            return render(request,'admisiones_sam/EditForms.html', {"form":form,"datos":datos})
    else:
        debug='5'
        form = FormsNewform(instance=datos)
        return render(request,'admisiones_sam/EditForms.html', {"form":form,"datos":datos})

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def KeepQuotas(request):
    
    cupos=Quotas.objects.all()
    
    consultagrados=Type_Grade.objects.all().order_by('id_typegrade')
    lista=[]
    
    for y in consultagrados:
        if y.typegrade in lista:
            pass
        else:
            lista.append(y.typegrade)
    
    if request.method == "POST":
        form = QuotaNewform(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            grado=item.grade
            if not Quotas.objects.filter(grade=item.grade).count()>0:
                item.save()
                form=QuotaNewform()
                return render(request, "admisiones_sam/KeepQuotas.html", {"form":form,"cupos":cupos,"lista":lista})
            else:
                return render(request,"admisiones_sam/QuotasNewMessage.html",{"grado":grado})
        else:
            print form.errors
            return render(request, "admisiones_sam/KeepQuotas.html", {"form":form,"cupos":cupos,"lista":lista})
    else:
        
        form = QuotaNewform()
        return render(request, "admisiones_sam/KeepQuotas.html", {"form":form,"cupos":cupos,"lista":lista})

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def DeleteQuotas(request, id_quotas): 
    cupos = Quotas.objects.get(pk = id_quotas)
    cupos.delete()
    return redirect('/admisiones/MantenerCupos')
   

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def EditQuotas(request, id_quotas):
  
    cuotas = Quotas.objects.get(pk = id_quotas)
    
    consultagrados=Type_Grade.objects.all().order_by('id_typegrade')
    lista=[]
    
    for y in consultagrados:
        if y.typegrade in lista:
            pass
        else:
            lista.append(y.typegrade)
            
    
    if request.method == "POST":
        form = QuotaNewform(request.POST, instance = cuotas)    
        if form.is_valid():
           form.save()
           return redirect('/admisiones/MantenerCupos')
        else:
            print form.errors
            return render(request,'admisiones_sam/EditQuotas.html', {"form":form,"cuotas":cuotas,"lista":lista})
    else:
        form = QuotaNewform(instance = cuotas)
        return render(request,'admisiones_sam/EditQuotas.html', {"form":form,"cuotas":cuotas,"lista":lista})        

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def KeepMails(request):
    
    mail=Mails.objects.all()
    
    if request.method == "POST":
        form = MailsNewform(request.POST)
        if form.is_valid():
            form.save()
           
            return redirect('/admisiones/MantenerMails')
            
        else:
            print form.errors
            return render(request, "admisiones_sam/KeepMails.html", {"form":form,"mail":mail})
    else:
        
        form = MailsNewform()
        return render(request, "admisiones_sam/KeepMails.html", {"form":form,"mail":mail})

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def DeleteMails(request, id_mails): 
    mail = Mails.objects.get(id_mails = id_mails)
    mail.delete()
    return redirect('/admisiones/MantenerMails')
   
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def EditMails(request, id_mails):
  
    mail = Mails.objects.get(id_mails = id_mails)
    
    if request.method == "POST":
        form = MailsNewform(request.POST, instance = mail)  
        if form.is_valid():
           form.save()
           return redirect('/admisiones/MantenerMails')
        else:
            print form.errors
            return render(request,'admisiones_sam/EditMails.html', {"form":form,"mail":mail})
    else:
        form = MailsNewform(instance = mail)
        return render(request,'admisiones_sam/EditMails.html', {"form":form,"mail":mail})    

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def KeepGrade(request):
    
    grados=Type_Grade.objects.all()
    
    if request.method == "POST":
        form = GradeNewform(request.POST)
        if form.is_valid():
            form.save()
           
            return redirect('/admisiones/MantenerGrados')
            
        else:
            print form.errors
            return render(request, "admisiones_sam/KeepGrade.html", {"form":form,"grados":grados})
    else:
        
        form = GradeNewform()
        return render(request, "admisiones_sam/KeepGrade.html", {"form":form,"grados":grados})

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def DeleteGrade(request, id_typegrade): 
    grado = Type_Grade.objects.get(id_typegrade = id_typegrade)
    grado.delete()
    return redirect('/admisiones/MantenerGrados')
   
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def EditGrade(request, id_typegrade):
  
    grado = Type_Grade.objects.get(id_typegrade = id_typegrade)
    
    if request.method == "POST":
        form = GradeNewform(request.POST, instance = grado) 
        if form.is_valid():
           form.save()
           return redirect('/admisiones/MantenerGrados')
        else:
            print form.errors
            return render(request,'admisiones_sam/EditGrade.html', {"form":form,"grado":grado})
    else:
        form = GradeNewform(instance = grado)
        return render(request,'admisiones_sam/EditGrade.html', {"form":form,"grado":grado})    

        
def CreateInvoices(request,id_applications):
    aplicacion=Applications.objects.get(id_applications=id_applications)
    if Invoices.objects.filter(id_applications=id_applications).count()>0:
        return redirect('CargaDocumentos',aplicacion.id_applications)
    elif request.method == "POST":
        form = InvoicesNewform(request.POST)
        
        aplicante=aplicacion.name_student +" "+ aplicacion.surname_student
        if form.is_valid():
            item = form.save(commit=False)
            item.id_applications=aplicacion
            if not Invoices.objects.filter(id_applications=id_applications).count()>0:
                datos=item.save()
                #return render(request, "admisiones_sam/InvoicesMessage.html", {"aplicante":aplicante})
                return redirect('CargaDocumentos',aplicacion.id_applications)
              
            else:
                return render(request, "admisiones_sam/InvoicesMessage2.html", {"datos":aplicacion})
            
        else:
            print form.errors
            form = InvoicesNewform()
            return render(request, "admisiones_sam/InvoicesCreate.html", {"form":form,"id_applications":id_applications})
    else:
        form = InvoicesNewform()
        return render(request, "admisiones_sam/InvoicesCreate.html", {"form":form,"id_applications":id_applications})


def NewInvoices(request,id_applications):
    aplicacion=Applications.objects.get(id_applications=id_applications)

    if request.method == "POST":
        form = InvoicesNewform(request.POST)
        
        aplicante=aplicacion.name_student +" "+ aplicacion.surname_student
        if form.is_valid():
            item = form.save(commit=False)
            item.id_applications=aplicacion
            if not Invoices.objects.filter(id_applications=id_applications).count()>0:
                datos=item.save()
                return render(request, "admisiones_sam/InvoicesNewMessage2.html", {"aplicante":aplicante})
                
            else:
                return render(request, "admisiones_sam/InvoicesNewMessage3.html", {"datos":aplicacion})
            
        else:
            print form.errors
            form = InvoicesNewform()
            return render(request, "admisiones_sam/InvoicesNew.html", {"form":form,"id_applications":id_applications})
    else:
        form = InvoicesNewform()
        return render(request, "admisiones_sam/InvoicesNew.html", {"form":form,"id_applications":id_applications})


def KeepInvoices(request):
    
    if request.method == "GET" and "q_aplicantes" in request.GET:
        aplicante=request.GET["q_aplicantes"]
        
     
        if aplicante:
            palabras = aplicante.split()
            
            aplicaciones=Applications.objects.get(reduce(operator.and_, (Q(name_student__icontains=x) | Q(surname_student__icontains=x) for x in palabras)))
            for estado in aplicaciones.state_applications_set.filter(final_date__isnull=True):
                if estado.state is not "Terminado" :
                    aplicaciones=Applications.objects.get(id_applications=estado.id_applications_id)
                   
                else:
                    aplicaciones=None
                    form = InvoicesNewform()
                    return render(request,'admisiones_sam/KeepInvoices.html',{"form":form})
       
            """ try:
                aplicaciones=Applications.objects.get(reduce(operator.and_, (Q(name_student__icontains=x) | Q(surname_student__icontains=x) for x in palabras)),~Q(state_applications__state = "Terminado"), state_applications__final_date__isnull = True)
                #request.session['id_applications']=aplicaciones.pk  #cookies
            except ObjectDoesNotExist:
                aplicaciones=None
                form = InvoicesNewform()
                return render(request,'admisiones_sam/KeepInvoices.html',{"form":form})"""
                
                
            if aplicaciones:
                try:
                    datos=Invoices.objects.get(id_applications=aplicaciones)
                    #request.session['id_invoice']=datos.id_invoice  #cookies
                    form = InvoicesNewform(instance=datos)  
                    
                    return render(request,'admisiones_sam/KeepInvoices.html',{"form":form,"aplicante":aplicaciones,"facturas":datos})
                except Invoices.DoesNotExist:
                    datos=None
                    form = InvoicesNewform()
                   
                    return render(request,'admisiones_sam/InvoicesSearchMessage.html', {"form":form,"aplicante":aplicante,"aplicaciones":aplicaciones})
            else:
                form = InvoicesNewform()
                return render(request,'admisiones_sam/KeepInvoices.html',{"form":form})
        else:
            
            form = InvoicesNewform()
            return render(request,'admisiones_sam/KeepInvoices.html',{"form":form})
           
    else:    
        form = InvoicesNewform()
        return render(request,'admisiones_sam/KeepInvoices.html', {"form":form})


@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def DeleteInvoices(request, id_invoice): 
    facturas = Invoices.objects.get(id_invoice = id_invoice)
    facturas.delete()
    return redirect('/admisiones/MantenerFacturas')
   

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def EditInvoices(request, id_invoice):
  
    facturas = Invoices.objects.get(id_invoice = id_invoice)
   
    if request.method == "POST":
        form = InvoicesNewform(request.POST, instance = facturas)   
        if form.is_valid():
           form.save()
           return redirect('/admisiones/MantenerFacturas')
        else:
            print form.errors
            return render(request,'admisiones_sam/EditInvoices.html', {"form":form,"factura":facturas.id_invoice})
    else:
        form = InvoicesNewform(instance = facturas)
        return render(request,'admisiones_sam/EditInvoices.html', {"form":form,"factura":facturas.id_invoice})       
 


def KeepTest(request):
    test=None
    mensaje = MailerMessage()
    
    lista=['Aprobado','Reprobado','Lista Espera']
    materias=['Inglés','Matemáticas','Lengua','Psicológicas','Ciencias','Evaluación Grupal','Evaluación Individual','Otros']
    contador=0
    formA = NewReportsTests()
    DetReportFormSet=modelformset_factory(Det_Reports,fields=('materia', 'value','observations'), extra=5)
    formB=DetReportFormSet(queryset=Det_Reports.objects.none())
    debug=0       
    if request.method == "GET" and "q_aplicantes" in request.GET:
        aplicante=request.GET["q_aplicantes"]
     
     
        if aplicante:
       
            palabras = aplicante.split()
            aplicaciones=Applications.objects.get(reduce(operator.and_, (Q(name_student__icontains=x) | Q(surname_student__icontains=x) for x in palabras)))
            #,~Q(state_applications__state = "Terminado"), state_applications__final_date__isnull = True
            
            test=Test.objects.filter(id_applications_id=aplicaciones.id_applications,state=True).extra(select={
                'date_test':'''SELECT admisiones_sam_type_tests.date_test
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'grade':'''SELECT admisiones_sam_type_tests.grade
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'time_test':'''SELECT admisiones_sam_type_tests.time_test
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'teacher':'''SELECT admisiones_sam_type_tests.teacher
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'type_test':'''SELECT admisiones_sam_type_tests.type_test
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                })
            
            if test:
            
                if not Reports_Tests.objects.filter(id_test_id=test).count()>0:
                
                    for x in test:
                        request.session['id_test']=x.pk  #cookies   
                         
                    return render(request,"admisiones_sam/KeepTest.html",{"form":formA,"form1":formB,"lista":lista,"test":test,"aplicante":aplicante,"materias":materias})
                else:
                   # mensaje no se puede ingresar registro de evaluacion ya se ha registrado uno.
                   return render(request,"admisiones_sam/KeepTestMessage.html")
            else:
                return render(request,"admisiones_sam/KeepTestMessage1.html")
        else:
            
            formA = NewReportsTests()
            return render(request,"admisiones_sam/KeepTest.html",{"form":formA,"form1":formB,"lista":lista,"test":test,"aplicante":aplicante,"materias":materias})
    
    elif request.method == "POST":
        debug=0
        
        formA=NewReportsTests(request.POST)
       
        id_evaluacion=request.session['id_test']  
        prueba=Test.objects.get(id_test=id_evaluacion)
        aplicante=Applications.objects.get(id_applications=prueba.id_applications_id)
 
        if formA.is_valid():
            debug=1
            a=formA.save(commit=False)
            a.date_report = date.today()
            a.user = request.user
            a.id_test_id  = prueba.id_test
            a.id_applications_id  = prueba.id_applications_id
            a.id_typetest_id  = prueba.id_typetest_id
            
            a.save()
            reporte=a.id_report
            formB=DetReportFormSet(request.POST)
            
            if formB.is_valid():
                debug=2 
                try:
                 
                    for x in formB:
                        try:
                            new_item = x.save(commit=False)
                            new_item.id_report = a
                            new_item.id_test_id = a.id_test_id
                            new_item.id_applications_id = a.id_applications_id
                            new_item.id_typetest_id = a.id_typetest_id
                            #if not new_item.materia== None :
                             #   try:
                            new_item.save()
                            #except:
                            #pass
                        except:
                            contador=contador+1
                            
                    if contador>0:
                        registros=Det_Reports.objects.filter(id_test_id=test.id_test)
                        for x in registros:
                            x.delete()
                        a.delete()
                        return redirect('/admisiones/MantenerEvaluaciones',{"contador":contador,"debug":debug})
                    else:
                        pass
                except:
                    a.delete()
                    return redirect('/admisiones/MantenerEvaluaciones',{"contador":contador})
                 
               
                if contador==0:
                    '''if a.state=='Aprobado' or a.state=='Aprobado/E.F.' :
                        debug=3
                        conta = AccountingDepartment(state='False',id_applications=Applications.objects.get(id_applications=prueba.id_applications_id))
                        conta.save()
                        old_estado=State_Applications.objects.get(id_applications = aplicante.id_applications, final_date__isnull=True)
                        old_estado.final_date=date.today()
                        old_estado.save()
                        new_estado = State_Applications(state='Contabilidad',final_date=None,observations='Registro en Contabilidad',id_applications=aplicante)
                        new_estado.save()
                        formB=DetReportFormSet(queryset=Det_Reports.objects.none())        
                        return redirect('/admisiones/MantenerEvaluaciones')
                    else:  ''' 
                    debug=4
                    old_estado=State_Applications.objects.get(id_applications = aplicante.id_applications, final_date__isnull=True)
                    old_estado.final_date=date.today()
                    old_estado.save()
                    new_estado = State_Applications(state='Evaluaciones',final_date=None,observations='Se mantiene en Evaluacione',id_applications=aplicante)
                    new_estado.save()
                    formB=DetReportFormSet(queryset=Det_Reports.objects.none()) 
                    
                    # Mensaje del sistema  a Admisiones informando que se ha calificado una evaluacion para que admisiones apruebe o repruebe las evaluaciones
                    mensaje.general_message('sorayamena@montebelloacademy.org','admisiones@montebelloacademy.org', 22, aplicante)                    
                    return redirect('/admisiones/MantenerEvaluaciones')
            else:
                
                print formA.errors
                print formB.errors
                return render(request,'admisiones_sam/KeepTest.html', {"form":formA,"form1":formB,"lista":lista,"test":test,"materias":materias,"debug":debug})
    
        else:
          
            print formA.errors
            print formB.errors
            return render(request,'admisiones_sam/KeepTest.html', {"form":formA,"form1":formB,"lista":lista,"test":test,"materias":materias,"debug":debug})
           
       
        
   
    else:    
      
        formA = NewReportsTests()
        return render(request,"admisiones_sam/KeepTest.html",{"form":formA,"form1":formB,"lista":lista,"test":test,"materias":materias,"debug":debug})



def SearchTest(request):
    test=None
    
    report=None
    if request.method == "GET" and "q_aplicantes" in request.GET:
        aplicante=request.GET["q_aplicantes"]
        
     
        if aplicante:
       
            palabras = aplicante.split()
            aplicaciones=Applications.objects.get(reduce(operator.and_, (Q(name_student__icontains=x) | Q(surname_student__icontains=x) for x in palabras)))
            
            
            
            test=Test.objects.filter(id_applications_id=aplicaciones.id_applications).extra(select={
                'date_test':'''SELECT admisiones_sam_type_tests.date_test
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'grade':'''SELECT admisiones_sam_type_tests.grade
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'time_test':'''SELECT admisiones_sam_type_tests.time_test
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'teacher':'''SELECT admisiones_sam_type_tests.teacher
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'type_test':'''SELECT admisiones_sam_type_tests.type_test
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'reporte': '''select  admisiones_sam_reports_tests.id_report
                from admisiones_sam_reports_tests
                where admisiones_sam_reports_tests.id_test_id=id_test'''
                })
            
            report=Reports_Tests.objects.filter(id_test=test)
            
                
            return render(request,'admisiones_sam/SearchTest.html', {"test":test,"aplicante":aplicante,"report":report})
        else:
            return render(request,'admisiones_sam/SearchTest.html')
    else:
        return render(request,'admisiones_sam/SearchTest.html')



def EditTest(request,id_test):
  
    lista=['Aprobado','Reprobado','Lista Espera']
    materias=[u'Inglés',u'Matemáticas',u'Lengua',u'Psicológicas',u'Ciencias',u'Evaluación Grupal',u'Evaluación Individual',u'Otros']
     
    if  Reports_Tests.objects.filter(id_test_id=id_test).count()>0:
        report=Reports_Tests.objects.get(id_test_id=id_test)
        detalle=Det_Reports.objects.filter(id_report_id=report.id_report)
            
        DetReportFormSet=modelformset_factory(Det_Reports,fields=('materia', 'value','observations'), extra=0)
        formB=DetReportFormSet(queryset=detalle)
        
        if request.method == "POST":
            formA = NewReportsTests(request.POST, instance = report)    
            formB=DetReportFormSet(request.POST)
            
            if formA.is_valid():
                a=formA.save(commit=False)
                aplica=a.id_applications_id
                aplicante=Applications.objects.get(id_applications=aplica)
                if formB.is_valid():
                    for x in formB:
                        x.save()
                    estado_a=a.state
                    a.save()
                    
                    #ingreso en contabilidad el registro del aplicante aprobado
                
                    if estado_a=='Aprobado' or estado_a=='Aprobado/E.F.':
                        conta = AccountingDepartment(state='False',id_applications=aplicante)
                        conta.save()
                        old_estado=State_Applications.objects.get(id_applications = aplicante, final_date__isnull=True)
                        old_estado.final_date=date.today()
                        old_estado.save()
                        new_estado = State_Applications(state='Contabilidad',final_date=None,observations='Registro en Contabilidad',id_applications=aplicante)
                        new_estado.save()
                    else:    
                        old_estado=State_Applications.objects.get(id_applications = aplicante, final_date__isnull=True)
                        old_estado.final_date=date.today()
                        old_estado.save()
                        new_estado = State_Applications(state='Evaluaciones',final_date=None,observations='Se mantiene en Evaluacione',id_applications=aplicante)
                        new_estado.save()
                        formB=DetReportFormSet(queryset=Det_Reports.objects.none())        
                        return redirect('/admisiones/ConsultarEvaluaciones')
                        
                return redirect('/admisiones/ConsultarEvaluaciones')
            else:
                print formA.errors
                print formB.errors
                return render(request,'admisiones_sam/EditarEvaluaciones.html', )
        else:
            formA = NewReportsTests( instance = report) 
           
            #formB = NewDetReports( instance = detalle)    
            return render(request,'admisiones_sam/EditTest.html', {"formA":formA,"formB":formB,"lista":lista,"materias":materias,"id_test":id_test})
    else:
        return render(request,'admisiones_sam/EditTestMessage.html')
        
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def TestStudent(request):

    if request.method == "GET" and "q_aplicantes" in request.GET:
        aplicante=request.GET["q_aplicantes"]
     
     
        if aplicante:
       
            palabras = aplicante.split()
            aplicaciones=Applications.objects.get(reduce(operator.and_, (Q(name_student__icontains=x) | Q(surname_student__icontains=x) for x in palabras)))
            
            test=Test.objects.filter(id_applications_id=aplicaciones.id_applications).extra(select={
                'date_test':'''SELECT admisiones_sam_type_tests.date_test
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'grade':'''SELECT admisiones_sam_type_tests.grade
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'time_test':'''SELECT admisiones_sam_type_tests.time_test
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'teacher':'''SELECT admisiones_sam_type_tests.teacher
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'type_test':'''SELECT admisiones_sam_type_tests.type_test
                FROM admisiones_sam_type_tests
                WHERE admisiones_sam_type_tests.id_typetest=id_typetest_id
                ''',
                'reporte': '''select  admisiones_sam_reports_tests.id_report
                from admisiones_sam_reports_tests
                where admisiones_sam_reports_tests.id_test_id=id_test'''
                })
                
                
            return render(request,'admisiones_sam/TestStudent.html', {"test":test,"aplicante":aplicante})
    
    elif request.method == 'POST': 
        
        estados=request.POST.getlist('estado')
        for x in estados:
        
            if not x== None :
                evaluacion=Test.objects.get(id_test=x)
                evaluacion.state='1'
                evaluacion.save()
    else:
        return render(request,'admisiones_sam/TestStudent.html')


@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def EditTestStudent(request,id_test):


    evaluacion = Test.objects.get(id_test = id_test)
    aplicante=Applications.objects.get(id_applications=evaluacion.id_applications_id)
    estado=State_Applications.objects.get(id_applications=evaluacion.id_applications_id,final_date__isnull=True)
    
    if request.method == "POST":
        form = Testsnewform(request.POST, instance = evaluacion)    
        if form.is_valid():
            form.save()
            nuevo_estado = State_Applications(state = "Documentos", initial_date=date.today(), observations = "Asigna nueva Fecha Evaluacion", id_applications = aplicante)
            nuevo_estado.save()
            estado.final_date = date.today()
            estado.save()
            return redirect('/admisiones/EvaluacionesEstudiante')
            
        else:
            print form.errors
            return render(request,'admisiones_sam/EditTestStudent.html', {"form":form,"evaluacion":evaluacion})
    else:
        form = Testsnewform(instance = evaluacion)
        return render(request,'admisiones_sam/EditTestStudent.html', {"form":form,"evaluacion":evaluacion})  
  
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def ApproveTest(request):
    mensaje = MailerMessage()
    lista=['Aprobado','Reprobado','Lista Espera']
    test=Reports_Tests.objects.filter(~Q(state = 'Aprobado'),~Q(state = 'Reprobado')).extra(select={
                'aplicante':'''SELECT concat(admisiones_sam_applications.surname_student,' ',admisiones_sam_applications.name_student)
                FROM admisiones_sam_applications
                WHERE admisiones_sam_applications.id_applications=id_applications_id
                ''',
                'tutor':'''SELECT concat(admisiones_sam_applications.surname_tutor,' ',admisiones_sam_applications.name_tutor)
                FROM admisiones_sam_applications
                WHERE admisiones_sam_applications.id_applications=id_applications_id
                ''',
               'codigo_banco':'''SELECT admisiones_sam_applications.bank_code
                FROM admisiones_sam_applications
                WHERE admisiones_sam_applications.id_applications=id_applications_id
                ''',
                'grado': '''SELECT concat(admisiones_sam_type_grade.typegrade,' / ', admisiones_sam_type_grade.typeschool)
                FROM admisiones_sam_applications,admisiones_sam_type_grade
                WHERE admisiones_sam_applications.id_applications=id_applications_id 
                and admisiones_sam_applications.applied_grade=admisiones_sam_type_grade.typegrade''',
                'cupo':''' SELECT admisiones_sam_quotas.available_students
                FROM admisiones_sam_applications,admisiones_sam_quotas
                WHERE admisiones_sam_applications.id_applications=id_applications_id
                and admisiones_sam_applications.applied_grade=admisiones_sam_quotas.grade'''})
  
  
    if request.method == "POST":
        for key, value in request.POST.items():
            
            if value=='Aprobado' or value=='Reprobado' or value=='Lista Espera':
                resultados=Reports_Tests.objects.get(id_report=key)
                #fecha_inicial = aplicacion.tours_set.get(id_applications = aplicacion , estado = True)
                aplicante=Applications.objects.get(id_applications=resultados.id_applications_id)
                #aplicante=aplicacion.reports_tests_set.get(id_applications=aplicacion)
 
                #.extra(select={
                #'observaciones':'''SELECT observations
                #FROM admisiones_sam_reports_tests
                #WHERE admisiones_sam_reports_tests.id_applications_id=id_applications
                #and admisiones_sam_reports_tests.state='Pendiente'
                #'''})
                
                
                cupos=Quotas.objects.get(grade=aplicante.applied_grade)            
            
                
                if value=='Aprobado':    #and cupos.available_students>0:
                    resultados.state=value
                    conta = AccountingDepartment(state='False',id_applications=Applications.objects.get(id_applications=aplicante.id_applications))
                  
                    old_estado=State_Applications.objects.get(id_applications = aplicante.id_applications, final_date__isnull=True)
                    old_estado.final_date=date.today()
                    
                    new_estado = State_Applications(state='Contabilidad',final_date=None,observations='Registro en Contabilidad',id_applications=aplicante)
                    conta.save()
                    old_estado.save()
                    new_estado.save()
                    resultados.save()
                    
                    #if aplicante.applied_grade==u'Inicial 1 (de 2 a 3 años)' or aplicante.applied_grade==u'Inicial 2 (de 3 a 4 años)' or aplicante.applied_grade==u'Inicial 2 (de 4 a 5 años) / Prekinder' : #or aplicante.applied_grade==u'1ro de Básica / Kinder':
                     #   pass
                    #else:       
                    #mensaje.application_message(aplicante.pk,13)
                    mensaje.general_message(aplicante.mail_tutor, "admisiones@montebelloacademy.org", 13, aplicante)
                        
                    # Mensaje del sistema  a Admisiones a Contabilidad para que vuelva a activar el codigo del banco para que pague la matricula
                    mensaje.general_message('jeaneth@montebelloacademy.org','admisiones@montebelloacademy.org', 23, aplicante)
   
                    
                    
                elif value=='Reprobado':
                    resultados.state=value
                    resultados.save()
                    
                    #if aplicante.applied_grade==u'Inicial 1 (de 2 a 3 años)' or aplicante.applied_grade==u'Inicial 2 (de 3 a 4 años)' or aplicante.applied_grade==u'Inicial 2 (de 4 a 5 años) / Prekinder' or aplicante.applied_grade==u'1ro de Básica / Kinder':
                    #    pass
                    #else:     
                    #mensaje.application_message(aplicante.pk,14)
                    mensaje.general_message(aplicante.mail_tutor, "admisiones@montebelloacademy.org", 14, aplicante)
                else:
                    resultados.state=value
                    resultados.save()
                    
            else: 
                pass
                
        return render(request,'admisiones_sam/ApproveTest.html', {"test":test,"lista":lista})
    
        
    else:
        return render(request,'admisiones_sam/ApproveTest.html', {"test":test,"lista":lista})
    
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)  

def SummaryTest(request,id_typetest_id,id_report): 
    
    
    reporte=Reports_Tests.objects.filter(id_report=id_report)
    materias=[u'Inglés',u'Matemáticas',u'Lengua',u'Psicológicas',u'Ciencias',u'Evaluación Grupal',u'Evaluación Individual',u'Otros']

    #report=Reports_Tests.objects.filter(~Q(state = 'Aprobado'),~Q(state = 'Reprobado'),id_typetest_id=id_typetest_id)
    detalle=Det_Reports.objects.filter(id_report_id=id_report)
            
    DetReportFormSet=modelformset_factory(Det_Reports,fields=('materia', 'value','observations'), extra=0)
    formB=DetReportFormSet(queryset=detalle)
    
    tipo=Type_Tests.objects.filter(id_typetest=id_typetest_id,).extra(select={
                'oportunidad':'''SELECT opportunity
                FROM admisiones_sam_test, admisiones_sam_reports_tests
                WHERE admisiones_sam_test.id_test=admisiones_sam_reports_tests.id_test_id
                and admisiones_sam_test.id_typetest_id=id_typetest
                and admisiones_sam_reports_tests.id_test_id=id_report
                and admisiones_sam_test.state=1
                ''',
                'estado': '''SELECT admisiones_sam_test.state
                FROM admisiones_sam_test, admisiones_sam_reports_tests
                WHERE admisiones_sam_test.id_test=admisiones_sam_reports_tests.id_test_id
                and admisiones_sam_test.id_typetest_id=id_typetest
                and admisiones_sam_reports_tests.id_test_id=id_report
                and admisiones_sam_test.state=1'''})
    
    
    if request.method == "POST":
            
        formB=DetReportFormSet(request.POST)
        
        return render(request,'admisiones_sam/SummaryTest.html',{"tipo":tipo,"formB":formB,"materias":materias})
    else:
        return render(request,'admisiones_sam/SummaryTest.html',{"tipo":tipo,"formB":formB,"materias":materias})
    
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
  
def KeepValidaDocuments(request):
    
    documentos=Det_Documents.objects.filter(state=0).extra(select={
                'nombre':'''SELECT concat(admisiones_sam_applications.surname_student,' ',admisiones_sam_applications.name_student)
                FROM admisiones_sam_applications
                WHERE admisiones_sam_applications.id_applications=id_applications_id
                ''',
                'grado':'''SELECT admisiones_sam_applications.applied_grade
                FROM admisiones_sam_applications
                WHERE admisiones_sam_applications.id_applications=id_applications_id
                '''}).values('nombre','grado').distinct().order_by('nombre')
                
          
                
    return render(request, "admisiones_sam/KeepValidaDocuments.html", {"documentos":documentos})



    
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)  

def KeepDocuments(request):    

    
    documentos=Documents_Type.objects.all().order_by('id_doctype')
    consultagrados=Type_Grade.objects.all().order_by('id_typegrade')
    lista=[]
    
    for y in consultagrados:
        if y.typegrade in lista:
            pass
        else:
            lista.append(y.typegrade)
    
    tipos=[u'Cédula Identidad / Partida Nacimiento (escanear dos lados)',
            u'Comprobante Pago / Transferencia',
            u'Reporte de Calificaciones y Comportamiento Año en curso',
            u'Reporte de Calificaciones y Comportamiento Año Anterior']
    if request.method == "POST":
        form = NewDocuments_Type(request.POST)
        if form.is_valid():
            item=form.save(commit=False)
            grado=item.grade
            documento=item.document
            if not Documents_Type.objects.filter(grade=grado,document=documento).count()>0:
                item.save()
                return redirect('/admisiones/MantenerDocumentos')
                
            else:
                return render(request,"admisiones_sam/DocumentsNewMessage.html",{"grado":grado,"documento":documento})
           
        else:
            print form.errors
            return render(request, "admisiones_sam/KeepDocuments.html", {"form":form,"documentos":documentos,"lista":lista,"tipos":tipos})
    else:
        
        form = NewDocuments_Type()
        return render(request, "admisiones_sam/KeepDocuments.html", {"form":form,"documentos":documentos,"lista":lista,"tipos":tipos})

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def DeleteDocuments(request, id_doctype): 
    documents = Documents_Type.objects.get(id_doctype = id_doctype)
    documents.delete()
    return redirect('/admisiones/MantenerDocumentos')
   
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def EditDocuments(request, id_doctype):
  
    documentos = Documents_Type.objects.get(id_doctype = id_doctype)
    consultagrados=Type_Grade.objects.all().order_by('id_typegrade')
    lista=[]
    
    for y in consultagrados:
        if y.typegrade in lista:
            pass
        else:
            lista.append(y.typegrade)
    
    tipos=[u'Cédula Identidad / Partida Nacimiento (escanear dos lados)',
            u'Comprobante Pago / Transferencia',
            u'Reporte de Calificaciones y Comportamiento Año en curso',
            u'Reporte de Calificaciones y Comportamiento Año Anterior']
    
    if request.method == "POST":
        form = NewDocuments_Type(request.POST, instance = documentos)   
        if form.is_valid():
            form.save()
            return redirect('/admisiones/MantenerDocumentos')
        else:
            print form.errors
            return render(request,'admisiones_sam/EditDocuments.html', {"form":form,"documentos":documentos,"lista":lista,"tipos":tipos})
    else:
        form = NewDocuments_Type(instance = documentos)
        return render(request,'admisiones_sam/EditDocuments.html', {"form":form,"documentos":documentos,"lista":lista,"tipos":tipos})  


def EditTypeApplicant(request, id_applications):
    aplicante = Applications.objects.get(pk = id_applications)
    if aplicante.type_student == "NEW":
        aplicante.type_student = "OLD"
        aplicante.save()
        
        estado=State_Applications.objects.get(id_applications = id_applications, final_date__isnull = True)
        nuevo_estado = State_Applications(state = "Error", initial_date=date.today(), observations = "Aplicación Hermanos", id_applications = aplicante)
        nuevo_estado.save()
        estado.final_date = date.today()
        estado.save()
        
    elif aplicante.type_student == "OLD":
        aplicante.type_student = "NEW"
        aplicante.save()
    return HttpResponseRedirect(request.META.get('HTTP_REFERER','/')) 


def ListUploadDocuments(request):
    aplicaciones = Applications.objects.filter(state_applications__state = "Documentos", state_applications__final_date = None).order_by('surname_student')
    sin_pendientes = []
    
    for x in aplicaciones:


        doc_pedidos = x.det_documents_set.all()

        if doc_pedidos.count() == 0:

            continue

        doc_pendientes = x.det_documents_set.filter(state = False)

        if doc_pendientes.count() == 0:

            sin_pendientes.append(x.pk)

    #return sin_pendientes
    aplicaciones = aplicaciones.exclude(pk__in = sin_pendientes)


    return render(request, 'admisiones_sam/ListUploadDocuments.html', {"aplicaciones":aplicaciones})




def UploadDocuments(request,id_applications):
    mensaje = MailerMessage()
    aplicacion=Applications.objects.get(id_applications=id_applications)
    documentos=Documents_Type.objects.filter(grade=aplicacion.applied_grade,estado=True).order_by()
    contador=documentos.count()
    
    DocumentsFormSet=modelformset_factory(Det_Documents,fields=('file','id_doctype',), extra=contador)
    formB=DocumentsFormSet(queryset=Det_Documents.objects.none(), initial=[{'id_doctype': x.pk} for x in documentos] )
    detalle=Det_Documents.objects.filter(id_applications = aplicacion.id_applications)
    n=0
    
    debug='1'
    if request.method == "POST":
        debug='2'
        if not detalle.count()>0:
            formB=DocumentsFormSet(request.POST, request.FILES)
            
            if formB.is_valid():
                debug='3'        
                for x in formB:
                    
                    debug='4'
                    new_item = x.save(commit=False)
                    new_item.id_doctype= documentos[n]
                    new_item.state = False
                    new_item.id_applications_id = aplicacion.id_applications
                   
                    n=n+1
                   
                    if not new_item.file== None:
                        #debug='5'
                        #try:
                        new_item.save()
                         #   debug='6'
                        #except:
                        #    pass
                            #debug='7'
                    else:
                        debug='8'
                        print formB.errors
                        
                        detalle = Det_Documents.objects.filter(id_applications = aplicacion.id_applications)
                        for y in detalle:
                            y.delete()
                        modelo='CargaDocumentos'
                        return render(request,'admisiones_sam/UploadNewMessage.html', {"aplicante":aplicacion,"modelo":modelo})
                    
                    
                
              
               
               # Mensaje del sistema  a Admisiones que el representante cargo los documentos
                mensaje.general_message('sorayamena@montebelloacademy.org','admisiones@montebelloacademy.org', 21, aplicacion)
                 
                return render(request, "admisiones_sam/InvoicesMessage.html", {"aplicante":aplicacion})
                               
            else:
                debug='9'
                print formB.errors
                return render(request,'admisiones_sam/UploadDocuments.html', {"form":formB,"documentos":documentos,"aplicante":aplicacion}) 
        else:
            return render(request,'admisiones_sam/UploadMessageExist.html', {"aplicante":aplicacion}) 
    
    else:
        debug=aplicacion.id_applications
        return render(request,'admisiones_sam/UploadDocuments.html', {"form":formB,"documentos":documentos,"aplicante":aplicacion})        


def UploadEditDocuments (request,id_applications):

    aplicacion=Applications.objects.get(id_applications=id_applications)
    documentos=Documents_Type.objects.filter(grade=aplicacion.applied_grade,estado=True).order_by()
    detalle=Det_Documents.objects.filter(id_applications = aplicacion.id_applications,state=False)
    contador=detalle.count()
    
    DocumentsFormSet=modelformset_factory(Det_Documents,fields=('file','id_doctype','observations',), extra=0)
    formB=DocumentsFormSet(queryset=detalle, initial=[{'id_doctype': x.pk} for x in detalle] )
    
    if request.method == "POST":
        formB=DocumentsFormSet(request.POST, request.FILES)
        if formB.is_valid():
            
            llaves = list(request.POST.keys())
            for x in llaves:
                if request.POST[x] == "":
                    print formB.errors
                    modelo='CambiaDocumentos'
                    return render(request,'admisiones_sam/UploadNewMessage3.html', {"aplicante":aplicacion.id_applications,"modelo":modelo})
            
                else:
                    for y in formB:
                        #path=y.as_table()
                        #os.remove(path)

                        new_item = y.save(commit=False)
                        new_item.save()
                    
            return render(request,'admisiones_sam/UploadMessage.html', {"Aplicacion":aplicacion,"path":formB})
        else:
            print formB.errors
            return render(request,'admisiones_sam/UploadEditDocuments.html', {"form":formB,"documentos":detalle,"aplicante":aplicacion}) 
    else:
        
        return render(request,'admisiones_sam/UploadEditDocuments.html', {"form":formB,"aplicante":aplicacion,"documentos":detalle})        

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def CheckDocuments(request):
    
    DocumentsFormSet=modelformset_factory(Det_Documents,fields=('file','id_doctype','state','observations',), extra=0)
    formB=DocumentsFormSet(queryset=Det_Documents.objects.none() )
    detalle=None
    
    estudiante=request.GET.get('q_estudiante', '')  #["q_estudiante"]
    #aplicacion=Applications.objects.get(id_applications=id_applications)
    
    
    
    if request.method == "GET" and "q_aplicantes" in request.GET:
        aplicante=request.GET["q_aplicantes"]
     
       
        if aplicante:
            
            palabras = aplicante.split()
 
            aplicaciones=Applications.objects.get(reduce(operator.and_, (Q(name_student__icontains=x) | Q(surname_student__icontains=x) for x in palabras)))
            for estado in aplicaciones.state_applications_set.filter(final_date__isnull=True):
                if estado.state is not "Terminado" :
                    
                    debug=aplicaciones
                  
                    detalle=Det_Documents.objects.filter(id_applications_id = aplicaciones)
                   
                    DocumentsFormSet=modelformset_factory(Det_Documents,fields=('file','id_doctype','state','observations',), extra=0)
                    formB=DocumentsFormSet(queryset=detalle, initial=[{'id_doctype': x.pk} for x in detalle] )
                   
                    return render(request,"admisiones_sam/CheckDocuments.html",{"form":formB,"detalle":detalle,"aplicante":aplicante,"debug":debug,"aplicaciones":aplicaciones})
                else:
                    detalle=None
                    return render(request,"admisiones_sam/CheckDocuments.html",{"form":formB,"detalle":detalle,"aplicante":aplicante,"debug":debug,"aplicaciones":aplicaciones})
     
        else:
        
           return render(request,"admisiones_sam/CheckDocuments.html",{"form":formB,"detalle":detalle})
           
    elif request.method == "POST":
        #return redirect('NuevaSolicitud')
     
        #documentos=request.POST.getlist('observaciones')
        for x in request.POST:
            prueba=x
            if not x== None :
                if x=='csrfmiddlewaretoken':
                    pass
                else:
                    observaciones=request.POST[x]
                    detalle_observaciones=Det_Documents.objects.get(id_detdoc=x)
                    detalle_observaciones.observations=observaciones
                    detalle_observaciones.save()
                
        return render(request,"admisiones_sam/CheckDocuments.html",{"form":formB})        
        
    else:
        return render(request,"admisiones_sam/CheckDocuments.html",{"form":formB,"aplicante":estudiante}) #"aplicante":aplicacion.surname_student+ " " + aplicacion.name_student})



def ViewDocuments(request):
    
    DocumentsFormSet=modelformset_factory(Det_Documents,fields=('file','id_doctype','state','observations',), extra=0)
    formB=DocumentsFormSet(queryset=Det_Documents.objects.none() )
    detalle=None
    
    if request.method == "GET" and "q_aplicantes" in request.GET:
        aplicante=request.GET["q_aplicantes"]
     
        
        if aplicante:
       
            palabras = aplicante.split()
            aplicaciones=Applications.objects.get(reduce(operator.and_, (Q(name_student__icontains=x) | Q(surname_student__icontains=x) for x in palabras)))
            
            detalle=Det_Documents.objects.filter(id_applications = aplicaciones.id_applications)
            
            DocumentsFormSet=modelformset_factory(Det_Documents,fields=('file','id_doctype','state','observations',), extra=0)
            formB=DocumentsFormSet(queryset=detalle, initial=[{'id_doctype': x.pk} for x in detalle] )
            #formB=NewDetDocuments(instance=detalle)
            return render(request,"admisiones_sam/ViewDocuments.html",{"form":formB,"detalle":detalle,"aplicante":aplicante})
           
        else:
           return render(request,"admisiones_sam/ViewDocuments.html",{"form":formB,"detalle":detalle})
           
    elif request.method == "POST":
        #return redirect('NuevaSolicitud')
     
        #documentos=request.POST.getlist('observaciones')
        for x in request.POST:
            prueba=x
            if not x== None :
                if x=='csrfmiddlewaretoken':
                    pass
                else:
                    observaciones=request.POST[x]
                    detalle_observaciones=Det_Documents.objects.get(id_detdoc=x)
                    detalle_observaciones.observations=observaciones
                    detalle_observaciones.save()
                
        return render(request,"admisiones_sam/ViewDocuments.html",{"form":formB})        
        
    else:
        return render(request,"admisiones_sam/ViewDocuments.html",{"form":formB})
        
        


def ChangeDocuments(request,id_detdoc):
    mensaje = MailerMessage()
    actualizo=id_detdoc
    debug='1'
    n=0
    c=0
    
    
    if actualizo:
        debug='2'
       
        detalle=Det_Documents.objects.get(id_detdoc = actualizo)
        aplicante=Applications.objects.get(id_applications=detalle.id_applications_id)
       
    
        try:
            pago_evaluacion=AccountingDepartmentTest.objects.get(id_applications_id=aplicante.id_applications,state=1)
        except AccountingDepartmentTest.DoesNotExist:
            pago_evaluacion=None
            
        if detalle.state==False:
            debug='3'
            detalle.state=True
            detalle.save()
            #return render(request,"admisiones_sam/CheckDocuments.html",{"debug":debug})
            documentos=Det_Documents.objects.filter(id_applications=detalle.id_applications_id)
    
            for x in documentos:
                n=n+1 #numero de documentos cargados
                if x.state == True:
                    c=c+1  #numero de documentos validados
            if n==c :   # si los dos son iguales se han validado todos los documentos.
                
                # Mensaje del sistema  a los representantes para informar que los documentos se validaron correctamente.
                #mensaje.application_message(aplicante.pk,28)
                mensaje.general_message(aplicante.mail_tutor, "admisiones@montebelloacademy.org", 28, aplicante)
                
                # Si se realizo el pago de las evaluaciones y se validaron los documentos mensaje a Admisiones para que agende fecha de evaluaciones.
                if pago_evaluacion != None:
                    mensaje.general_message('sorayamena@montebelloacademy.org','admisiones@montebelloacademy.org', 30, aplicante)
                else:
                    pass
                
                
            else:
                pass
            
            return HttpResponseRedirect(request.META.get('HTTP_REFERER','/')) 
        else:
            debug='4'
            detalle.state=False
            detalle.save()
            return HttpResponseRedirect(request.META.get('HTTP_REFERER','/'))
            #return render(request,"admisiones_sam/CheckDocuments.html",{"debug":debug})
        
        
    else:
        debug='5'
        return HttpResponseRedirect(request.META.get('HTTP_REFERER','/'))
        #return render(request,"admisiones_sam/CheckDocuments.html",{"debug":debug})
        

def SendDocuments(request):

    documentos=Applications.objects.filter(id_applications__in=Det_Documents.objects.filter(state=False).values('id_applications_id').distinct()).extra(select={
            'grade':'''select admisiones_sam_type_grade.typeschool 
            from admisiones_sam_type_grade
            where admisiones_sam_type_grade.typegrade =applied_grade  ''',}).order_by('surname_student')
    
    
    return render(request,"admisiones_sam/SendDocuments.html",{"documentos":documentos})


def MailChangeDocuments(request,id_applications):
    mensaje = MailerMessage()
    aplicacion=Applications.objects.get(id_applications=id_applications)
    #mensaje.application_message(id_applications,8) #envio de mail a representantes para volver a cambiar los documentos de ingreso admisiones 
    mensaje.general_message(aplicacion.mail_tutor, "admisiones@montebelloacademy.org", 8, aplicacion)
    return HttpResponseRedirect(request.META.get('HTTP_REFERER','/'))
    
    
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def Administration (request):
    
    consultaestados=State_Applications.objects.all().distinct()
    estados=[]
    for xx in consultaestados:
        if xx.state in estados:
            pass
        else:
            estados.append(xx.state)
  
    consultafechastour=Type_Tours.objects.filter(state=True).order_by('date_tour').distinct()
    fechas=['Otra Fecha']
    
    for y in consultafechastour:
        if y.date_tour.strftime("%Y-%m-%d") in fechas:
            pass
        else:
            fechas.append(y.date_tour.strftime("%Y-%m-%d"))
    
 
    consultafechastest=Type_Tests.objects.filter(state=True).order_by('date_test').distinct()
    #.order_by('date_test')
    fechast=[]
    
    for z in consultafechastest:
        if z.date_test.strftime("%Y-%m-%d") in fechast:
            pass
        else:
            fechast.append(z.date_test.strftime("%Y-%m-%d"))
    
    solicitud=Applications.objects.extra(select={
            'grade':'''SELECT admisiones_sam_type_grade.typeschool 
            from admisiones_sam_type_grade 
            where admisiones_sam_type_grade.typegrade = applied_grade''',
            'fecha_tour':'''SELECT admisiones_sam_type_tours.date_tour
            FROM admisiones_sam_tours,admisiones_sam_type_tours
            WHERE admisiones_sam_tours.id_typetour_id=admisiones_sam_type_tours.id_typetour
            and id_applications_id=id_applications 
            and admisiones_sam_tours.estado='1'
             ''',
             'fecha_pruebas':'''SELECT admisiones_sam_type_tests.date_test
            FROM admisiones_sam_test,admisiones_sam_type_tests
            WHERE admisiones_sam_test.id_typetest_id=admisiones_sam_type_tests.id_typetest
            and id_applications_id=id_applications
            and admisiones_sam_test.state='1' ''',
            'fecha_resultados':'''SELECT date_report
            FROM admisiones_sam_reports_tests,admisiones_sam_test
            WHERE admisiones_sam_reports_tests.id_test_id=admisiones_sam_test.id_test
            and admisiones_sam_reports_tests.id_applications_id=id_applications
            and admisiones_sam_test.state=1''',
            'resultado':'''SELECT admisiones_sam_reports_tests.state
            from admisiones_sam_reports_tests,admisiones_sam_test
            WHERE admisiones_sam_reports_tests.id_test_id=admisiones_sam_test.id_test
            and admisiones_sam_reports_tests.id_applications_id=id_applications
            and admisiones_sam_test.state=1''',
            'estado_solicitud':'''SELECT state
            FROM admisiones_sam_state_applications
            WHERE final_date is null
            AND id_applications_id=id_applications''',
            'cuota':''' SELECT  admisiones_sam_quotas.available_students
            FROM admisiones_sam_quotas
            WHERE admisiones_sam_quotas.grade= applied_grade''',
            'pago':'''SELECT state
            FROM admisiones_sam_accountingdepartmenttest
            WHERE admisiones_sam_accountingdepartmenttest.id_applications_id=id_applications
            ''',

            }).order_by('-id_applications')#(Lower('surname_student'))
    #return solicitud
        
    #consulta para validar los aspirantes por grado#
            
    aplicantes=Applications.objects.values('applied_grade').order_by().annotate(Count('applied_grade'))
    #estados_grados=State_Applications.objects.filter(final_date__isnull=True).filter(~Q(state='Terminado')).values_list('id_applications',flat=True)
    estados_grados=State_Applications.objects.filter(final_date__isnull=True).exclude(state='Terminado').values_list('id_applications',flat=True)
    aplicado=aplicantes.filter(id_applications__in=estados_grados)

    documentos = Documents_Type.objects.values('grade').order_by().annotate(Count('grade'))

    validados = Det_Documents.objects.filter(state = True).values("id_applications__pk").order_by().annotate(Count('id_applications'))
    periodo_lectivo=PeriodSchool.objects.filter(per_state=True)
    #return estados_grados


    grados_sol=[]
    for x in solicitud:
        if x.grade in grados_sol:
            pass
        else:
            grados_sol.append(x.grade)

    grados_sys = Type_Grade.objects.all().order_by('pk').values_list('typeschool', flat = True)
    grados = []
    for x in grados_sys:
        if x in grados_sol:
            grados.append(x)
        else:
            pass

            
    if request.method == "GET" and "q_grado" in request.GET:
    
        estado=request.GET["q_estado"]
        aplicante=request.GET["q_aplicante"]
        
        grado=request.GET["q_grado"]
        fecha=request.GET["q_fecha"]
        fechat=request.GET["q_fechat"]
        periodo=request.GET["q_school_period"]
           
        #### problemas en la cconsulta para varios periodods lectivos
        if grado:
            grado_sys = Type_Grade.objects.get(typeschool = grado)   
            aplicantes=solicitud.filter(applied_grade=grado_sys.typegrade)

            #aplicantes=solicitud.extra(where=['grade = grado'])
            return render(request,'admisiones_sam/Administration.html',{"aplicantes":aplicantes,"aplicado":aplicado, "documentos":documentos, "validados":validados})
        
        elif aplicante:
            palabras = aplicante.split()
            aplicantes=solicitud.filter(reduce(operator.and_, (Q(surname_student__icontains=x) | Q(name_student__icontains=x) for x in palabras)))
            #aplicantes=solicitud.filter(name_student__icontains=aplicante)
            return render(request,'admisiones_sam/Administration.html',{"aplicantes":aplicantes,"aplicado":aplicado, "documentos":documentos, "validados":validados})
        
        elif estado:
            aplicantes = solicitud.filter(state_applications__state = estado, state_applications__final_date = None)
            #aplicantes=solicitud.extra(where=["estado_solicitud = %s"], params=[estado])
            return render(request,'admisiones_sam/Administration.html',{"aplicantes":aplicantes,"aplicado":aplicado, "documentos":documentos, "validados":validados})
        
        elif fecha:
            if fecha=='Otra Fecha':

                aplicantes = solicitud.filter(tours = None)
                
                #aplicantes=solicitud.extra(where=["fecha_tour is null"])
                return render(request,'admisiones_sam/Administration.html',{"aplicantes":aplicantes,"aplicado":aplicado, "documentos":documentos, "validados":validados})
            else:
                #tour_sys = Type_Tours.objects.get(date_tour = fecha)
                #tour_sys2= Tours.objects.get(id_typetour=tour_sys)
                aplicantes = solicitud.filter(tours__id_typetour__date_tour = fecha,tours__estado=True)
                #aplicantes=solicitud.extra(where=["fecha_tour = %s"], params=[fecha])
                return render(request,'admisiones_sam/Administration.html',{"aplicantes":aplicantes,"aplicado":aplicado, "documentos":documentos, "validados":validados})
        elif fechat:

            aplicantes = solicitud.filter(test__id_typetest__date_test = fechat, test__id_typetest__state = True).distinct()
            #aplicantes=solicitud.extra(where=["fecha_pruebas = %s"], params=[fechat])
            return render(request,'admisiones_sam/Administration.html',{"aplicantes":aplicantes,"aplicado":aplicado, "documentos":documentos, "validados":validados})
        
        elif periodo:
            periodos=PeriodSchool.objects.get(per_name = periodo)
            aplicantes = solicitud.filter(school_period=periodos.per_id)
            return render(request,'admisiones_sam/Administration.html',{"aplicantes":aplicantes,"aplicado":aplicado, "documentos":documentos, "validados":validados})
        else:
            solicitud = solicitud.exclude(state_applications__state = "Terminado", state_applications__final_date = None)
            #solicitud = solicitud.extra(where=["estado_solicitud is not 'Terminado'"])
            return render(request,'admisiones_sam/Administration.html',{"aplicantes":solicitud,"estados":estados,"grados":grados,"fechas":fechas,"aplicado":aplicado, "documentos":documentos, "validados":validados,"periodo_lectivo":periodo_lectivo})
    else:


        solicitud = solicitud.exclude(state_applications__state =  "Terminado", state_applications__final_date = None)
        #solicitud = solicitud.extra(where=["estado_solicitud is not 'Terminado'"])
        return render(request,'admisiones_sam/Administration.html',{"aplicantes":solicitud,"estados":estados,"grados":grados,"fechas":fechas,"fechast":fechast,"aplicado":aplicado, "documentos":documentos, "validados":validados,"periodo_lectivo":periodo_lectivo})
    
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def ApproveAccounting  (request):
    mensaje = MailerMessage()
    aplicantes=AccountingDepartment.objects.filter(~Q(state=True)).extra(select={
            'nombre':'''SELECT concat(admisiones_sam_applications.surname_student,' ',admisiones_sam_applications.name_student)
            FROM admisiones_sam_applications
            WHERE admisiones_sam_applications.id_applications=id_applications_id
            ''',
            'grade':'''SELECT admisiones_sam_applications.applied_grade
            FROM admisiones_sam_applications
            WHERE admisiones_sam_applications.id_applications=id_applications_id
            ''',}).extra(order_by=['nombre'])
            
   
    if request.method == 'POST': 
        conta=request.POST.getlist('estado')
        for x in conta:
        
            if not x== None :
                valor="observaciones_"+ str( x)
                observa=request.POST.get(valor)
                
                estado=AccountingDepartment.objects.get(id_accounting=x)
                estado.state='1'
                estado.observations=observa
                estado.save()
                
                #Valida los cupos
                aplicante=Applications.objects.get(id_applications=estado.id_applications_id)
                cupos=Quotas.objects.get(grade=aplicante.applied_grade)
                
                if cupos.old_students==cupos.max_students: #no hay cupos
                    return render(request,"admisiones_sam/QuotasMessage.html",{"nivel":aplicante.applied_grade})    
                else:
                    
                    
                    #Ingreso el registro en secretaria
                    
                    
                    secretaria = AcademicSecretary(id_applications=aplicante,state='0')
                    secretaria.save()
                    old_estado=State_Applications.objects.get(id_applications = aplicante.id_applications, final_date__isnull=True)
                    old_estado.final_date=date.today()
                    old_estado.save()
                    new_estado = State_Applications(state='Secretaria',final_date=None,observations='Registro en Secretaria',id_applications=aplicante)
                    new_estado.save()
                    
                    cupos.old_students=cupos.old_students+1
                    cupos.available_students=cupos.available_students-1
                    cupos.save()
                    
                    #mensaje.application_message(aplicante.pk,15)
                    mensaje.general_message(aplicante.mail_tutor, "admisiones@montebelloacademy.org", 15, aplicante)
                    
                    # Mensaje del sistema desde contabilidad a admisiones informando el pago de la matricula
                    mensaje.general_message('sorayamena@montebelloacademy.org','admisiones@montebelloacademy.org', 24, aplicante)
                    # Mensaje del sistema desde contabilidad a secretaria informando el pago de la matricula y la generacion del formulario de infromacion
                    mensaje.general_message('marysilva@montebelloacademy.org','admisiones@montebelloacademy.org', 25, aplicante)
   
                #fin de validación
      
        aplicantes=AccountingDepartment.objects.filter(~Q(state=True)).extra(select={
        'nombre':'''SELECT concat(admisiones_sam_applications.surname_student,' ',admisiones_sam_applications.name_student)
        FROM admisiones_sam_applications
        WHERE admisiones_sam_applications.id_applications=id_applications_id
        ''',
        'grade':'''SELECT admisiones_sam_applications.applied_grade
        FROM admisiones_sam_applications
        WHERE admisiones_sam_applications.id_applications=id_applications_id
        ''',}).extra(order_by=['nombre'])
            
        return render(request,"admisiones_sam/ApproveAccounting.html",{"aplicantes":aplicantes,"conta":estado})    
    else:          
        return render(request,"admisiones_sam/ApproveAccounting.html",{"aplicantes":aplicantes})

        
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def ApproveAccountingTest  (request):
    mensaje = MailerMessage()
    n=0
    c=0
    aplicacion=None
    aplicantes=AccountingDepartmentTest.objects.filter(state=0).extra(select={
            'nombre':'''SELECT concat(admisiones_sam_applications.surname_student,' ',admisiones_sam_applications.name_student)
            FROM admisiones_sam_applications
            WHERE admisiones_sam_applications.id_applications=id_applications_id
            ''',
            'grade':'''SELECT admisiones_sam_applications.applied_grade
            FROM admisiones_sam_applications
            WHERE admisiones_sam_applications.id_applications=id_applications_id
            ''',}).extra(order_by=['nombre'])
            
            
     
    if request.method == 'POST': 
        conta=request.POST.getlist('estado')
        
        for x in conta:
            
            if not x== None :
                n=0
                c=0
                
                apli="id_aplicacion_"+ str( x)
                aplicacion=request.POST.get(apli)
                               
                documentos=Det_Documents.objects.filter(id_applications=aplicacion)
                aplicante=Applications.objects.get(id_applications=aplicacion)
                
                valor="observaciones_"+ str( x)
                observa=request.POST.get(valor)
               
                
                estado=AccountingDepartmentTest.objects.get(id_accountingtest=x)
                estado.state='1'
                estado.observations=observa
                estado.save()

                for y in documentos:
                    n=n+1 #numero de documentos cargados
                    if y.state == True:
                        c=c+1  #numero de documentos validados
                if n==c :   # si los dos son iguales se han validado todos los documentos.
                    # Si se realizo el pago de las evaluaciones y se validaron los documentos mensaje a Admisiones para que agende fecha de evaluaciones.
                    mensaje.general_message('admisiones@montebelloacademy.org','admisiones@montebelloacademy.org', 30, apli)
                else:
                    # Mensaje del sistema a admisiones informando solo el pago de las evaluaciones 
                    mensaje.general_message('sorayamena@montebelloacademy.org','admisiones@montebelloacademy.org', 29, apli)
                    
                    
             
        aplicantes=AccountingDepartmentTest.objects.filter(state=0).extra(select={
        'nombre':'''SELECT concat(admisiones_sam_applications.name_student,' ',admisiones_sam_applications.surname_student)
        FROM admisiones_sam_applications
        WHERE admisiones_sam_applications.id_applications=id_applications_id
        ''',
        'grade':'''SELECT admisiones_sam_applications.applied_grade
        FROM admisiones_sam_applications
        WHERE admisiones_sam_applications.id_applications=id_applications_id
        ''',}).extra(order_by=['nombre'])
        #order_by(Lower('id_applications_id'))
        return render(request,"admisiones_sam/ApproveAccountingTest.html",{"aplicantes":aplicantes}) 
                   
    else:          
        return render(request,"admisiones_sam/ApproveAccountingTest.html",{"aplicantes":aplicantes})
        
        
@login_required(login_url = '/user/login/')
@user_passes_test(is_admisiones)  
def ApproveSecretary(request): 
    aplicantes= AcademicSecretary.objects.filter(state='1').extra(select={
            'name':'''SELECT concat(admisiones_sam_applications.name_student,' ',admisiones_sam_applications.surname_student)
            FROM admisiones_sam_applications
            WHERE admisiones_sam_applications.id_applications=id_applications_id
            ''',
            'grade':'''SELECT admisiones_sam_applications.applied_grade
            FROM admisiones_sam_applications
            WHERE admisiones_sam_applications.id_applications=id_applications_id
            ''',}).order_by(Lower('id_applications_id'))
    
    


    if request.method == 'POST': 
        secre=request.POST.getlist('estado')
        
        for x in secre:
        
            if not x== None :
                valor="observaciones_"+ str( x)
                observa=request.POST.get(valor)
                
                estado=AcademicSecretary.objects.get(id_secretary=x)
                estado.state='2'
                estado.observations=observa
                estado.save()
                
                #Ingreso el estado a Finalizado
                aplicante=Applications.objects.get(id_applications=estado.id_applications_id)
                
                old_estado=State_Applications.objects.get(id_applications = aplicante.id_applications, final_date__isnull=True)
                old_estado.final_date=date.today()
                old_estado.save()
                new_estado = State_Applications(state='Finalizado',final_date=None,observations='Proceso Concluido',id_applications=aplicante)
                new_estado.save()

                estudiante = Students.objects.get(id_applications = aplicante)
                estudiante = Student.objects.get(user = estudiante.id_students)

                try:
                    periodo_lectivo = Periodo_Lectivo.objects.get(inicio__year = aplicante.school_period.per_startdate.strftime("%Y"), fin__year = aplicante.school_period.per_enddate.strftime("%Y"))
                except Periodo_Lectivo.DoesNotExist:
                    periodo_lectivo = Periodo_Lectivo.objects.create(inicio = aplicante.school_period.per_startdate, fin = aplicante.school_period.per_enddate)
                try:
                    nivel = Nivel.objects.get(nombre_aplicaciones = aplicante.applied_grade, activo = True)

                except Nivel.DoesNotExist:
                    
                    try:
                        periodo_actual = Periodo_Lectivo.objects.get(actual = True)
                        clase = Clase.objects.get(periodo_lectivo = periodo_actual, nivel__nombre_admisiones = aplicante.applied_grade)
                        nivel = clase.nivel
                    except:
                        nivel = Nivel.objects.create(nombre= aplicante.applied_grade, nombre_ministerio= aplicante.applied_grade, nombre_aplicaciones = aplicante.applied_grade, activo = True)

                Aptitud_Matricula.objects.create(estudiante = estudiante, nivel = nivel, periodo_lectivo = periodo_lectivo, tipo = "N")     
                
            else:          
                return render(request,"admisiones_sam/ApproveSecretary.html",{"aplicantes":aplicantes}) 
        return render(request,"admisiones_sam/ApproveSecretary.html",{"aplicantes":aplicantes,"conta":estado,"secre":secre})    
    else:          
        return render(request,"admisiones_sam/ApproveSecretary.html",{"aplicantes":aplicantes}) 
   
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def InvoiceMessageSearch (request):
    try:
        id_invoice=request.session['id_invoice']  
    except:
        return redirect('MantenerFacturas')
        
    formulario=Invoices.objects.get(id_invoice=id_invoice)
    
    del request.session["id_invoice"]
    return render(request,"admisiones_sam/InvoicesSearchMessage.html", {"formulario":formulario})
    

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def FormsMessageSave (request):
    try:
        id_forms=request.session['id_forms']  
    except:
        return redirect('EditaFormulario')
        
    formulario=Forms.objects.get(id_forms=id_forms)
    
    del request.session["id_forms"]
    return render(request,"admisiones_sam/FormsMessageSave.html", {"formulario":formulario})
    

def ReturnHome(request):

    return redirect('http://www.montebelloacademy.org')
        

def AppMessage (request):
    try:
        id_formulario=request.session['id_formulario']  
    except:
        return redirect('NuevaSolicitud')
        
    aplicacion=Applications.objects.filter(id_applications=id_formulario).extra(select={'date':'''
                select  admisiones_sam_type_tours.date_tour
                from admisiones_sam_type_tours, admisiones_sam_tours
                where admisiones_sam_tours.id_typetour_id=admisiones_sam_type_tours.id_typetour
                and admisiones_sam_tours.id_applications_id=id_applications
                and admisiones_sam_tours.estado='1'  ''',})
    del request.session["id_formulario"]
    return render(request,"admisiones_sam/AppMessage.html", {"Aplicacion":aplicacion})
    

def AppMessageCH (request):
    try:
        id_formulario=request.session['id_formulario']  
    except:
        return redirect('NuevaSolicitudHermanos')
        
    aplicacion=Applications.objects.filter(id_applications=id_formulario).extra(select={'date':'''
                select  admisiones_sam_type_tours.date_tour
                from admisiones_sam_type_tours, admisiones_sam_tours
                where admisiones_sam_tours.id_typetour_id=admisiones_sam_type_tours.id_typetour
                and admisiones_sam_tours.id_applications_id=id_applications
                and admisiones_sam_tours.estado='1'  ''',})
                                
    del request.session["id_formulario"]
    return render(request,"admisiones_sam/AppMessageCH.html", {"Aplicacion":aplicacion})
    
    

def AppMessageState (request):
    try:
        id_formulario=request.session['id_formulario']  
    except:
        return redirect('NuevaSolicitud')
        
    aplicacion=Applications.objects.get(id_applications=id_formulario)
   
    del request.session["id_formulario"]
    return render(request,"admisiones_sam/AppMessageState.html", {"Aplicacion":aplicacion})
    

def ListApplicant(request):
    
    if request.is_ajax():
        results = []
        q = request.GET.get('term', '')
        q = q.upper()
        palabras = q.split()
        
        aplicantes = Applications.objects.filter(reduce(operator.and_, (Q(surname_student__icontains=x) | Q(name_student__icontains=x) for x in palabras)))[:20]


        for x in aplicantes:
            x_json = {}
            x_json["value"] = x.surname_student + " " + x.name_student
            if x_json in results:
                pass
            else:
                results.append(x_json)
        data = json.dumps(results)
    else:
        data = 'fail'
        
    mimetype = 'application/json'
    return HttpResponse(data, mimetype)
    

def ListStudent(request):
    
    if request.is_ajax():
        q = request.GET.get('term', '')
        palabras = q.split()
        
        
        estudiantes = Forms.objects.filter(reduce(operator.or_, (Q(name_student__icontains=x) | Q(surname_student__icontains=x) | Q(identity_student__icontains=x)  for x in palabras)))[:20]
        results = []

        for x in estudiantes:
            x_json = {}
            x_json["value"] =  x.name_student + " " + x.surname_student + " " + str(x.identity_student)
            if x_json in results:
                pass
            else:
                results.append(x_json)
        data = json.dumps(results)
    else:
        data = 'fail'
        
    mimetype = 'application/json'
    return HttpResponse(data, mimetype)


def ListUserStudents(request):
    
    if request.is_ajax():
        q = request.GET.get('term', '')
        palabras = q.split()
       
        usuarios = Students.objects.filter(reduce(operator.and_, (Q(id_students__first_name__icontains=x) | Q(id_students__father_last_name__icontains=x) | Q(id_students__mother_last_name__icontains=x)| Q(id_students__identity__icontains=x)  for x in palabras)))[:20]
        results = []

        for x in usuarios:
            x_json = {}
            x_json["value"] =  x.id_students.first_name + " " + x.id_students.father_last_name + " " + x.id_students.mother_last_name+ " " + str(x.id_students.identity)
            if x_json in results:
                pass
            else:
                results.append(x_json)
        data = json.dumps(results)
    else:
        data = 'fail'
        
    mimetype = 'application/json'
    return HttpResponse(data, mimetype)  

    


def SearchApplicant(request):
    if request.method == "GET" and "q_aplicante" in request.GET:
        q_aplicante = request.GET["q_aplicante"]
        if q_aplicante:
            aplicantes = Applications.objects.filter(aplicante = q_aplicante)
            lista_aplicantes = []
            return render(request, "admisiones/BuscaAplicantes", {"aplicantes":q_aplicante} )
        else:
            return render(request, "admisiones/BuscaAplicantes",{})
    else:
        return render(request, "admisiones/BuscaAplicantes",{})
        
        
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def SearchStudent(request):
    if request.method == "GET" and "q_estudiante" in request.GET:
        q_estudiante = request.GET["q_estudiante"]
        
        if q_estudiante:
            estudiantes = Forms.objects.filter(estudiante = q_estudiante)
            lista_estudiantes = []
           
            return render(request, "admisiones/BuscaEstudiantes", {"estudiantes":q_estudiante} )
        
        else:
            return render(request, "admisiones/BuscaEstudiantes",{})


    else:
        return render(request, "admisiones/BuscaEstudiantes",{})


@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)
def ReportListTour(request): 
                    
    
        
    consultafechastour=Type_Tours.objects.filter(date_tour__gte = date.today(),state=True).order_by('date_tour')
    fechas=[]
    
    for y in consultafechastour:
        if y.date_tour in fechas:
            pass
        else:
            fechas.append(y.date_tour.strftime("%Y-%m-%d"))
            
            
            
    if request.method == "GET" and "q_fecha" in request.GET:
        fecha=request.GET["q_fecha"]
        if fecha:
             return redirect('ListaTourxlsx', fecha)   
        else:
            return render(request,'admisiones_sam/ReportListTours.html',{"fechas":fechas})         
    else:
        return render(request,'admisiones_sam/ReportListTours.html',{"fechas":fechas})       
                
            
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)        
def ReportListTour_xlsx(request,fecha):
    context = RequestContext(request)
    
    
    query=Applications.objects.filter(tours__id_typetour__date_tour=fecha,tours__id_typetour__state=True,tours__estado=True)
    
    lista=[]
    for x in query:
    
        grado=Type_Grade.objects.get(typegrade=x.applied_grade)
        
        campos=[x.tours_set.get(estado=True).id_typetour.date_tour,
                x.surname_student+' '+x.name_student,
                grado.typegrade+' / '+grado.typeschool, 
                x.surname_tutor+' '+x.name_tutor,
                x.phone_tutor,
                x.cell_tutor,
                x.mail_tutor]
        lista.append(campos)
   
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Lista_Tours.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Aplicaciones"
    row_num = 0

    columns = [
            (u"FECHA", 30),
            (u"ESTUDIANTE", 70),
            (u"GRADO APLICA", 70),
            (u"REPRESENTANTE", 70),
            (u"TELÉFONO", 70),
            (u"CELULAR", 70),
            (u"CORREO", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        #c.style.font.bold = True
        
        # set column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num+1)].width = columns[col_num][1]

    for x in lista:
    
        row_num += 1
        row=[]
        
        for item in x:
            row.append(item)
        
        
        for col_num in xrange(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                #c.style.alignment.wrap_text = True

    wb.save(response)
    return response


@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def ReportListTest(request): 
                   
    consultafechastest=Type_Tests.objects.filter(date_test__gte = date.today(),state=True).order_by('date_test')
    
    fechas=[]
    
    #query=Applications.objects.filter(test__id_typetest__date_test='2017-03-31',test__id_typetest__state=True)
    
    for y in consultafechastest:
        if y.date_test.strftime("%Y-%m-%d") in fechas:
            pass
        else:
            fechas.append(y.date_test.strftime("%Y-%m-%d"))
            
            
            
    if request.method == "GET" and "q_fecha" in request.GET:
        fecha=request.GET["q_fecha"]
        if fecha:
             return redirect('ListaEvaluacionesxlsx', fecha)   
        else:
            return render(request,'admisiones_sam/ReportListTest.html',{"fechas":fechas})         
    else:
        return render(request,'admisiones_sam/ReportListTest.html',{"fechas":fechas})       
                      

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)    

                 
def ReportListTest_xlsx(request,fecha):
    context = RequestContext(request)
    
    query=Applications.objects.filter(test__id_typetest__date_test=fecha,test__id_typetest__state=True,test__state=True)
    
    lista=[]
    for x in query:
        grado=Type_Grade.objects.get(typegrade=x.applied_grade)
        
        if str(x.test_set.get(state=True).id_typetest.date_test)==fecha:
            
            campos=[x.test_set.get(state=True).id_typetest.date_test,
                    grado.typegrade+' / '+grado.typeschool,
                    x.surname_student+' '+x.name_student,
                    x.surname_tutor+' '+x.name_tutor,
                    x.phone_tutor,
                    x.cell_tutor,
                    x.mail_tutor,
                    x.test_set.get(state=True).id_typetest.type_test,
                    x.test_set.get(state=True).id_typetest.time_test,
                    x.test_set.get(state=True).id_typetest.teacher]
                    
            #campos=[x.name_student,x.applied_grade,x.id_applications,x.surname_student,x.gender_student,x.name_tutor]
            lista.append(campos)
        else:
            pass
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Lista_Evaluaciones.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Aplicaciones"
    row_num = 0

    columns = [
            (u"FECHA", 30),
            (u"GRADO", 70),
            (u"ESTUDIANTE", 70),
            (u"REPRESENTANTE", 70),
            (u"TELÉFONO", 70),
            (u"CELULAR", 70),
            (u"CORREO", 70),
            (u"TIPO DE EVALUACIÓN", 70),
            (u"HORA", 30),
            (u"EVALUADOR", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        #c.style.font.bold = True
        
        # set column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num+1)].width = columns[col_num][1]

    for x in lista:
    
        row_num += 1
        row=[]
        
        for item in x:
            row.append(item)
        
        
        for col_num in xrange(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                #c.style.alignment.wrap_text = True

    wb.save(response)
    return response

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)


def SummaryApplications(request): 
    numero_aplicaciones=Applications.objects.filter(date_application__lte = '2017-03-20').values('applied_grade').annotate(total=Count('applied_grade')).order_by('total')

    if request.method == "GET" and "q_fecha" in request.GET:
        fecha=request.GET["q_fecha"]
        if fecha:
             return redirect('ResumenAplicacionesxlsx', fecha)   
        else:
            return render(request,'admisiones_sam/SummaryApplications.html')         
    else:
        return render(request,'admisiones_sam/SummaryApplications.html',{"numero_aplicaciones":numero_aplicaciones})       
                      

@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)  

#def
                   
def SummaryApplications_xlsx(request,fecha):
    context = RequestContext(request)
    
    numero_aplicaciones=Applications.objects.filter(date_application__lte = fecha).values('applied_grade').annotate(total=Count('applied_grade')).order_by('applied_grade')
   
    lista=[]
    
    for x in numero_aplicaciones:
        total=x["total"]
        
       # revisar tours 
      
       #def
        numero_tours=Applications.objects.filter(date_application__lte = fecha,
                                                    tours__estado=1,
                                                    tours__id_typetour__date_tour__lte=fecha,
                                                    applied_grade = x["applied_grade"]).distinct().count()
        
  
        try:
       
            numero_evaluaciones=Applications.objects.filter(date_application__lte = fecha,
                                                    #state_applications__state='Evaluaciones',
                                                    test__id_typetest__date_test__lte=fecha,
                                                    applied_grade = x["applied_grade"]).distinct().count()
        except:
            numero_evaluaciones=0
            
        try:
            evaluaciones_aprobadas=Applications.objects.filter(date_application__lte = fecha,
                                                            reports_tests__state__in=['Aprobado','Aprobado/E.F.'],
                                                            reports_tests__date_report__lte=fecha,
                                                            applied_grade = x["applied_grade"]).distinct().count()
                                                            #.values('applied_grade').annotate(total=Count('applied_grade')).group_by('applied_grade').order_by('applied_grade')
        except:
            evaluaciones_aprobadas=0
           
        
        try:    
            cupo_reservado=Applications.objects.filter(date_application__lte = fecha,
                                                    accountingdepartment__state='True',
                                                    applied_grade = x["applied_grade"]).distinct().count()
        except:
            cupo_reservado=0
            
            
        try:    
            cupos=Quotas.objects.get(grade=x["applied_grade"])
            cupos_total=cupos.max_students
            cupos_otorgados=cupos.old_students
            cupos_disponibles=cupos.available_students
        except:
            cupos_total=0
            cupos_otorgados=0
            cupos_disponibles=0
        
        grado=Type_Grade.objects.get(typegrade=x["applied_grade"])
       
        
        campos=[grado.typegrade+' / '+grado.typeschool,
                x["total"],
                numero_tours,
                #round((numero_tours/1)*100,2),
                round(float(numero_tours)/float(total)*100,2),
                numero_evaluaciones,
                round(float(numero_evaluaciones)/float(numero_tours)*100,2),
                evaluaciones_aprobadas,
               # round((evaluaciones_aprobadas/numero_evaluaciones)*100,2),
                cupo_reservado,
                cupos_total,
                cupos_otorgados,
                cupos_disponibles]
                
        #campos=[x.name_student,x.applied_grade,x.id_applications,x.surname_student,x.gender_student,x.name_tutor]
        lista.append(campos)
        
        
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Consolidado_Admisiones.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = "Consolidado"
    row_num = 0

    columns = [
            (u"GRADO", 40),
            (u"NUMERO DE APLICACIONES", 25),
            (u"ASISTENCIA A TOUR", 25),
            (u"% ASISTENCIA TOUR",25),
            (u"NUMERO DE EVALUADOS", 25),
            (u"% EVALUADOS", 25),
            (u"EVALUACIONES APROBADAS", 25),
            #(u"% APROBADOS",25),
            (u"RESERVARON CUPO", 20),
            (u"CUPOS TOTALES", 20),
            (u"ESTUDIANTES ANTIGUOS", 20),
            (u"CUPOS DISPONIBLES", 20),
            
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        #c.style.font.bold = True
        
        # set column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num+1)].width = columns[col_num][1]

    for x in lista:
    
        row_num += 1
        row=[]
        
        for item in x:
            row.append(item)
        
        
        for col_num in xrange(len(row)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = row[col_num]
                #c.style.alignment.wrap_text = True

    wb.save(response)
    return response

    
'''borrar'''
@login_required(login_url = '/admin/login/')
@user_passes_test(is_admisiones)

def prueba4 (request):
    fechas_disponibles=Type_Tours.objects.filter(state=True,date_tour__lt=date.today()).aggregate(Max('date_tour'))
    numero_fechas=Type_Tours.objects.filter(state=True,date_tour__gt=date.today()).order_by('date_tour').count()
    #return render(request,'admisiones_sam/prueba4.html')
    return render(request,'admisiones_sam/prueba4.html',{"fechas_disponibles":fechas_disponibles})       
    
    
    
